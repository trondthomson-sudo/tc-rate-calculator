"""
TC-rate calculator for a live fish carrier (brønnbåt) — Streamlit app.

Tabs:
    1. Vessel TC-rate    — required EBITDA (capex x yield) + vessel opex
    2. Lease spread       — leased equipment funded via bank debt, leased
                             out to the customer at a higher required return
    3. Combined TC-rate   — vessel TC-rate + the customer-facing lease
                             payment, on daily / monthly / annual basis

Run locally with:
    pip install streamlit pandas
    streamlit run tc_rate_app.py
"""

import datetime
import json
import os
import re
from io import BytesIO

import altair as alt
import pandas as pd
import streamlit as st


@st.cache_data
def _month_to_weeks_map():
    """{month_index (0=Jan..11=Dec): [ISO week numbers]} — derived once
    from a fixed reference year (2025, an arbitrary ordinary year; no
    real dates are used anywhere, only the week-number pattern), so
    month <-> week-number translation is stable and every week belongs
    to exactly one month (the month containing that ISO week's Thursday
    — the standard ISO 8601 rule for which month 'owns' a week that
    spans a month boundary). @st.cache_data means this 365-iteration
    date loop runs exactly once per app deployment rather than being
    rebuilt from scratch on every call — this gets called deep inside
    the monthly revenue/cost engine (once per month, per spot service,
    and that whole model runs twice per script pass), so an uncached
    version here was a real, measurable source of slow reruns."""
    year = 2025
    month_weeks = {m: [] for m in range(12)}
    seen_weeks = set()
    d = datetime.date(year, 1, 1)
    while d.year == year:
        iso_week = d.isocalendar()[1]
        if iso_week not in seen_weeks:
            iso_weekday = d.isocalendar()[2]  # 1=Mon..7=Sun
            thursday = d + datetime.timedelta(days=(4 - iso_weekday))
            if thursday.year == year:
                month_weeks[thursday.month - 1].append(iso_week)
                seen_weeks.add(iso_week)
        d += datetime.timedelta(days=1)
    return month_weeks

# ---------------------------------------------------------------------------
# Page setup
# ---------------------------------------------------------------------------

st.set_page_config(
    page_title="TC-rate calculator — Live Fish Carrier",
    page_icon="⚓",
    layout="wide",
)

# ---------------------------------------------------------------------------
# Auto-load a saved configuration, if one has been committed to the repo.
# This is what makes "your well-qualified inputs" the default everyone sees —
# see the "Configuration" section in the sidebar to save/load one.
# ---------------------------------------------------------------------------

CONFIG_FILE = "default_config.json"
CONFIG_EXCLUDE_KEYS = {"unlocked", "unlock_password_input", "_config_loaded", "_config_status", "config_uploader"}


def _is_excluded_key(key: str) -> bool:
    """Buttons (like the opex/service/price/voyage-cost '✕' remove
    buttons, and '+ Add phase'-style add buttons) can't have their
    session_state pre-set — Streamlit raises a
    StreamlitValueAssignmentNotAllowedError if you try, since buttons are
    trigger-only widgets. File uploader keys shouldn't be restored either.
    Match on "_remove_" and "_add_" anywhere in the key (not just as a
    prefix) so this catches every button naming pattern used across the
    app (remove_{i}, service_remove_{i}, smolt_add_phase,
    harvest_add_phase, and any future ones), rather than needing a new
    prefix added here every time a new button is built. Note: "_add_"
    specifically (underscores on both sides) — NOT a bare "add" substring
    match, which would incorrectly exclude legitimate restorable inputs
    like "spot_additional_capex_depreciation" (contains "add" as part of
    "additional", not as its own "_add_" token)."""
    if key in CONFIG_EXCLUDE_KEYS:
        return True
    if key.startswith("remove_") or "_remove_" in key:
        return True
    if key.startswith("add_") or "_add_" in key:
        return True
    return False


def _apply_config(config_dict):
    for k, v in config_dict.items():
        if _is_excluded_key(k):
            continue
        try:
            st.session_state[k] = v
        except Exception:
            pass  # skip any single key Streamlit won't allow, rather than crash the whole app


# Defensive cleanup, run every single pass (cheap — just a key filter):
# strip any session_state entries for button-only keys (remove_*,
# *_remove_*, add_*, *_add_*) that may have been set incorrectly by an
# older app version's config file, or any other means. Buttons are
# trigger-only widgets and can never legitimately hold a stored value;
# setting one raises StreamlitValueAssignmentNotAllowedError — but only
# when the button widget itself is created, not at assignment time, so
# _apply_config's own try/except above can't catch it. This runs
# unconditionally so any already-poisoned session (from before
# _is_excluded_key was widened to also exclude these keys) gets cleaned
# up immediately, without needing a brand-new session.
for _k in [
    k for k in list(st.session_state.keys())
    if k.startswith("remove_") or "_remove_" in k or k.startswith("add_") or "_add_" in k
]:
    del st.session_state[_k]

_config_status = None
if "_config_loaded" not in st.session_state:
    st.session_state["_config_loaded"] = True
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE, "r") as f:
                _apply_config(json.load(f))
            _config_status = f"✅ Loaded {CONFIG_FILE} ({os.path.getsize(CONFIG_FILE)} bytes) from {os.path.abspath(CONFIG_FILE)}"
        except Exception as e:
            _config_status = f"❌ Found {CONFIG_FILE} but failed to load it: {e}"
    else:
        _config_status = f"⚠️ {CONFIG_FILE} not found at {os.path.abspath(CONFIG_FILE)} (cwd: {os.getcwd()})"
    st.session_state["_config_status"] = _config_status

st.title("⚓ TC-rate calculator — live fish carrier")

st.caption(
    "Vessel TC-rate, leased equipment financing, and the combined total — "
    "all on a daily / monthly / annual basis."
)

currency = st.text_input("Currency", value="NOK", key="currency_input").strip() or "NOK"

# ---------------------------------------------------------------------------
# Sidebar controls (refresh, reset, save/load configuration)
# ---------------------------------------------------------------------------

def _request_rerun():
    """Defer a rerun to the very end of the script, after every tab's
    widgets have had a chance to render normally in this pass. Calling
    st.rerun() directly mid-script cuts the current pass short before
    later tabs' widgets ever run — Streamlit can then treat those
    never-reached widgets as 'not seen this run' and wipe their saved
    values (the exact mechanism stateful_number_input's shadow-key
    protection works around for individual widgets). Deferring the
    actual rerun avoids the whole class of bug at its root, for every
    widget, without needing per-widget protection."""
    st.session_state["_pending_rerun"] = True


with st.sidebar:
    if st.session_state.get("spot_market_enabled", False):
        st.warning(
            "⚠️ **Spot market is ON.** The Financial Statements, Sources & "
            "Uses, and every downstream tab are running on spot-market "
            "revenue (Spot market tab), not the TC-rate. Turn it off there "
            "if that's not intended."
        )

with st.sidebar:
    st.subheader("🔄 Refresh")
    if st.button("Refresh calculations"):
        _request_rerun()
    st.caption(
        "A few figures (e.g. Tab 1's Sources & Uses guideline) are computed "
        "one script pass behind live edits elsewhere. If a number looks "
        "stale after changing inputs, click here instead of switching tabs."
    )
    if st.button("⚠️ Reset to script defaults (clears everything)"):
        for _k in list(st.session_state.keys()):
            del st.session_state[_k]
        _request_rerun()
    st.caption(
        "Wipes all inputs — including anything typed in this session — "
        "and rebuilds the app entirely from the script's own hardcoded "
        "defaults, bypassing any leftover or corrupted session state. Use "
        "this if numbers look wrong in a way Refresh doesn't fix (e.g. "
        "fields showing their bare minimum value instead of the intended "
        "default)."
    )

# All inputs are editable — no password gate. Kept as a single flag
# (rather than removing it) since it's threaded through every widget's
# disabled= argument across the whole app; always False now.
locked = False

with st.sidebar:
    st.divider()
    st.subheader("💾 Configuration")
    st.caption(
        "Save your current inputs as the default everyone sees."
    )

    config_to_save = {k: v for k, v in dict(st.session_state).items() if not _is_excluded_key(k)}
    config_json = json.dumps(config_to_save, default=str, indent=2)
    st.download_button(
        "Save current inputs as default",
        data=config_json,
        file_name=CONFIG_FILE,
        mime="application/json",
    )
    st.caption(
        f"Downloads **{CONFIG_FILE}** — commit it to the same GitHub repo "
        f"as `tc_rate_app.py` (same folder) and the app will auto-load it "
        f"for every visitor from then on."
    )

    uploaded_config = st.file_uploader("Or load a saved configuration", type="json", key="config_uploader")
    if uploaded_config is not None:
        try:
            loaded = json.load(uploaded_config)
            _apply_config(loaded)
            st.success("Configuration loaded.")
            _request_rerun()
        except Exception as e:
            st.error(f"Couldn't load that file: {e}")

# ---------------------------------------------------------------------------
# Shared helpers
# ---------------------------------------------------------------------------


def format_nok(n: float) -> str:
    """1234567 -> '1 234 567'"""
    return f"{n:,.0f}".replace(",", " ")


def parse_nok(s: str) -> float:
    """'1 234 567' or '1234567' -> 1234567.0"""
    cleaned = re.sub(r"[^\d.\-]", "", s or "")
    try:
        return float(cleaned) if cleaned else 0.0
    except ValueError:
        return 0.0


def nok_input(label: str, state_key: str, default: float, key: str, disabled: bool = False) -> float:
    """A text input that displays with thousand separators and re-formats
    itself every time the value changes (on Enter / click-away), rather
    than freezing on whatever was last typed."""
    if state_key not in st.session_state:
        st.session_state[state_key] = default
    if key not in st.session_state:
        st.session_state[key] = format_nok(st.session_state[state_key])

    def _on_change():
        value = parse_nok(st.session_state[key])
        st.session_state[state_key] = value
        st.session_state[key] = format_nok(value)

    st.text_input(label, key=key, on_change=_on_change, disabled=disabled)
    return st.session_state[state_key]


def stateful_number_input(label, key=None, value=0.0, **kwargs):
    """Drop-in replacement for st.number_input, protected against Streamlit
    wiping the widget's saved value when it briefly isn't rendered (e.g. the
    lock/unlock buttons call st.rerun() from partway through the script,
    before Tab 1's widgets run — Streamlit treats any widget key that wasn't
    freshly re-registered as 'gone' and clears it). A shadow key that's never
    itself a widget's key= holds the true value and survives that cleanup,
    the same protection nok_input already has via its separate state_key."""
    shadow_key = f"__shadow_{key}"
    if shadow_key not in st.session_state:
        st.session_state[shadow_key] = value
    if key not in st.session_state:
        st.session_state[key] = st.session_state[shadow_key]
    result = st.number_input(label, key=key, **kwargs)
    st.session_state[shadow_key] = result
    return result


def stateful_toggle(label, key=None, value=False, **kwargs):
    """Same shadow-key protection as stateful_number_input, for st.toggle."""
    shadow_key = f"__shadow_{key}"
    if shadow_key not in st.session_state:
        st.session_state[shadow_key] = value
    if key not in st.session_state:
        st.session_state[key] = st.session_state[shadow_key]
    result = st.toggle(label, key=key, **kwargs)
    st.session_state[shadow_key] = result
    return result


def fmt(n):
    return format_nok(n)  # currency shown once, via the field at the top — not repeated per number


def show_table(df: pd.DataFrame, label_col: str = None, decimal_cols: list = None, **kwargs):
    """Display a numeric dataframe with right-aligned, thousands-separated
    columns (Streamlit only right-aligns numeric dtype columns — pre-formatted
    strings stay left-aligned regardless of styling, hence keeping values raw
    here and letting column_config handle the display format).
    If label_col is given, that column becomes the index (kept as text).
    decimal_cols: optional list of column names to show with one decimal
    place instead of the default integer formatting (e.g. a "Months"
    column where 7.5 matters and rounding to 7 or 8 would be misleading)."""
    display_df = df.set_index(label_col) if label_col else df
    decimal_cols = decimal_cols or []
    config = {
        col: st.column_config.NumberColumn(format="%.1f" if col in decimal_cols else "%,d")
        for col in display_df.columns
    }
    st.dataframe(display_df, column_config=config, **kwargs)


NOK_AXIS_FORMAT = ",.0f"  # Vega-Lite doesn't support space thousands; comma is the closest built-in


def formatted_line_chart(df: pd.DataFrame, x_col: str, y_cols: list, height: int = 300):
    """Line chart with comma-separated axis labels and tooltips (Streamlit's
    built-in st.line_chart shows raw unformatted numbers on hover)."""
    melted = df.reset_index().melt(id_vars=[x_col], value_vars=y_cols, var_name="Series", value_name="Value")
    chart = (
        alt.Chart(melted)
        .mark_line()
        .encode(
            x=alt.X(f"{x_col}:Q", title=x_col),
            y=alt.Y("Value:Q", title=None, axis=alt.Axis(format=NOK_AXIS_FORMAT)),
            color=alt.Color("Series:N", title=None),
            tooltip=[
                alt.Tooltip(f"{x_col}:Q", title=x_col),
                alt.Tooltip("Series:N", title="Series"),
                alt.Tooltip("Value:Q", title="Value", format=NOK_AXIS_FORMAT),
            ],
        )
        .properties(height=height)
        .interactive()
    )
    st.altair_chart(chart, width="stretch")


def formatted_bar_chart(df: pd.DataFrame, category_col: str, value_col: str, height: int = 300):
    """Horizontal bar chart with comma-separated axis labels, tooltips, and
    a value label printed on each bar (space-thousands, matching the rest
    of the app's number formatting)."""
    plot_df = df.reset_index()
    plot_df["_label"] = plot_df[value_col].apply(format_nok)

    base = alt.Chart(plot_df).encode(
        y=alt.Y(f"{category_col}:N", title=None, sort=None),
        x=alt.X(f"{value_col}:Q", title=None, axis=alt.Axis(format=NOK_AXIS_FORMAT)),
    )
    bars = base.mark_bar().encode(
        tooltip=[
            alt.Tooltip(f"{category_col}:N", title=category_col),
            alt.Tooltip(f"{value_col}:Q", title=value_col, format=NOK_AXIS_FORMAT),
        ],
    )
    labels = base.mark_text(align="left", dx=5, color="black").encode(text="_label:N")
    chart = (bars + labels).properties(height=height)
    st.altair_chart(chart, width="stretch")


def annuity_monthly_payment(principal_nok: float, annual_rate_pct: float, num_months: int) -> float:
    """Level monthly annuity payment. Nominal monthly rate = annual rate / 12."""
    if num_months <= 0:
        return 0.0
    r = (annual_rate_pct / 100) / 12
    if r == 0:
        return principal_nok / num_months
    return principal_nok * r / (1 - (1 + r) ** -num_months)


def amortization_balances(principal_nok: float, annual_rate_pct: float, num_months: int, payment_basis_months: int = None) -> list:
    """Closing balance for each month (used to chart the paydown).

    If payment_basis_months is given and differs from num_months, the level
    payment is computed over the LONGER payback basis, but the schedule
    only runs for num_months — leaving a genuine, non-zero residual
    balance at the end (e.g. a lease rental priced on a 7-year payback but
    only actually collected over a 5-year customer contract)."""
    basis_months = payment_basis_months if payment_basis_months else num_months
    r = (annual_rate_pct / 100) / 12
    payment = annuity_monthly_payment(principal_nok, annual_rate_pct, basis_months)
    balance = principal_nok
    balances = []
    for month in range(1, num_months + 1):
        interest = balance * r
        principal_paid = payment - interest
        balance = balance - principal_paid
        if month == num_months and basis_months == num_months:
            balance = 0.0  # only force to zero when the schedule fully amortizes within num_months
        balances.append(balance)
    return balances


def amortization_schedule_full(principal_nok: float, annual_rate_pct: float, num_months: int, payment_basis_months: int = None) -> list:
    """Month-by-month: opening balance, payment, finance cost (interest),
    amortization (principal), closing balance.

    Same payment_basis_months mechanic as amortization_balances() above —
    see that docstring."""
    basis_months = payment_basis_months if payment_basis_months else num_months
    r = (annual_rate_pct / 100) / 12
    payment = annuity_monthly_payment(principal_nok, annual_rate_pct, basis_months)
    balance = principal_nok
    rows = []
    for month in range(1, num_months + 1):
        interest = balance * r
        principal_paid = payment - interest
        closing = balance - principal_paid
        if month == num_months and basis_months == num_months:
            closing = 0.0  # only force to zero when the schedule fully amortizes within num_months
        rows.append({
            "Month": month,
            "Opening balance": balance,
            "Finance cost": interest,
            "Amortization": principal_paid,
            "Payment": payment,
            "Closing balance": closing,
        })
        balance = closing
    return rows


# ---------------------------------------------------------------------------
# Session state — opex line items (persist across reruns)
# ---------------------------------------------------------------------------

if "opex_items" not in st.session_state:
    st.session_state.opex_items = [
        {"name": "Crewing", "value_nok": 19_737_162.08},
        {"name": "Insurance", "value_nok": 2_548_000.0},
        {"name": "Other vessel opex", "value_nok": 3_250_000.0},
    ]


def add_opex_item():
    st.session_state.opex_items.append({"name": "New item", "value_nok": 0.0})


def remove_opex_item(index):
    st.session_state.opex_items.pop(index)


def _on_opex_value_change(index):
    raw = st.session_state[f"value_{index}"]
    value = parse_nok(raw)
    st.session_state.opex_items[index]["value_nok"] = value
    st.session_state[f"value_{index}"] = format_nok(value)


# ---------------------------------------------------------------------------
# Tabs
# ---------------------------------------------------------------------------

tab_vessel, tab_opex, tab_construction, tab_spot, tab_lease, tab_combined, tab_financials, tab_investment, tab_summary, tab_inputs = st.tabs(
    ["Vessel TC-rate", "Vessel opex", "Construction capex", "Spot market", "Lease spread", "Combined TC-rate", "Financial Statements", "Investment Analysis", "Summary", "Inputs"]
)

# ===========================================================================
# TAB 1 — Vessel TC-rate
# ===========================================================================
with tab_vessel:
    left, right = st.columns([1, 1.4], gap="large")

    with left:
        st.subheader("Capital & return")
        capex_nok = nok_input("Capex (NOK)", "capex_nok", 800_000_000.0, key="capex_input", disabled=locked)
        ebitda_yield_pct = stateful_number_input(
            "EBITDA-yield (%)", min_value=0.0, value=12.0, step=0.1, key="ebitda_yield", disabled=locked
        )
        operating_days = stateful_number_input(
            "Operating days / year", min_value=1, value=365, step=1, key="operating_days", disabled=locked
        )

        st.subheader("Vessel opex (annual, NOK)")
        opex_linked_to_vessel_opex_tab = stateful_toggle(
            "Link Crewing / Insurance / Other vessel opex to the Vessel opex tab?",
            value=True, key="opex_linked_to_vessel_opex_tab", disabled=locked
        )
        st.caption(
            "When on, these three line items are driven automatically by "
            "the detailed build-up on the Vessel opex tab — their value "
            "fields below are disabled, since editing happens there "
            "instead. Turn this off to set them manually here again "
            "(e.g. for a quick standalone estimate without touching the "
            "detailed build-up)."
        )
        _linked_opex_names = {"Crewing", "Insurance", "Other vessel opex"}
        for i, item in enumerate(st.session_state.opex_items):
            _this_item_is_linked = opex_linked_to_vessel_opex_tab and item["name"] in _linked_opex_names
            c1, c2, c3 = st.columns([2.2, 1.6, 0.4])
            with c1:
                item["name"] = st.text_input(
                    "Name", value=item["name"], key=f"name_{i}", label_visibility="collapsed",
                    disabled=locked
                )
            with c2:
                if f"value_{i}" not in st.session_state:
                    st.session_state[f"value_{i}"] = format_nok(item["value_nok"])
                st.text_input(
                    "Value (NOK)",
                    key=f"value_{i}",
                    label_visibility="collapsed",
                    on_change=_on_opex_value_change,
                    args=(i,),
                    disabled=locked or _this_item_is_linked,
                )
            with c3:
                st.button("✕", key=f"remove_{i}", on_click=remove_opex_item, args=(i,), disabled=locked)

        st.button("+ Add opex line item", on_click=add_opex_item, disabled=locked)

        opex_total = sum(item["value_nok"] for item in st.session_state.opex_items)
        st.markdown(f"**Total vessel opex:** {format_nok(opex_total)} NOK")

        st.subheader("Depreciation & maintenance")
        depreciation_rate_pct = stateful_number_input(
            "Vessel depreciation rate, annual (%)", min_value=0.0, value=2.5, step=0.1,
            key="depreciation_rate", disabled=locked
        )
        st.caption(
            "Straight-line, % of original vessel capex per year. Leased "
            "equipment (Tab 2) is depreciated separately, straight-line over "
            "the same number of months as its financing term — see the "
            "Lease spread tab."
        )
        annual_maintenance_capex_nok = nok_input(
            "Annual maintenance capex (NOK)", "maintenance_capex_nok", 4_300_000.0,
            key="maintenance_capex_input", disabled=locked
        )

    # --- calculations ---
    required_ebitda_annual = capex_nok * (ebitda_yield_pct / 100)
    vessel_tc_annual = required_ebitda_annual + opex_total
    vessel_tc_daily = vessel_tc_annual / operating_days if operating_days else 0
    vessel_tc_monthly = vessel_tc_annual / 12

    with right:
        header_col, req_col = st.columns([2, 1.3])
        with header_col:
            st.subheader("Rate build-up")
        with req_col:
            st.markdown(
                f"""
                <div style="text-align:right; padding-top:6px;">
                    <span style="font-size:12px; color:#888; text-transform:uppercase; letter-spacing:0.05em;">
                        TC-requirement, annual
                    </span><br>
                    <span style="font-size:22px; font-weight:700; color:#1a1a1a;">
                        {fmt(vessel_tc_annual)}
                    </span>
                </div>
                """,
                unsafe_allow_html=True,
            )

        chart_rows = [{"Component": "Required EBITDA", "NOK (annual)": required_ebitda_annual}]
        for item in st.session_state.opex_items:
            if item["value_nok"] > 0:
                chart_rows.append({"Component": item["name"], "NOK (annual)": item["value_nok"]})
        chart_df = pd.DataFrame(chart_rows)
        formatted_bar_chart(chart_df, "Component", "NOK (annual)")

        st.subheader("TC-rate")
        results_df = pd.DataFrame(
            [
                {
                    "Component": "Required EBITDA",
                    "Daily": required_ebitda_annual / operating_days,
                    "Monthly": required_ebitda_annual / 12,
                    "Annual": required_ebitda_annual,
                },
                {
                    "Component": "Vessel opex",
                    "Daily": opex_total / operating_days,
                    "Monthly": opex_total / 12,
                    "Annual": opex_total,
                },
                {
                    "Component": "TC-rate",
                    "Daily": vessel_tc_daily,
                    "Monthly": vessel_tc_monthly,
                    "Annual": vessel_tc_annual,
                },
            ]
        )
        show_table(results_df, "Component", width="stretch")

        m1, m2, m3 = st.columns(3)
        m1.metric("TC-rate, daily", fmt(vessel_tc_daily))
        m2.metric("TC-rate, monthly", fmt(vessel_tc_monthly))
        m3.metric("TC-rate, annual", fmt(vessel_tc_annual))

        st.caption(
            "**Scope:** this TC-rate covers capital return and vessel opex only. "
            "Fuel, lubrication oil, port fees, and other spot expenses are excluded "
            "and typically sit for charterer's account."
        )

    # =======================================================================
    # Debt financing (vessel)
    # =======================================================================
    st.divider()
    st.subheader("Debt financing")
    st.caption(
        "Debt sized off Year 1 EBITDA, at an implied LTV against capex. "
        "Finance cost = swap rate + credit spread, charged monthly on the "
        "outstanding balance; principal is repaid quarterly (straight-line), "
        "not monthly — the typical structure for shipping debt."
    )

    debt_left, debt_right = st.columns([1, 1.4], gap="large")

    with debt_left:
        debt_multiple = stateful_number_input(
            "Debt multiple (x Year 1 EBITDA)", min_value=0.0, value=6.0, step=0.5,
            key="debt_multiple", disabled=locked
        )
        amortization_years = stateful_number_input(
            "Amortization profile (years)", min_value=1, max_value=30, value=12, step=1,
            key="amortization_years", disabled=locked
        )
        swap_rate_pct = stateful_number_input(
            "Swap rate, annual (%)", min_value=0.0, value=4.0, step=0.1, key="swap_rate", disabled=locked
        )
        credit_spread_pct = stateful_number_input(
            "Credit spread, annual (%)", min_value=0.0, value=3.5, step=0.1, key="credit_spread", disabled=locked
        )

        debt_nok = debt_multiple * required_ebitda_annual
        implied_ltv_pct = (debt_nok / capex_nok * 100) if capex_nok else 0.0
        implied_equity_nok = capex_nok - debt_nok
        finance_cost_rate_pct = swap_rate_pct + credit_spread_pct

        st.markdown(f"**Debt:** {format_nok(debt_nok)} NOK")
        st.markdown(f"**Implied LTV:** {implied_ltv_pct:,.1f}%".replace(",", " "))
        st.markdown(f"**Implied equity (capex − debt):** {format_nok(implied_equity_nok)} NOK")
        st.caption(
            "This covers the vessel purchase only. For an additional "
            "operational cash buffer — sized off the worst point in the "
            "monthly cash flow — see the Financial Statements tab."
        )

        st.markdown("**Finance cost summary**")
        finance_cost_summary_df = pd.DataFrame(
            [
                {"Component": "Swap rate", "Rate (%)": swap_rate_pct},
                {"Component": "Credit spread", "Rate (%)": credit_spread_pct},
                {"Component": "Total finance cost", "Rate (%)": finance_cost_rate_pct},
            ]
        )
        st.dataframe(
            finance_cost_summary_df.set_index("Component"),
            column_config={"Rate (%)": st.column_config.NumberColumn(format="%.2f")},
            width="stretch",
        )

    # --- build monthly schedule: interest monthly, principal quarterly ---
    amortization_months = int(amortization_years) * 12
    quarterly_amortization_nok = debt_nok / (int(amortization_years) * 4) if amortization_years else 0.0
    monthly_rate = (finance_cost_rate_pct / 100) / 12

    debt_schedule = []
    balance = debt_nok
    for month in range(1, amortization_months + 1):
        opening_balance = balance
        monthly_finance_cost = opening_balance * monthly_rate
        is_quarter_end = (month % 3 == 0)
        principal_paid = quarterly_amortization_nok if is_quarter_end else 0.0
        # guard the very last month against floating point residue
        if month == amortization_months:
            principal_paid = opening_balance
        closing_balance = opening_balance - principal_paid

        debt_schedule.append({
            "Month": month,
            "Opening balance": opening_balance,
            "Monthly finance cost": monthly_finance_cost,
            "Quarterly amortization": principal_paid,
            "Closing balance": closing_balance,
        })
        balance = closing_balance

    with debt_right:
        first_year_finance_cost = sum(row["Monthly finance cost"] for row in debt_schedule[:12])
        first_month_finance_cost = debt_schedule[0]["Monthly finance cost"] if debt_schedule else 0.0

        m1, m2, m3 = st.columns(3)
        m1.metric("Monthly finance cost (month 1)", fmt(first_month_finance_cost))
        m2.metric("Quarterly amortization", fmt(quarterly_amortization_nok))
        m3.metric("Finance cost, Year 1", fmt(first_year_finance_cost))

        chart_df = pd.DataFrame(debt_schedule)[["Month", "Closing balance"]]
        formatted_line_chart(chart_df, "Month", ["Closing balance"])

        st.markdown("**Monthly schedule** (finance cost monthly, amortization quarterly)")
        debt_schedule_df = pd.DataFrame(debt_schedule)
        show_table(debt_schedule_df, "Month", width="stretch", height=300)

    st.divider()
    st.subheader("Sources & uses")
    su = st.session_state.get("_sources_uses")
    if su is None:
        st.caption(
            "Computing... this figure comes from the Financial Statements tab "
            "and will appear after the page finishes loading."
        )
        cover_operational_funding = False
        operational_equity_nok = 0.0
        operational_debt_nok = 0.0
    else:
        deficit_guideline = abs(su["min_cash_balance"]) if su["min_cash_balance"] < 0 else 0.0

        st.markdown("**Operational funding requirement**")
        st.caption(
            f"Guideline: {fmt(deficit_guideline)}, from the worst point in the monthly "
            f"cash flow assuming **zero** operational funding (month {su['min_cash_month']}). "
            f"This is a fixed target — it doesn't move depending on how much you "
            f"choose to fund below." if deficit_guideline > 0 else
            "Cash flow never dips negative — no operational funding is required."
        )
        cover_operational_funding = stateful_toggle(
            "Cover the operational funding requirement?", value=False,
            key="cover_operational_funding", disabled=locked
        )

        if cover_operational_funding and deficit_guideline > 0:
            operational_equity_raw = nok_input(
                "Operational funding — equity portion (NOK)", "operational_equity_nok", 0.0,
                key="operational_equity_input", disabled=locked
            )
            # IMPORTANT: never write a corrected/capped value back into this
            # widget's own state. The guideline above is computed on Tab 4
            # from a run forced to ZERO operational funding, so — unlike the
            # old design — it no longer moves depending on how much you fund
            # here; it only changes if some other input (capex, opex,
            # escalators, contracts, refinancing) was just edited on another
            # tab, in which case it's one script pass behind until you
            # switch tabs. Keep the raw typed amount untouched in the box
            # regardless, and only cap it for the calculations below,
            # recomputed fresh from the latest guideline every pass — so a
            # temporarily-stale guideline can never permanently discard
            # what the user typed.
            operational_equity_nok = min(operational_equity_raw, deficit_guideline)
            operational_debt_nok = deficit_guideline - operational_equity_nok
            if operational_equity_raw > deficit_guideline:
                st.caption(
                    f"⚠️ You entered {fmt(operational_equity_raw)}, which is more than the "
                    f"guideline ({fmt(deficit_guideline)}). Only {fmt(operational_equity_nok)} "
                    f"is being applied below — the remainder isn't needed unless another "
                    f"input on a different tab changes the guideline."
                )
            st.caption(
                f"The remaining {fmt(operational_debt_nok)} is automatically funded as "
                f"additional debt, at the same swap + credit spread rate and amortization "
                f"profile as the vessel debt above."
            )
        else:
            operational_equity_nok = 0.0
            operational_debt_nok = 0.0

        uncovered = 0.0 if cover_operational_funding else deficit_guideline

        uses_col, sources_col = st.columns(2)
        with uses_col:
            st.markdown("**Uses**")
            uses_df = pd.DataFrame([
                {"Item": "Vessel capex", "Amount": capex_nok},
                {"Item": "Equipment capex" + ("" if su["lease_enabled"] else " (off)"), "Amount": su["equipment_debt"] + su["equipment_equity"]},
                {"Item": "Operational funding — equity portion", "Amount": operational_equity_nok},
                {"Item": "Operational funding — debt portion", "Amount": operational_debt_nok},
                {"Item": "Uncovered operational fund requirement", "Amount": uncovered},
                {"Item": "Total uses", "Amount": capex_nok + su["equipment_debt"] + su["equipment_equity"] + operational_equity_nok + operational_debt_nok + uncovered},
            ])
            show_table(uses_df, "Item", width="stretch")
        with sources_col:
            st.markdown("**Sources**")
            sources_df = pd.DataFrame([
                {"Item": "Vessel debt", "Amount": su["vessel_debt"]},
                {"Item": "Equipment debt" + ("" if su["lease_enabled"] else " (off)"), "Amount": su["equipment_debt"]},
                {"Item": "Operational funding debt", "Amount": operational_debt_nok},
                {"Item": "Vessel equity", "Amount": su["vessel_equity"]},
                {"Item": "Equipment equity" + ("" if su["lease_enabled"] else " (off)"), "Amount": su["equipment_equity"]},
                {"Item": "Operational funding equity", "Amount": operational_equity_nok},
                {"Item": "Uncovered operational funding (equity or debt, TBD)", "Amount": uncovered},
                {"Item": "Total sources", "Amount": su["vessel_debt"] + su["equipment_debt"] + operational_debt_nok + su["vessel_equity"] + su["equipment_equity"] + operational_equity_nok + uncovered},
            ])
            show_table(sources_df, "Item", width="stretch")

        if uncovered > 0:
            st.warning(
                f"**{fmt(uncovered)} of the operational cash-deficit guideline is not "
                f"funded.** Turn on 'Cover the operational funding requirement?' above "
                f"to raise it as equity, debt, or a mix of both — otherwise it's left "
                f"to group liquidity to absorb."
            )

        # store for Tab 4 (cash flow / balance sheet) and Tab 5 (IRR) to pick up
        st.session_state["_operational_funding"] = {
            "equity": operational_equity_nok,
            "debt": operational_debt_nok,
        }

        if su["min_cash_balance"] < 0:
            st.caption(
                f"Operational equity buffer guideline: {fmt(abs(su['min_cash_balance']))}, "
                f"from the worst point in the monthly cash flow (month {su['min_cash_month']}) "
                f"— see the Financial Statements tab to adjust."
            )
        st.caption(
            "This section mirrors the Financial Statements tab. The guideline "
            "itself no longer depends on your funding choice above, so it "
            "won't shift as you adjust the equity/debt split. If a different "
            "input changes elsewhere (capex, opex, escalators, contracts, "
            "refinancing), this number auto-refreshes on its own — no need "
            "to switch tabs or click Refresh manually."
        )

# ===========================================================================
# TAB 1.5 — Vessel opex (detailed bottom-up build-up: crew wages, Norwegian
#           net wage scheme, insurance, other vessel cost — replicated from
#           the uploaded Opex workbook)
# ===========================================================================
with tab_opex:
    st.subheader("Vessel opex — detailed build-up")
    st.caption(
        "Bottom-up crew cost, insurance, and other vessel cost build-up, "
        "replicated from the uploaded Opex workbook — including the "
        "Norwegian net wage scheme for seafarers (nettolønnsordningen): "
        "the state refunds the employer's payroll tax on Norwegian "
        "seafarers' wages, up to a cap per crew member per year."
    )

    opexfolder_vessel_value = capex_nok
    st.metric("Vessel value (for insurance calc)", fmt(opexfolder_vessel_value))
    st.caption(
        "Linked directly to Tab 1's Capex — always the same number, "
        "edited there, not here."
    )

    st.markdown("**Wages** (by rank)")
    if "opexfolder_wages" not in st.session_state:
        st.session_state.opexfolder_wages = [
            {"rank": "Captain", "headcount": 2, "monthly_wage": 99_900.0, "tax_rate_pct": 24.0},
            {"rank": "Chief Officer", "headcount": 2, "monthly_wage": 89_500.0, "tax_rate_pct": 24.0},
            {"rank": "Chief Engineer", "headcount": 2, "monthly_wage": 83_000.0, "tax_rate_pct": 24.0},
            {"rank": "Able Seaman", "headcount": 4, "monthly_wage": 59_300.0, "tax_rate_pct": 22.0},
            {"rank": "Cook", "headcount": 2, "monthly_wage": 59_300.0, "tax_rate_pct": 22.0},
            {"rank": "Apprentice", "headcount": 2, "monthly_wage": 23_000.0, "tax_rate_pct": 22.0},
            {"rank": "Deck Cadet", "headcount": 0, "monthly_wage": 30_000.0, "tax_rate_pct": 22.0},
            {"rank": "Second Engineer", "headcount": 2, "monthly_wage": 69_300.0, "tax_rate_pct": 22.0},
            {"rank": "First Mate", "headcount": 2, "monthly_wage": 79_000.0, "tax_rate_pct": 24.0},
            {"rank": "Ordinary Seaman", "headcount": 2, "monthly_wage": 59_300.0, "tax_rate_pct": 22.0},
        ]

    def _add_wage_row():
        st.session_state.opexfolder_wages.append(
            {"rank": "New rank", "headcount": 0, "monthly_wage": 0.0, "tax_rate_pct": 22.0}
        )

    def _remove_wage_row(index):
        st.session_state.opexfolder_wages.pop(index)

    def _on_wage_monthly_change(index):
        raw = st.session_state[f"opexfolder_wage_monthly_{index}"]
        value = parse_nok(raw)
        st.session_state.opexfolder_wages[index]["monthly_wage"] = value
        st.session_state[f"opexfolder_wage_monthly_{index}"] = format_nok(value)

    whdr = st.columns([1.6, 0.8, 1.2, 1.0, 0.4])
    whdr[0].markdown("**Rank**")
    whdr[1].markdown("**#**")
    whdr[2].markdown("**Monthly wage (NOK)**")
    whdr[3].markdown("**Tax rate (%)**")

    for i, item in enumerate(st.session_state.opexfolder_wages):
        c1, c2, c3, c4, c5 = st.columns([1.6, 0.8, 1.2, 1.0, 0.4])
        with c1:
            item["rank"] = st.text_input(
                "Rank", value=item["rank"], key=f"opexfolder_wage_rank_{i}", label_visibility="collapsed", disabled=locked
            )
        with c2:
            item["headcount"] = st.number_input(
                "#", min_value=0, value=item["headcount"], step=1,
                key=f"opexfolder_wage_headcount_{i}", label_visibility="collapsed", disabled=locked
            )
        with c3:
            _monthly_wage_display_key = f"opexfolder_wage_monthly_{i}"
            if _monthly_wage_display_key not in st.session_state:
                st.session_state[_monthly_wage_display_key] = format_nok(item["monthly_wage"])
            st.text_input(
                "Monthly wage", key=_monthly_wage_display_key, label_visibility="collapsed",
                on_change=_on_wage_monthly_change, args=(i,), disabled=locked
            )
        with c4:
            item["tax_rate_pct"] = st.number_input(
                "Tax rate (%)", min_value=0.0, max_value=100.0, value=item["tax_rate_pct"], step=1.0,
                key=f"opexfolder_wage_taxrate_{i}", label_visibility="collapsed", disabled=locked
            )
        with c5:
            st.button("✕", key=f"opexfolder_wage_remove_{i}", on_click=_remove_wage_row, args=(i,), disabled=locked)

    st.button("+ Add rank", key="opexfolder_wage_add_rank", on_click=_add_wage_row, disabled=locked)

    st.markdown("**Rotation & net wage scheme inputs**")
    rc1, rc2, rc3, rc4 = st.columns(4)
    with rc1:
        opexfolder_days_on_board = stateful_number_input(
            "Days on board /yr /crew member", min_value=0.0, value=182.5, step=0.5,
            key="opexfolder_days_on_board", disabled=locked
        )
    with rc2:
        opexfolder_trips_per_year = stateful_number_input(
            "# of travels /yr /crew member", min_value=0.0, value=13.0, step=1.0,
            key="opexfolder_trips_per_year", disabled=locked
        )
    with rc3:
        opexfolder_max_grant_per_head = nok_input(
            "Max grant /head/yr (NOK)", "opexfolder_max_grant_per_head", 216_000.0,
            key="opexfolder_max_grant_input", disabled=locked
        )
    with rc4:
        opexfolder_1g = nok_input(
            "1G — Norwegian base amount (NOK)", "opexfolder_1g", 118_620.0,
            key="opexfolder_1g_input", disabled=locked
        )
    opexfolder_7_1g = opexfolder_1g * 7.1
    st.caption(
        f"7.1G threshold = 1G x 7.1 = {fmt(opexfolder_7_1g)}. Grant = MIN(max grant/head x "
        f"headcount, tax on that rank's total wages). Salary above 7.1G (floored at 0) "
        f"feeds the OTP (7.1G-12G) social cost line below."
    )

    # --- per-rank computation, mirroring the workbook's Calculations table ---
    _opexfolder_wage_calc = []
    for item in st.session_state.opexfolder_wages:
        _yearly_salary = item["monthly_wage"] * 12
        _yearly_wages_total = item["headcount"] * _yearly_salary
        _tax = _yearly_wages_total * (item["tax_rate_pct"] / 100)
        _max_grant = opexfolder_max_grant_per_head * item["headcount"]
        _grant = min(_max_grant, _tax)
        _salary_above_71g = (_yearly_salary - opexfolder_7_1g) * item["headcount"]
        _salary_above_71g_floored = max(0.0, _salary_above_71g)
        _opexfolder_wage_calc.append({
            "Rank": item["rank"], "#": item["headcount"], "Yearly wages (total)": _yearly_wages_total,
            "Tax": _tax, "Max grant": _max_grant, "Grant": _grant,
            "Salary above 7.1G (floored)": _salary_above_71g_floored,
        })

    opexfolder_total_headcount = sum(item["headcount"] for item in st.session_state.opexfolder_wages)
    opexfolder_total_yearly_wages = sum(r["Yearly wages (total)"] for r in _opexfolder_wage_calc)
    opexfolder_total_grant = sum(r["Grant"] for r in _opexfolder_wage_calc)
    opexfolder_total_salary_above_71g = sum(r["Salary above 7.1G (floored)"] for r in _opexfolder_wage_calc)

    st.markdown("**Grant scheme calculation** (by rank)")
    wage_calc_df = pd.DataFrame(_opexfolder_wage_calc)
    show_table(wage_calc_df, "Rank", width="stretch")
    wc1, wc2, wc3 = st.columns(3)
    wc1.metric("Total headcount", f"{opexfolder_total_headcount:.0f}")
    wc2.metric("Total yearly wages", fmt(opexfolder_total_yearly_wages))
    wc3.metric("Total grant (refund)", fmt(opexfolder_total_grant))

    with st.expander("Show the equation behind each column above"):
        _example = _opexfolder_wage_calc[0] if _opexfolder_wage_calc else None
        _example_input = st.session_state.opexfolder_wages[0] if st.session_state.opexfolder_wages else None
        _example_text = ""
        if _example is not None and _example_input is not None:
            _yearly_salary_ex = _example_input["monthly_wage"] * 12
            _example_text = f"""
**Worked example — {_example_input['rank']}** ({_example_input['headcount']} crew, {fmt(_example_input['monthly_wage'])}/month, {_example_input['tax_rate_pct']:.0f}% tax rate):

- Yearly salary (per person) = monthly wage × 12 = {fmt(_example_input['monthly_wage'])} × 12 = {fmt(_yearly_salary_ex)}
- Yearly wages (total) = headcount × yearly salary = {_example_input['headcount']} × {fmt(_yearly_salary_ex)} = **{fmt(_example['Yearly wages (total)'])}**
- Tax = tax rate × yearly wages (total) = {_example_input['tax_rate_pct']:.0f}% × {fmt(_example['Yearly wages (total)'])} = **{fmt(_example['Tax'])}**
- Max grant = max grant/head ({fmt(opexfolder_max_grant_per_head)}) × headcount = {fmt(opexfolder_max_grant_per_head)} × {_example_input['headcount']} = **{fmt(_example['Max grant'])}**
- Grant = MIN(Max grant, Tax) = MIN({fmt(_example['Max grant'])}, {fmt(_example['Tax'])}) = **{fmt(_example['Grant'])}**
- Salary above 7.1G = MAX(0, (yearly salary − 7.1G threshold) × headcount) = MAX(0, ({fmt(_yearly_salary_ex)} − {fmt(opexfolder_7_1g)}) × {_example_input['headcount']}) = **{fmt(_example['Salary above 7.1G (floored)'])}**
"""
        st.markdown(
            _example_text
            + "\nEach rank runs through the same five equations; the table above just "
              "shows the result for every rank at once. **Grant** is the Norwegian net "
              "wage scheme refund — the state reimburses the employer's payroll tax on "
              "Norwegian seafarers' wages, capped per head — and **Salary above 7.1G** "
              "feeds directly into the OTP (7.1G-12G) line in the Social costs table below."
        )

    st.markdown("**Social costs** (breakdown)")
    st.caption(
        "Each line uses the same formula type as the source workbook — "
        "provisions and travel cost scale with headcount and rotation; "
        "AGA/MPK/OTP/AFP scale with total wages (or wages above 7.1G for "
        "the second OTP line); the rest are flat per-head costs. Rates "
        "are all editable."
    )
    if "opexfolder_social_rates" not in st.session_state:
        st.session_state.opexfolder_social_rates = {
            "provisions_per_head_day": 280.0,
            "travel_cost_per_trip": 9_100.0,
            "aga_pct": 5.1,
            "mpk_pct": 3.3,
            "otp_12g_pct": 3.0,
            "otp_71g_12g_pct": 12.0,
            "afp_pct": 2.6,
            "health_insurance_per_head": 2_500.0,
            "travel_insurance_per_head": 1_000.0,
            "nautilus_per_head": 7_152.0,
            "lo_nho_per_head": 600.0,
            "norsk_maritim_per_head": 12_000.0,
            "gruppeliv_per_head": 1_020.0,
            "certificates_per_head": 1_500.0,
            "other_personnel_per_head": 3_000.0,
        }
    _sr = st.session_state.opexfolder_social_rates

    sc1, sc2, sc3 = st.columns(3)
    with sc1:
        _sr["provisions_per_head_day"] = stateful_number_input(
            "Provisions (NOK/head/day aboard)", min_value=0.0, value=_sr["provisions_per_head_day"], step=10.0,
            key="opexfolder_sr_provisions", disabled=locked
        )
        _sr["travel_cost_per_trip"] = stateful_number_input(
            "Travel cost (NOK/trip/head)", min_value=0.0, value=_sr["travel_cost_per_trip"], step=100.0,
            key="opexfolder_sr_travel", disabled=locked
        )
        _sr["aga_pct"] = stateful_number_input(
            "AGA (%, of total wages)", min_value=0.0, value=_sr["aga_pct"], step=0.1,
            key="opexfolder_sr_aga", disabled=locked
        )
        _sr["mpk_pct"] = stateful_number_input(
            "Employer premium MPK (%, of total wages)", min_value=0.0, value=_sr["mpk_pct"], step=0.1,
            key="opexfolder_sr_mpk", disabled=locked
        )
        _sr["otp_12g_pct"] = stateful_number_input(
            "OTP inntil 12G (%, of total wages)", min_value=0.0, value=_sr["otp_12g_pct"], step=0.1,
            key="opexfolder_sr_otp12g", disabled=locked
        )
    with sc2:
        _sr["otp_71g_12g_pct"] = stateful_number_input(
            "OTP 7.1G-12G (%, of salary above 7.1G)", min_value=0.0, value=_sr["otp_71g_12g_pct"], step=0.1,
            key="opexfolder_sr_otp71g", disabled=locked
        )
        _sr["afp_pct"] = stateful_number_input(
            "AFP (%, of total wages)", min_value=0.0, value=_sr["afp_pct"], step=0.1,
            key="opexfolder_sr_afp", disabled=locked
        )
        _sr["health_insurance_per_head"] = stateful_number_input(
            "Health insurance (NOK/head/yr)", min_value=0.0, value=_sr["health_insurance_per_head"], step=100.0,
            key="opexfolder_sr_health", disabled=locked
        )
        _sr["travel_insurance_per_head"] = stateful_number_input(
            "Travel insurance (NOK/head/yr)", min_value=0.0, value=_sr["travel_insurance_per_head"], step=100.0,
            key="opexfolder_sr_travelins", disabled=locked
        )
        _sr["nautilus_per_head"] = stateful_number_input(
            "Nautilus insurance (NOK/head/yr)", min_value=0.0, value=_sr["nautilus_per_head"], step=100.0,
            key="opexfolder_sr_nautilus", disabled=locked
        )
    with sc3:
        _sr["lo_nho_per_head"] = stateful_number_input(
            "LO/NHO/Kystrederiene (NOK/head/yr)", min_value=0.0, value=_sr["lo_nho_per_head"], step=100.0,
            key="opexfolder_sr_lonho", disabled=locked
        )
        _sr["norsk_maritim_per_head"] = stateful_number_input(
            "Norsk Maritim Kompetanse (NOK/head/yr)", min_value=0.0, value=_sr["norsk_maritim_per_head"], step=100.0,
            key="opexfolder_sr_nmk", disabled=locked
        )
        _sr["gruppeliv_per_head"] = stateful_number_input(
            "Gruppelivsforsikring (NOK/head/yr)", min_value=0.0, value=_sr["gruppeliv_per_head"], step=10.0,
            key="opexfolder_sr_gruppeliv", disabled=locked
        )
        _sr["certificates_per_head"] = stateful_number_input(
            "Certificates crew (NOK/head/yr)", min_value=0.0, value=_sr["certificates_per_head"], step=100.0,
            key="opexfolder_sr_certs", disabled=locked
        )
        _sr["other_personnel_per_head"] = stateful_number_input(
            "Other personnel cost (NOK/head/yr)", min_value=0.0, value=_sr["other_personnel_per_head"], step=100.0,
            key="opexfolder_sr_otherpersonnel", disabled=locked
        )

    _opexfolder_social_costs = {
        "Provisions": _sr["provisions_per_head_day"] * opexfolder_total_headcount * opexfolder_days_on_board,
        "Travel cost": _sr["travel_cost_per_trip"] * opexfolder_trips_per_year * opexfolder_total_headcount,
        "AGA": opexfolder_total_yearly_wages * (_sr["aga_pct"] / 100),
        "Employer premium MPK": opexfolder_total_yearly_wages * (_sr["mpk_pct"] / 100),
        "OTP (inntil 12G)": opexfolder_total_yearly_wages * (_sr["otp_12g_pct"] / 100),
        "OTP (7.1G-12G)": opexfolder_total_salary_above_71g * (_sr["otp_71g_12g_pct"] / 100),
        "AFP": opexfolder_total_yearly_wages * (_sr["afp_pct"] / 100),
        "Health insurance": _sr["health_insurance_per_head"] * opexfolder_total_headcount,
        "Travel insurance": _sr["travel_insurance_per_head"] * opexfolder_total_headcount,
        "Nautilus insurance": _sr["nautilus_per_head"] * opexfolder_total_headcount,
        "LO/NHO/Kystrederiene": _sr["lo_nho_per_head"] * opexfolder_total_headcount,
        "Norsk Maritim Kompetanse": _sr["norsk_maritim_per_head"] * opexfolder_total_headcount,
        "Gruppelivsforsikring": _sr["gruppeliv_per_head"] * opexfolder_total_headcount,
        "Certificates crew": _sr["certificates_per_head"] * opexfolder_total_headcount,
        "Other personnel cost": _sr["other_personnel_per_head"] * opexfolder_total_headcount,
    }
    opexfolder_total_social_costs = sum(_opexfolder_social_costs.values())
    social_costs_df = pd.DataFrame(
        [{"Item": k, "Annual cost (NOK)": v} for k, v in _opexfolder_social_costs.items()]
        + [{"Item": "Total", "Annual cost (NOK)": opexfolder_total_social_costs}]
    )
    show_table(social_costs_df, "Item", width="stretch")

    with st.expander("Show the equation behind each line above"):
        st.markdown(
            f"""
Five formula types, each applied consistently to every line of that type:

**1) Per head, per day aboard** — rate × total headcount × days on board/yr
- Provisions = {fmt(_sr['provisions_per_head_day'])} × {opexfolder_total_headcount:.0f} × {opexfolder_days_on_board:.1f} = **{fmt(_opexfolder_social_costs['Provisions'])}**

**2) Per trip, per head** — rate × trips/yr × total headcount
- Travel cost = {fmt(_sr['travel_cost_per_trip'])} × {opexfolder_trips_per_year:.0f} × {opexfolder_total_headcount:.0f} = **{fmt(_opexfolder_social_costs['Travel cost'])}**

**3) % of total annual wages** — rate% × total yearly wages ({fmt(opexfolder_total_yearly_wages)})
- AGA = {_sr['aga_pct']:.1f}% × {fmt(opexfolder_total_yearly_wages)} = **{fmt(_opexfolder_social_costs['AGA'])}**
- Employer premium MPK = {_sr['mpk_pct']:.1f}% × {fmt(opexfolder_total_yearly_wages)} = **{fmt(_opexfolder_social_costs['Employer premium MPK'])}**
- OTP (inntil 12G) = {_sr['otp_12g_pct']:.1f}% × {fmt(opexfolder_total_yearly_wages)} = **{fmt(_opexfolder_social_costs['OTP (inntil 12G)'])}**
- AFP = {_sr['afp_pct']:.1f}% × {fmt(opexfolder_total_yearly_wages)} = **{fmt(_opexfolder_social_costs['AFP'])}**

**4) % of wages above the 7.1G threshold** — rate% × total salary above 7.1G, floored at 0 per rank ({fmt(opexfolder_total_salary_above_71g)})
- OTP (7.1G-12G) = {_sr['otp_71g_12g_pct']:.1f}% × {fmt(opexfolder_total_salary_above_71g)} = **{fmt(_opexfolder_social_costs['OTP (7.1G-12G)'])}**

**5) Flat per head, per year** — rate × total headcount ({opexfolder_total_headcount:.0f})
- Health insurance = {fmt(_sr['health_insurance_per_head'])} × {opexfolder_total_headcount:.0f} = **{fmt(_opexfolder_social_costs['Health insurance'])}**
- Travel insurance = {fmt(_sr['travel_insurance_per_head'])} × {opexfolder_total_headcount:.0f} = **{fmt(_opexfolder_social_costs['Travel insurance'])}**
- Nautilus insurance = {fmt(_sr['nautilus_per_head'])} × {opexfolder_total_headcount:.0f} = **{fmt(_opexfolder_social_costs['Nautilus insurance'])}**
- LO/NHO/Kystrederiene = {fmt(_sr['lo_nho_per_head'])} × {opexfolder_total_headcount:.0f} = **{fmt(_opexfolder_social_costs['LO/NHO/Kystrederiene'])}**
- Norsk Maritim Kompetanse = {fmt(_sr['norsk_maritim_per_head'])} × {opexfolder_total_headcount:.0f} = **{fmt(_opexfolder_social_costs['Norsk Maritim Kompetanse'])}**
- Gruppelivsforsikring = {fmt(_sr['gruppeliv_per_head'])} × {opexfolder_total_headcount:.0f} = **{fmt(_opexfolder_social_costs['Gruppelivsforsikring'])}**
- Certificates crew = {fmt(_sr['certificates_per_head'])} × {opexfolder_total_headcount:.0f} = **{fmt(_opexfolder_social_costs['Certificates crew'])}**
- Other personnel cost = {fmt(_sr['other_personnel_per_head'])} × {opexfolder_total_headcount:.0f} = **{fmt(_opexfolder_social_costs['Other personnel cost'])}**

"Total yearly wages" and "Total salary above 7.1G" both come from the Grant scheme
table above — see that table's own **"Show the equation"** panel for how those two
figures are built, per rank, from headcount, monthly wage, and the 7.1G threshold.
            """
        )

    st.markdown("**Estimated use of temps**")
    opexfolder_temps_pct = stateful_number_input(
        "Estimated use of temps (%, of total wages)", min_value=0.0, value=2.5, step=0.5,
        key="opexfolder_temps_pct", disabled=locked
    )
    opexfolder_temps_cost = opexfolder_total_yearly_wages * (opexfolder_temps_pct / 100)

    opexfolder_crewcost_total = (
        opexfolder_total_yearly_wages + opexfolder_total_social_costs
        + opexfolder_temps_cost - opexfolder_total_grant
    )
    st.markdown(
        f"**Total crew cost** = Wages ({fmt(opexfolder_total_yearly_wages)}) + Social costs "
        f"({fmt(opexfolder_total_social_costs)}) + Temps ({fmt(opexfolder_temps_cost)}) − "
        f"Grant ({fmt(opexfolder_total_grant)}) = **{fmt(opexfolder_crewcost_total)}**"
    )

    st.divider()
    st.markdown("**Insurance**")
    ic1, ic2, ic3 = st.columns(3)
    with ic1:
        opexfolder_hull_machinery_pct = stateful_number_input(
            "Hull Machinery rate (%, of 80% of vessel value)", min_value=0.0, value=0.3, step=0.05,
            key="opexfolder_hull_pct", disabled=locked
        )
        opexfolder_hi_pct = stateful_number_input(
            "HI rate (%, of 20% of vessel value)", min_value=0.0, value=0.16, step=0.01,
            key="opexfolder_hi_pct", disabled=locked
        )
    with ic2:
        opexfolder_fi_pct = stateful_number_input(
            "FI rate (%, of 20% of vessel value)", min_value=0.0, value=0.16, step=0.01,
            key="opexfolder_fi_pct", disabled=locked
        )
        opexfolder_pi_flat = nok_input(
            "P&I (NOK, flat)", "opexfolder_pi_flat", 100_000.0, key="opexfolder_pi_input", disabled=locked
        )
    with ic3:
        opexfolder_war_flat = nok_input(
            "War & Equipment (NOK, flat)", "opexfolder_war_flat", 16_000.0, key="opexfolder_war_input", disabled=locked
        )

    _opexfolder_insurance = {
        "Hull Machinery": opexfolder_vessel_value * 0.8 * (opexfolder_hull_machinery_pct / 100),
        "HI": opexfolder_vessel_value * 0.2 * (opexfolder_hi_pct / 100),
        "FI": opexfolder_vessel_value * 0.2 * (opexfolder_fi_pct / 100),
        "P&I": opexfolder_pi_flat,
        "War & Equipment": opexfolder_war_flat,
    }
    opexfolder_insurance_total = sum(_opexfolder_insurance.values())
    insurance_df = pd.DataFrame(
        [{"Item": k, "Annual cost (NOK)": v} for k, v in _opexfolder_insurance.items()]
        + [{"Item": "Total", "Annual cost (NOK)": opexfolder_insurance_total}]
    )
    show_table(insurance_df, "Item", width="stretch")

    st.divider()
    st.markdown("**Other vessel cost** (cost-code breakdown)")
    if "opexfolder_other_vessel_cost" not in st.session_state:
        st.session_state.opexfolder_other_vessel_cost = [
            {"code": 4100, "name": "Ship general", "value_nok": 450_000.0},
            {"code": 4110, "name": "Ship repair & maintenance", "value_nok": 2_800_000.0},
            {"code": 4200, "name": "Hull", "value_nok": 0.0},
            {"code": 4300, "name": "Equipment for cargo", "value_nok": 0.0},
            {"code": 4400, "name": "Ship equipment", "value_nok": 0.0},
            {"code": 4500, "name": "Equipment for crew and passenger", "value_nok": 0.0},
            {"code": 4600, "name": "Machinery main components", "value_nok": 0.0},
            {"code": 4700, "name": "System for machinery", "value_nok": 0.0},
            {"code": 4800, "name": "Ship common system", "value_nok": 0.0},
        ]

    def _add_other_cost_row():
        st.session_state.opexfolder_other_vessel_cost.append({"code": 0, "name": "New item", "value_nok": 0.0})

    def _remove_other_cost_row(index):
        st.session_state.opexfolder_other_vessel_cost.pop(index)

    def _on_other_cost_value_change(index):
        raw = st.session_state[f"opexfolder_other_value_{index}"]
        value = parse_nok(raw)
        st.session_state.opexfolder_other_vessel_cost[index]["value_nok"] = value
        st.session_state[f"opexfolder_other_value_{index}"] = format_nok(value)

    ohdr = st.columns([0.7, 1.8, 1.2, 0.4])
    ohdr[0].markdown("**Code**")
    ohdr[1].markdown("**Name**")
    ohdr[2].markdown("**Annual cost (NOK)**")

    for i, item in enumerate(st.session_state.opexfolder_other_vessel_cost):
        c1, c2, c3, c4 = st.columns([0.7, 1.8, 1.2, 0.4])
        with c1:
            item["code"] = st.number_input(
                "Code", min_value=0, value=item["code"], step=100,
                key=f"opexfolder_other_code_{i}", label_visibility="collapsed", disabled=locked
            )
        with c2:
            item["name"] = st.text_input(
                "Name", value=item["name"], key=f"opexfolder_other_name_{i}", label_visibility="collapsed", disabled=locked
            )
        with c3:
            _other_value_display_key = f"opexfolder_other_value_{i}"
            if _other_value_display_key not in st.session_state:
                st.session_state[_other_value_display_key] = format_nok(item["value_nok"])
            st.text_input(
                "Annual cost", key=_other_value_display_key, label_visibility="collapsed",
                on_change=_on_other_cost_value_change, args=(i,), disabled=locked
            )
        with c4:
            st.button("✕", key=f"opexfolder_other_remove_{i}", on_click=_remove_other_cost_row, args=(i,), disabled=locked)

    st.button("+ Add cost code", key="opexfolder_other_add_code", on_click=_add_other_cost_row, disabled=locked)

    opexfolder_other_vessel_cost_total = sum(item["value_nok"] for item in st.session_state.opexfolder_other_vessel_cost)
    st.markdown(f"**Total other vessel cost:** {fmt(opexfolder_other_vessel_cost_total)}")

    st.divider()
    st.markdown("**Summary**")
    opexfolder_grand_total = opexfolder_crewcost_total + opexfolder_insurance_total + opexfolder_other_vessel_cost_total
    sm1, sm2, sm3, sm4 = st.columns(4)
    sm1.metric("Crew cost", fmt(opexfolder_crewcost_total))
    sm2.metric("Insurance", fmt(opexfolder_insurance_total))
    sm3.metric("Other vessel cost", fmt(opexfolder_other_vessel_cost_total))
    sm4.metric("Total vessel opex", fmt(opexfolder_grand_total))

    _opexfolder_push_map = {
        "Crewing": opexfolder_crewcost_total,
        "Insurance": opexfolder_insurance_total,
        "Other vessel opex": opexfolder_other_vessel_cost_total,
    }

    def _sync_opex_totals_to_tab1():
        for _name, _value in _opexfolder_push_map.items():
            _existing = next((it for it in st.session_state.opex_items if it["name"] == _name), None)
            if _existing is not None:
                _existing["value_nok"] = _value
            else:
                st.session_state.opex_items.append({"name": _name, "value_nok": _value})
            # Tab 1's own "value_{i}" text widgets already rendered earlier
            # in this same pass, so their session_state can't be SET
            # directly here (Streamlit forbids assigning to an
            # already-instantiated widget's key within the same run —
            # same class of issue as the button-key bug fixed earlier).
            # Deleting the stale entry instead is safe, and lets Tab 1's
            # own "if key not in session_state: initialize from
            # item['value_nok']" logic correctly re-populate it fresh on
            # the next pass (triggered by _request_rerun() below).
            for _idx, _it in enumerate(st.session_state.opex_items):
                if _it["name"] == _name and f"value_{_idx}" in st.session_state:
                    del st.session_state[f"value_{_idx}"]

    if opex_linked_to_vessel_opex_tab:
        st.caption(
            "**Linked** — Crewing, Insurance, and Other vessel opex on Tab "
            "1 update automatically from the totals above (creating them "
            "if they don't already exist). Tab 1's value fields for these "
            "three are disabled while linked; turn the link off on Tab 1 "
            "to edit them manually there again."
        )
        def _opex_value_matches(name, target_value, tol=1.0):
            current = next((it["value_nok"] for it in st.session_state.opex_items if it["name"] == name), None)
            if current is None:
                return False
            return abs(current - target_value) <= tol

        _tab1_currently_matches = all(
            _opex_value_matches(_name, _value) for _name, _value in _opexfolder_push_map.items()
        )
        _opex_sync_retry_count = st.session_state.get("_opex_sync_retry_count", 0)
        if not _tab1_currently_matches and _opex_sync_retry_count < 4:
            st.session_state["_opex_sync_retry_count"] = _opex_sync_retry_count + 1
            _sync_opex_totals_to_tab1()
            _request_rerun()
        else:
            st.session_state["_opex_sync_retry_count"] = 0
    else:
        st.caption(
            "**Not linked** — Tab 1's Crewing/Insurance/Other vessel opex "
            "are independently editable there and won't update from the "
            "totals above automatically. Turn the link on there, or use "
            "the button below for a one-off manual sync without turning "
            "the link on."
        )
        if st.button("Push totals to Tab 1's vessel opex (one-off)", disabled=locked):
            _sync_opex_totals_to_tab1()
            st.success("Pushed to Tab 1 — Crewing, Insurance, and Other vessel opex updated.")
            _request_rerun()

# ===========================================================================
# TAB 1.75 — Construction capex (currency-converted capex build-up, and the
#            equity-first installment waterfall with two-tranche debt —
#            construction debt for installments 1..N-1, take-out debt for
#            the final installment. Core waterfall only for now; day-count
#            fee accrual — commitment/counter-guarantee/take-out-guarantee
#            fees — is a planned follow-up, not built here yet.)
# ===========================================================================
with tab_construction:
    st.subheader("Construction capex & finance")
    st.caption(
        "Currency-converted capex build-up, and the installment waterfall "
        "to the yard — equity funds each installment first, up to "
        "however much is available; once equity runs out, construction "
        "debt covers the rest. The final installment is different: it's "
        "funded entirely by a separate take-out debt facility (which "
        "refinances/replaces the construction debt at delivery), not by "
        "the equity-first waterfall. Works with anywhere from 2 to 6 "
        "installments — construction debt only ever covers up to "
        "whichever installment sits second-to-last, however many there are."
    )

    st.markdown("**FX rates** (NOK per unit of foreign currency)")
    fxc1, fxc2, fxc3, fxc4, fxc5 = st.columns(5)
    with fxc1:
        construction_fx_eur = stateful_number_input(
            "EUR/NOK", min_value=0.0, value=11.65, step=0.05,
            key="construction_fx_eur", disabled=locked
        )
    with fxc2:
        construction_fx_usd = stateful_number_input(
            "USD/NOK", min_value=0.0, value=10.70, step=0.05,
            key="construction_fx_usd", disabled=locked
        )
    with fxc3:
        construction_fx_gbp = stateful_number_input(
            "GBP/NOK", min_value=0.0, value=13.50, step=0.05,
            key="construction_fx_gbp", disabled=locked
        )
    with fxc4:
        construction_fx_cad = stateful_number_input(
            "CAD/NOK", min_value=0.0, value=7.70, step=0.05,
            key="construction_fx_cad", disabled=locked
        )
    with fxc5:
        construction_fx_clp = stateful_number_input(
            "CLP/NOK", min_value=0.0, value=0.011, step=0.001, format="%.4f",
            key="construction_fx_clp", disabled=locked
        )
    _construction_fx_lookup = {
        "NOK": 1.0, "EUR": construction_fx_eur, "USD": construction_fx_usd,
        "GBP": construction_fx_gbp, "CAD": construction_fx_cad, "CLP": construction_fx_clp,
    }

    st.markdown("**Capex items subject to the installment schedule** (yard contract and related)")
    st.caption(
        "These items sum to the 'installment capex' that gets split "
        "across the installment schedule below. Each item has its own "
        "currency, converted to NOK at the rates above."
    )
    if "construction_capex_items" not in st.session_state:
        st.session_state.construction_capex_items = [
            {"name": "Yard contract", "currency": "EUR", "amount": 59_000_000.0},
            {"name": "RO plant", "currency": "NOK", "amount": 4_000_000.0},
            {"name": "FLS", "currency": "NOK", "amount": 4_000_000.0},
            {"name": "Delicer", "currency": "EUR", "amount": 0.0},
            {"name": "Contingencies", "currency": "EUR", "amount": 0.0},
        ]

    def _add_construction_capex_item():
        st.session_state.construction_capex_items.append({"name": "New item", "currency": "NOK", "amount": 0.0})

    def _remove_construction_capex_item(index):
        st.session_state.construction_capex_items.pop(index)

    def _on_construction_capex_amount_change(index):
        raw = st.session_state[f"construction_capex_amount_{index}"]
        value = parse_nok(raw)
        st.session_state.construction_capex_items[index]["amount"] = value
        st.session_state[f"construction_capex_amount_{index}"] = format_nok(value)

    cchdr = st.columns([1.8, 1.0, 1.4, 1.4, 0.4])
    cchdr[0].markdown("**Name**")
    cchdr[1].markdown("**Currency**")
    cchdr[2].markdown("**Amount (in currency)**")
    cchdr[3].markdown("**NOK value**")

    _construction_installment_capex_nok = 0.0
    _construction_capex_items_nok = []
    for i, item in enumerate(st.session_state.construction_capex_items):
        c1, c2, c3, c4, c5 = st.columns([1.8, 1.0, 1.4, 1.4, 0.4])
        with c1:
            item["name"] = st.text_input(
                "Name", value=item["name"], key=f"construction_capex_name_{i}", label_visibility="collapsed", disabled=locked
            )
        with c2:
            item["currency"] = st.selectbox(
                "Currency", ["NOK", "EUR", "USD", "GBP", "CAD", "CLP"],
                index=["NOK", "EUR", "USD", "GBP", "CAD", "CLP"].index(item["currency"]),
                key=f"construction_capex_currency_{i}", label_visibility="collapsed", disabled=locked
            )
        with c3:
            _capex_amount_display_key = f"construction_capex_amount_{i}"
            if _capex_amount_display_key not in st.session_state:
                st.session_state[_capex_amount_display_key] = format_nok(item["amount"])
            st.text_input(
                "Amount", key=_capex_amount_display_key, label_visibility="collapsed",
                on_change=_on_construction_capex_amount_change, args=(i,), disabled=locked
            )
        with c4:
            _item_nok_value = item["amount"] * _construction_fx_lookup.get(item["currency"], 1.0)
            st.markdown(f"<div style='padding-top:8px'>{fmt(_item_nok_value)}</div>", unsafe_allow_html=True)
        with c5:
            st.button("✕", key=f"construction_capex_remove_{i}", on_click=_remove_construction_capex_item, args=(i,), disabled=locked)
        _construction_installment_capex_nok += _item_nok_value
        _construction_capex_items_nok.append({"name": item["name"], "nok_value": _item_nok_value})

    st.button("+ Add capex item", key="construction_capex_add_item", on_click=_add_construction_capex_item, disabled=locked)
    st.markdown(f"**Total installment capex:** {fmt(_construction_installment_capex_nok)}")

    st.markdown("**Non-installment capitalized costs** (always equity-funded, spread evenly across installments)")
    st.caption(
        "Project management, legal costs, and similar — capitalized, but "
        "not part of the yard-contract waterfall. Always equity, split "
        "evenly across however many installments exist below."
    )
    if "construction_other_costs" not in st.session_state:
        st.session_state.construction_other_costs = [
            {"name": "Project management", "amount_nok": 9_000_000.0},
            {"name": "Legal & owner's supply", "amount_nok": 4_500_000.0},
        ]

    def _add_construction_other_cost():
        st.session_state.construction_other_costs.append({"name": "New item", "amount_nok": 0.0})

    def _remove_construction_other_cost(index):
        st.session_state.construction_other_costs.pop(index)

    def _on_construction_other_cost_change(index):
        raw = st.session_state[f"construction_other_amount_{index}"]
        value = parse_nok(raw)
        st.session_state.construction_other_costs[index]["amount_nok"] = value
        st.session_state[f"construction_other_amount_{index}"] = format_nok(value)

    ochdr = st.columns([2.4, 1.6, 0.4])
    ochdr[0].markdown("**Name**")
    ochdr[1].markdown("**Amount (NOK)**")

    _construction_other_capex_nok = 0.0
    for i, item in enumerate(st.session_state.construction_other_costs):
        c1, c2, c3 = st.columns([2.4, 1.6, 0.4])
        with c1:
            item["name"] = st.text_input(
                "Name", value=item["name"], key=f"construction_other_name_{i}", label_visibility="collapsed", disabled=locked
            )
        with c2:
            _other_amount_display_key = f"construction_other_amount_{i}"
            if _other_amount_display_key not in st.session_state:
                st.session_state[_other_amount_display_key] = format_nok(item["amount_nok"])
            st.text_input(
                "Amount", key=_other_amount_display_key, label_visibility="collapsed",
                on_change=_on_construction_other_cost_change, args=(i,), disabled=locked
            )
        with c3:
            st.button("✕", key=f"construction_other_remove_{i}", on_click=_remove_construction_other_cost, args=(i,), disabled=locked)
        _construction_other_capex_nok += item["amount_nok"]

    st.button("+ Add item", key="construction_other_add_item", on_click=_add_construction_other_cost, disabled=locked)
    st.markdown(f"**Total non-installment capex:** {fmt(_construction_other_capex_nok)}")

    _construction_total_capitalized_nok = _construction_installment_capex_nok + _construction_other_capex_nok
    st.markdown(f"**Total capitalized cost (excl. construction finance cost & guarantee premium):** {fmt(_construction_total_capitalized_nok)}")

    _construction_finance_cost_prev_pass = st.session_state.get("_construction_finance_cost_total", 0.0)
    _construction_guarantee_premium_prev_pass = st.session_state.get("_construction_guarantee_premium_total", 0.0)
    st.markdown(
        f"**Total capitalized cost, incl. construction finance cost & guarantee premium:** "
        f"{fmt(_construction_total_capitalized_nok + _construction_finance_cost_prev_pass + _construction_guarantee_premium_prev_pass)}"
    )
    st.caption(
        "Construction finance cost and the guarantee premium are both "
        "computed further down (they depend on the installment schedule "
        "and debt waterfall below), so this figure is one script pass "
        "behind after changing debt, installment, or fee-rate inputs — "
        "it catches up automatically; switch tabs or click Refresh in "
        "the sidebar if it looks stale."
    )

    st.divider()
    st.markdown("**Debt sizing**")
    construction_total_debt_nok = nok_input(
        "Total debt (NOK)", "construction_total_debt_nok", 576_000_000.0,
        key="construction_total_debt_input", disabled=locked
    )
    _construction_yard_contract_item = next(
        (it for it in st.session_state.construction_capex_items if it["name"] == "Yard contract"), None
    )
    _construction_yard_contract_nok = (
        _construction_yard_contract_item["amount"] * _construction_fx_lookup.get(_construction_yard_contract_item["currency"], 1.0)
        if _construction_yard_contract_item is not None else 0.0
    )
    _construction_ltv1 = (construction_total_debt_nok / _construction_yard_contract_nok) if _construction_yard_contract_nok else 0.0
    _construction_ltv2 = (construction_total_debt_nok / _construction_total_capitalized_nok) if _construction_total_capitalized_nok else 0.0
    dc1, dc2, dc3 = st.columns(3)
    dc1.metric("LTV1 — debt / yard contract", f"{_construction_ltv1:.1%}")
    dc2.metric("LTV2 — debt / total capitalized cost", f"{_construction_ltv2:.1%}")
    _construction_total_equity_nok = _construction_total_capitalized_nok - construction_total_debt_nok
    dc3.metric("Total equity for capex", fmt(_construction_total_equity_nok))
    if _construction_total_equity_nok < 0:
        st.warning(
            f"⚠️ Debt ({fmt(construction_total_debt_nok)}) exceeds total capitalized cost "
            f"({fmt(_construction_total_capitalized_nok)}) — equity would be negative. Check your inputs."
        )

    st.divider()
    st.markdown("**Installment schedule** (2 to 6 installments — the last one is always take-out financing)")
    st.caption(
        "Each installment is a share of the installment capex above — "
        "shares should sum to 100%. The last row is always treated as "
        "take-out financing: funded entirely by a dedicated take-out "
        "debt facility, not the equity-first waterfall used for every "
        "other installment. Add or remove rows freely between 2 and 6."
    )
    if "construction_installments" not in st.session_state:
        st.session_state.construction_installments = [
            {"share_pct": 20.0, "month": 0.0},
            {"share_pct": 20.0, "month": 7.5},
            {"share_pct": 20.0, "month": 15.0},
            {"share_pct": 20.0, "month": 22.5},
            {"share_pct": 20.0, "month": 30.0},
        ]

    def _add_construction_installment():
        if len(st.session_state.construction_installments) < 6:
            st.session_state.construction_installments.append({"share_pct": 0.0, "month": 0.0})

    def _remove_construction_installment(index):
        if len(st.session_state.construction_installments) > 2:
            st.session_state.construction_installments.pop(index)

    _construction_n_installments = len(st.session_state.construction_installments)
    ishdr = st.columns([1.6, 1.0, 1.0, 0.4])
    ishdr[0].markdown("**Installment**")
    ishdr[1].markdown("**Share of installment capex (%)**")
    ishdr[2].markdown("**Month**")

    for i, inst in enumerate(st.session_state.construction_installments):
        is_last = (i == _construction_n_installments - 1)
        _label = "Take-out-financing" if is_last else f"{i + 1}{'st' if i == 0 else 'nd' if i == 1 else 'rd' if i == 2 else 'th'} yard-installment"
        c1, c2, c3, c4 = st.columns([1.6, 1.0, 1.0, 0.4])
        with c1:
            st.markdown(f"<div style='padding-top:8px'>{_label}</div>", unsafe_allow_html=True)
        with c2:
            inst["share_pct"] = st.number_input(
                "Share (%)", min_value=0.0, max_value=100.0, value=inst["share_pct"], step=1.0,
                key=f"construction_installment_share_{i}", label_visibility="collapsed", disabled=locked
            )
        with c3:
            inst["month"] = st.number_input(
                "Month", min_value=0.0, value=inst["month"], step=0.5,
                key=f"construction_installment_month_{i}", label_visibility="collapsed", disabled=locked
            )
        with c4:
            if _construction_n_installments > 2:
                st.button("✕", key=f"construction_installment_remove_{i}", on_click=_remove_construction_installment, args=(i,), disabled=locked)

    _install_add_col1, _install_add_col2 = st.columns([1, 3])
    with _install_add_col1:
        if _construction_n_installments < 6:
            st.button("+ Add installment", key="construction_installment_add_row", on_click=_add_construction_installment, disabled=locked)
        else:
            st.caption("Maximum of 6 installments reached.")

    _construction_share_sum = sum(inst["share_pct"] for inst in st.session_state.construction_installments)
    if abs(_construction_share_sum - 100.0) > 0.5:
        st.warning(f"⚠️ Shares sum to {_construction_share_sum:.1f}%, not 100%.")

    # --- the waterfall: equity funds each installment (except the last)
    # first, up to whatever's left; debt covers the rest. The last
    # installment is funded 100% by take-out debt, by definition — not
    # subject to the equity-first rule, regardless of how much equity
    # happens to remain at that point. Uses "installment capex minus total
    # debt" as its OWN equity pool — NOT total equity (which also nets
    # off PM/Legal's equity draw) — since PM/Legal is tracked separately
    # and doesn't compete with the yard-installment waterfall for the
    # same funding pool (matches the source file's own G41 = E41-H41,
    # using installment-capex specifically, not the overall capitalized
    # cost). ---
    _construction_installment_equity_pool = _construction_installment_capex_nok - construction_total_debt_nok
    _construction_remaining_equity = _construction_installment_equity_pool
    _construction_waterfall_rows = []
    _construction_construction_debt_draws = []
    _construction_takeout_debt_amount = 0.0
    for i, inst in enumerate(st.session_state.construction_installments):
        is_last = (i == _construction_n_installments - 1)
        _label = "Take-out-financing" if is_last else f"{i + 1}{'st' if i == 0 else 'nd' if i == 1 else 'rd' if i == 2 else 'th'} yard-installment"
        _amount = _construction_installment_capex_nok * (inst["share_pct"] / 100)
        if is_last:
            _equity_portion = 0.0
            _debt_portion = _amount
            _construction_takeout_debt_amount = _amount
        else:
            _equity_portion = min(_construction_remaining_equity, _amount)
            _debt_portion = _amount - _equity_portion
            _construction_remaining_equity -= _equity_portion
            _construction_construction_debt_draws.append(_debt_portion)
        _construction_waterfall_rows.append({
            "Installment": _label, "Month": inst["month"], "Amount": _amount,
            "Equity": _equity_portion, "Debt": _debt_portion,
        })

    waterfall_df = pd.DataFrame(_construction_waterfall_rows)
    show_table(waterfall_df, "Installment", width="stretch")

    _construction_construction_debt_commitment = sum(_construction_construction_debt_draws)
    st.markdown("**Debt tranches**")
    tc1, tc2, tc3 = st.columns(3)
    tc1.metric("Construction debt commitment", fmt(_construction_construction_debt_commitment))
    tc2.metric("Take-out debt", fmt(_construction_takeout_debt_amount))
    tc3.metric("Total debt (check)", fmt(_construction_construction_debt_commitment + _construction_takeout_debt_amount))
    st.caption(
        "Construction debt commitment = total debt minus the take-out "
        "installment's amount — it only ever covers installments 1 "
        "through the second-to-last, however many there are. Take-out "
        "debt is the separate facility that funds the final installment "
        "and refinances the construction debt at delivery."
    )
    _construction_debt_check = _construction_construction_debt_commitment + _construction_takeout_debt_amount
    if abs(_construction_debt_check - construction_total_debt_nok) > 1.0:
        st.warning(
            f"⚠️ Construction debt + take-out debt ({fmt(_construction_debt_check)}) doesn't "
            f"match total debt input ({fmt(construction_total_debt_nok)}) — this can happen if "
            f"equity is negative or shares don't sum to 100%; check the inputs above."
        )

    st.divider()
    st.markdown("**Construction finance cost**")
    st.caption(
        "Interest on the construction debt drawn at each installment, "
        "accruing (simple interest, not compounding) from that "
        "installment's own month until take-out (the last installment's "
        "month) — the construction debt gets repaid/refinanced at that "
        "point, so nothing accrues beyond it. Capitalized: added to "
        "total capitalized cost below, since interest during "
        "construction is normally added to the vessel's cost basis "
        "rather than expensed as incurred. Rate is set independently "
        "here, not linked to Tab 1 — useful since construction finance "
        "is often in a different currency (e.g. a EUR loan) with its "
        "own market rate."
    )
    cfc1, cfc2 = st.columns(2)
    with cfc1:
        construction_swap_rate_pct = stateful_number_input(
            "Swap rate (%/yr)", min_value=0.0, value=4.0, step=0.1,
            key="construction_swap_rate", disabled=locked
        )
    with cfc2:
        construction_credit_spread_pct = stateful_number_input(
            "Credit spread (%/yr)", min_value=0.0, value=3.5, step=0.1,
            key="construction_credit_spread", disabled=locked
        )
    _construction_finance_cost_rate_pct = construction_swap_rate_pct + construction_credit_spread_pct
    st.caption(f"Total construction finance cost rate: {_construction_finance_cost_rate_pct:.2f}%/yr.")

    # --- computed AFTER the waterfall above has already fixed each
    # installment's debt draw — finance cost here is purely a function of
    # that fixed waterfall, never fed back into the debt sizing itself.
    # Feeding it back would create exactly the kind of circular reference
    # that appears to be what broke in the source workbook (#REF! errors
    # cascading from this same section). ---
    _construction_takeout_month = st.session_state.construction_installments[-1]["month"]
    _construction_finance_cost_rows = []
    _construction_finance_cost_total = 0.0
    for i, _row in enumerate(_construction_waterfall_rows):
        if i == _construction_n_installments - 1:
            continue  # take-out installment itself doesn't accrue construction finance cost
        _months_outstanding = max(0.0, _construction_takeout_month - _row["Month"])
        _interest = _row["Debt"] * (_construction_finance_cost_rate_pct / 100) * (_months_outstanding / 12)
        _construction_finance_cost_rows.append({
            "Installment": _row["Installment"], "Debt drawn": _row["Debt"],
            "Months outstanding": _months_outstanding, "Finance cost": _interest,
        })
        _construction_finance_cost_total += _interest

    if _construction_finance_cost_rows:
        finance_cost_df = pd.DataFrame(
            _construction_finance_cost_rows
            + [{"Installment": "Total", "Debt drawn": None, "Months outstanding": None, "Finance cost": _construction_finance_cost_total}]
        )
        show_table(finance_cost_df, "Installment", decimal_cols=["Months outstanding"], width="stretch")
    st.metric("Total construction finance cost", fmt(_construction_finance_cost_total))

    # --- store for the earlier "Total capitalized cost, incl. construction
    # finance cost" display (above, before the waterfall this depends on
    # has even run) — self-healing: if it just changed, trigger one more
    # pass so that earlier display catches up immediately rather than
    # waiting for the user to switch tabs or click Refresh. ---
    _construction_finance_cost_changed = (
        abs(st.session_state.get("_construction_finance_cost_total", 0.0) - _construction_finance_cost_total) > 1.0
    )
    st.session_state["_construction_finance_cost_total"] = _construction_finance_cost_total
    _construction_fc_retry_count = st.session_state.get("_construction_fc_retry_count", 0)
    if _construction_finance_cost_changed and _construction_fc_retry_count < 4:
        st.session_state["_construction_fc_retry_count"] = _construction_fc_retry_count + 1
        _request_rerun()
    else:
        st.session_state["_construction_fc_retry_count"] = 0

    st.divider()
    st.markdown("**Guarantee premium on unutilized construction debt**")
    st.caption(
        "The construction debt commitment is a facility, not a lump sum "
        "— only the part actually drawn earns interest (the finance "
        "cost above); the undrawn remainder still costs a guarantee "
        "premium for keeping the facility available. As each "
        "installment draws more of the facility, the unutilized balance "
        "steps down, so the premium accrues period by period at a "
        "shrinking base — reaching zero once the facility is fully "
        "drawn (which happens at the last yard-installment, one before "
        "take-out)."
    )
    construction_guarantee_premium_pct = stateful_number_input(
        "Guarantee premium rate (%/yr, of unutilized construction debt)", min_value=0.0, value=1.35, step=0.05,
        key="construction_guarantee_premium_rate", disabled=locked
    )

    _construction_guarantee_rows = []
    _construction_guarantee_premium_total = 0.0
    _construction_cumulative_drawn = 0.0
    for i in range(_construction_n_installments - 1):  # excludes take-out itself
        _construction_cumulative_drawn += _construction_waterfall_rows[i]["Debt"]
        _unutilized = max(0.0, _construction_construction_debt_commitment - _construction_cumulative_drawn)
        _period_start_month = st.session_state.construction_installments[i]["month"]
        _period_end_month = st.session_state.construction_installments[i + 1]["month"]
        _period_months = max(0.0, _period_end_month - _period_start_month)
        _premium_this_period = _unutilized * (construction_guarantee_premium_pct / 100) * (_period_months / 12)
        _construction_guarantee_rows.append({
            "Period": f"After {_construction_waterfall_rows[i]['Installment']}",
            "Cumulative drawn": _construction_cumulative_drawn,
            "Unutilized": _unutilized,
            "Months in period": _period_months,
            "Guarantee premium": _premium_this_period,
        })
        _construction_guarantee_premium_total += _premium_this_period

    if _construction_guarantee_rows:
        guarantee_df = pd.DataFrame(
            _construction_guarantee_rows
            + [{"Period": "Total", "Cumulative drawn": None, "Unutilized": None, "Months in period": None, "Guarantee premium": _construction_guarantee_premium_total}]
        )
        show_table(guarantee_df, "Period", decimal_cols=["Months in period"], width="stretch")
    st.metric("Total guarantee premium", fmt(_construction_guarantee_premium_total))

    # --- same one-pass-behind, self-healing storage as finance cost above ---
    _construction_guarantee_changed = (
        abs(st.session_state.get("_construction_guarantee_premium_total", 0.0) - _construction_guarantee_premium_total) > 1.0
    )
    st.session_state["_construction_guarantee_premium_total"] = _construction_guarantee_premium_total
    _construction_gp_retry_count = st.session_state.get("_construction_gp_retry_count", 0)
    if _construction_guarantee_changed and _construction_gp_retry_count < 4:
        st.session_state["_construction_gp_retry_count"] = _construction_gp_retry_count + 1
        _request_rerun()
    else:
        st.session_state["_construction_gp_retry_count"] = 0

    _construction_total_capitalized_incl_finance_nok = (
        _construction_total_capitalized_nok + _construction_finance_cost_total + _construction_guarantee_premium_total
    )
    st.markdown(
        f"**Total capitalized cost, incl. construction finance cost & guarantee premium:** "
        f"{fmt(_construction_total_capitalized_incl_finance_nok)}"
    )
    st.caption(
        "This is the full implied vessel capex once construction "
        "finance cost and the guarantee premium are both capitalized on "
        "top — carried through into the Sources & uses check below, "
        "funded by additional equity. Neither is fed back into the "
        "installment waterfall or debt sizing above, though — the "
        "waterfall only depends on installment capex and total debt, "
        "both fixed before either fee is even calculated, so there's no "
        "circularity in adding them here. If you want this larger "
        "figure reflected as Tab 1's vessel capex, update it there "
        "directly."
    )

    st.divider()
    st.markdown("**Sources & uses check**")
    st.caption(
        "Includes construction finance cost and the guarantee premium "
        "as part of the capitalized cost of completing the vessel — "
        "both funded by additional equity, since debt stays fixed at "
        "whatever you've set above (same treatment as Project "
        "management/Legal). This is what pushes the total above the "
        "installment-capex-only figure."
    )
    _construction_other_equity_nok = _construction_other_capex_nok  # always equity, per the caption above
    _construction_installment_equity_used = _construction_installment_equity_pool - _construction_remaining_equity
    su1, su2 = st.columns(2)
    with su1:
        st.markdown("**Uses**")
        uses_df = pd.DataFrame(
            [{"Item": _it["name"], "Amount": _it["nok_value"]} for _it in _construction_capex_items_nok]
            + [
                {"Item": "Non-installment capex (PM, legal, etc.)", "Amount": _construction_other_capex_nok},
                {"Item": "Construction finance cost", "Amount": _construction_finance_cost_total},
                {"Item": "Guarantee premium", "Amount": _construction_guarantee_premium_total},
                {"Item": "Total uses", "Amount": _construction_total_capitalized_incl_finance_nok},
            ]
        )
        show_table(uses_df, "Item", width="stretch")
    with su2:
        st.markdown("**Sources**")
        sources_df = pd.DataFrame([
            {"Item": "Construction debt", "Amount": _construction_construction_debt_commitment},
            {"Item": "Take-out debt", "Amount": _construction_takeout_debt_amount},
            {"Item": "Equity (installments)", "Amount": _construction_installment_equity_used},
            {"Item": "Equity (non-installment costs)", "Amount": _construction_other_equity_nok},
            {"Item": "Equity (construction finance cost)", "Amount": _construction_finance_cost_total},
            {"Item": "Equity (guarantee premium)", "Amount": _construction_guarantee_premium_total},
            {"Item": "Unused equity", "Amount": max(0.0, _construction_remaining_equity)},
            {"Item": "Total sources", "Amount": (
                _construction_construction_debt_commitment + _construction_takeout_debt_amount
                + _construction_installment_equity_used + _construction_other_equity_nok
                + _construction_finance_cost_total + _construction_guarantee_premium_total
                + max(0.0, _construction_remaining_equity)
            )},
        ])
        show_table(sources_df, "Item", width="stretch")

# ===========================================================================
# TAB 1b — Spot market (alternative revenue basis, with transparent
#          voyage cost recovery)
# ===========================================================================
with tab_spot:
    st.subheader("Spot market revenue")
    st.caption(
        "An alternative to the TC-rate above: revenue is a market day-rate "
        "for vessel capacity, plus separate charge lines that recover "
        "voyage costs (fuel, lube oil, port fees, waste, farledsavgift, pH "
        "adjustment) from the customer — shown transparently against what "
        "each actually costs. Unlike a TC charter, these voyage costs sit "
        "for the **owner's** account, not the charterer's, which is why "
        "they need to be billed back explicitly rather than assumed away."
    )

    spot_market_enabled = stateful_toggle(
        "Use spot-market revenue instead of the TC-rate in the Financial Statements",
        value=False, key="spot_market_enabled", disabled=locked
    )
    if not spot_market_enabled:
        st.info(
            "Spot market is currently **off** — the Financial Statements tab "
            "still uses the vessel TC-rate (Tab 1). Inputs below are still "
            "editable so this is ready whenever you switch it on."
        )

    st.markdown("**Utilization & service mix**")
    st.caption(
        "Not every calendar day earns revenue — utilization sets what share "
        "of operating days are actually working. Of those working days, the "
        "service mix splits them across job types (treatment, smolt "
        "transport, harvest transport, etc.), each at its own day-rate. "
        "Vessel opex (crewing — Tab 1) is unaffected by utilization, since "
        "crew cost doesn't stop during idle days; voyage costs below scale "
        "with utilization the same way revenue does, since idle days don't "
        "burn fuel or incur port fees either."
    )

    st.markdown("**Service mix** (% of activity, by service type)")
    st.caption(
        "Set overall utilization once, then split it across services as a "
        "percentage — days/year are computed automatically from these two "
        "numbers rather than typed in directly. Same underlying math as "
        "before (days = utilization x operating days x share), just entered "
        "the way it's easiest to work through with the board or commercial "
        "team: '65% utilization, split 70/20/10', not a calculator."
    )

    # Default weekly activity plan — the Weekly activity calendar's own
    # starting point further down the page. Defined here, early, so
    # even the very first pass (before the calendar section has
    # rendered even once) can derive a sensible implied utilization
    # from it directly, rather than flashing a hardcoded placeholder.
    # (activity, occupancy_pct) per week, W1 through W52 — every week
    # defaults to 90% occupancy regardless of activity (including Idle
    # weeks, though occupancy has no downstream effect for those).
    _default_calendar_weeks = (
        [("Treatment of fish", 90.0)] * 2
        + [("Idle", 90.0)] * 7
        + [("Transport", 90.0)] * 4
        + [("Treatment of fish", 90.0)] * 13
        + [("Idle", 90.0)] * 5
        + [("Treatment of fish", 90.0)] * 4
        + [("Transport", 90.0)] * 4
        + [("Treatment of fish", 90.0)] * 12
        + [("Idle", 90.0)] * 1
    )

    if st.session_state.get("spot_calendar_enabled", True):
        _calendar_implied_utilization_pct = st.session_state.get("_calendar_implied_utilization_pct")
        if _calendar_implied_utilization_pct is None:
            # First-ever pass — derive directly from the default plan
            # above so there's no flash of a placeholder before the
            # calendar has rendered once.
            _calendar_implied_utilization_pct = sum(
                occ for (act, occ) in _default_calendar_weeks if act != "Idle"
            ) / 52
        spot_utilization_pct = _calendar_implied_utilization_pct
        st.metric("Utilization (implied from Monthly activity calendar)", f"{spot_utilization_pct:.1f}%")
        st.caption(
            "Set by the Monthly activity calendar further down this page "
            "(active months x occupancy, ÷ 12) — not a separate input "
            "while the calendar is on. Turn the calendar off there to set "
            "utilization manually again."
        )
    else:
        spot_utilization_pct = stateful_number_input(
            "Utilization (%)", min_value=0.0, max_value=100.0, value=65.0, step=1.0,
            key="spot_utilization_pct", disabled=locked
        )
    _working_days_annual_target = operating_days * (spot_utilization_pct / 100)
    st.caption(f"= {fmt(_working_days_annual_target)} working days/year, out of {fmt(operating_days)} operating days (Tab 1) — this is the Year 1 baseline; see the Capacity schedule below to change it for later years.")

    if "spot_service_items" not in st.session_state:
        st.session_state.spot_service_items = [
            {"name": "Treatment of fish", "share_pct": 70.0, "rate_nok_day": 720_000.0, "escalator_pct": 3.0, "priced_at_baseline": False},
            {"name": "Transport", "share_pct": 20.0, "rate_nok_day": 456_000.0, "escalator_pct": 2.0, "priced_at_baseline": False},
            {"name": "Other", "share_pct": 10.0, "rate_nok_day": 456_000.0, "escalator_pct": 2.0, "priced_at_baseline": False},
        ]

    st.markdown("**Fixed Voyage opex (annual budget, spread over working days)**")
    st.caption(
        "Unlike Tab 1's crewing/vessel opex — which is fixed and applies "
        "every calendar day, since crew salaries don't stop when idle — "
        "this treats fixed voyage opex as a separate **annual** budget "
        "that gets spread over however many days are actually worked. "
        "Defaults to 0 — this is a fuse for future overhead specific to "
        "running the spot-trade business (e.g. a dedicated spot-trade "
        "commercial hire), not currently modeled elsewhere. Change the "
        "annual figure, or change utilization above, and the resulting "
        "day-rate updates automatically either way — it's the day-rate "
        "that's derived, not the annual figure. While spot mode is active, "
        "this **replaces** Tab 1's opex for the vessel's P&L (Tab 1's own "
        "figures stay as they are, for the TC-mode scenario)."
    )
    spot_opex_annual_nok = nok_input(
        "Fixed Voyage opex (NOK/year, total)", "spot_opex_annual_nok", 1_500_000.0,
        key="spot_opex_annual_input", disabled=locked
    )
    _opex_esc_col1, _opex_esc_col2 = st.columns(2)
    with _opex_esc_col1:
        spot_opex_escalator_pct = stateful_number_input(
            "Fixed Voyage opex escalator (%/yr)", min_value=-100.0, value=2.0, step=0.5,
            key="spot_opex_escalator", disabled=locked
        )
    with _opex_esc_col2:
        spot_variable_opex_escalator_pct = stateful_number_input(
            "Variable Voyage opex escalator (%/yr)", min_value=-100.0, value=2.0, step=0.5,
            key="spot_variable_opex_escalator", disabled=locked,
            help="Drives the Transport/Other/Treatment build-up tools' own "
                 "voyage costs (fuel, additional opex/hr) — decoupled from "
                 "both the Fixed line above and each segment's own revenue "
                 "escalator on the Service mix table, so cost and revenue "
                 "can grow at different rates."
        )
    spot_opex_rate_nok_day = (
        spot_opex_annual_nok / _working_days_annual_target if _working_days_annual_target else 0.0
    )
    st.metric(
        "Implied day-rate (derived)",
        fmt(spot_opex_rate_nok_day) + "/day",
        help=f"{fmt(spot_opex_annual_nok)}/year ÷ {fmt(_working_days_annual_target)} working days/year."
    )

    st.markdown("**Year 1–12 planning** (utilization by year, and revenue indexation by segment)")
    st.caption(
        "Utilization set individually for each year — no staging concept, "
        "just type the number you expect for that year. Defaults to "
        f"{spot_utilization_pct:.1f}% (matching the Year 1 baseline above "
        "— implied by the Monthly activity calendar, if that's on) for "
        "every year; adjust freely, e.g. ramping up as you add capacity. "
        "Indexation below is one flat rate per segment (not per-year) — "
        "each segment's day-rate compounds at its own rate from Year 2 "
        "onward, same escalation pattern used everywhere else in this "
        "model."
    )

    if "spot_utilization_by_year" not in st.session_state:
        st.session_state.spot_utilization_by_year = [spot_utilization_pct] * 11

    st.markdown("Utilization (%), Year 2–12")
    _util_year_cols = st.columns(11)
    for _yi in range(11):
        with _util_year_cols[_yi]:
            st.session_state.spot_utilization_by_year[_yi] = st.number_input(
                f"Year {_yi + 2}", min_value=0.0, max_value=100.0,
                value=st.session_state.spot_utilization_by_year[_yi], step=1.0,
                key=f"spot_util_year_{_yi}", disabled=locked
            )
    spot_utilization_by_year = st.session_state.spot_utilization_by_year

    st.markdown("**Fixed Voyage opex — today's value, Year 2–12**")
    st.caption(
        "Each year's own real (today's money) Fixed Voyage opex — default "
        "matches Year 1's baseline above. Change a later year's value to "
        "reflect a real cost change (e.g. hiring a second person): typing "
        "2,000,000 for Year 3 means 2,000,000 in today's money, which the "
        "Financial Statements then show as 2,000,000 x (1 + Fixed Voyage "
        "opex escalator)² in Year 3's nominal terms — the same single "
        "escalator set above, just applied to whatever real value is "
        "typed for that specific year, not a new escalation clock."
    )
    if "spot_fixed_opex_real_by_year" not in st.session_state:
        st.session_state.spot_fixed_opex_real_by_year = [spot_opex_annual_nok] * 11

    def _on_fixed_opex_year_change(index):
        raw = st.session_state[f"spot_fixed_opex_year_{index}"]
        value = parse_nok(raw)
        st.session_state.spot_fixed_opex_real_by_year[index] = value
        st.session_state[f"spot_fixed_opex_year_{index}"] = format_nok(value)

    _fixed_opex_year_cols = st.columns(11)
    for _yi in range(11):
        with _fixed_opex_year_cols[_yi]:
            _fixed_opex_display_key = f"spot_fixed_opex_year_{_yi}"
            if _fixed_opex_display_key not in st.session_state:
                st.session_state[_fixed_opex_display_key] = format_nok(st.session_state.spot_fixed_opex_real_by_year[_yi])
            st.text_input(
                f"Year {_yi + 2}", key=_fixed_opex_display_key,
                on_change=_on_fixed_opex_year_change, args=(_yi,), disabled=locked
            )
    spot_fixed_opex_real_by_year = st.session_state.spot_fixed_opex_real_by_year

    st.markdown("**Additional spot capex — Year 2–12** (on top of the NOK 4.3m TC-equivalent maintenance capex)")
    st.caption(
        "Spot trading takes slightly more wear on the vessel than a "
        "steady TC charter — this is the extra capex on top of what the "
        "TC-operation would incur (Tab 1's maintenance capex, applied "
        "unconditionally either way). Punch in a real (today's money) "
        "figure for whichever year it's needed. Added to the vessel's "
        "asset value on the balance sheet and depreciated on its own "
        "schedule (set below), separate from the vessel/maintenance "
        "capex rate — see the Financial Statements tab's Asset register "
        "for the resulting P&L/cash flow/balance sheet lines."
    )
    spot_additional_capex_depreciation_pct = stateful_number_input(
        "Additional spot capex depreciation rate (%/yr)", min_value=0.1, max_value=100.0, value=5.0, step=0.5,
        key="spot_additional_capex_depreciation", disabled=locked,
        help="Default 5%/yr = 20-year useful life. Separate from the "
             "vessel's own rate (Tab 1), since spot-specific capex "
             "additions may have a genuinely different useful life."
    )
    if spot_additional_capex_depreciation_pct > 0:
        st.caption(f"= {100/spot_additional_capex_depreciation_pct:.0f}-year implied useful life.")

    if "spot_additional_capex_by_year" not in st.session_state:
        st.session_state.spot_additional_capex_by_year = [1_000_000.0] * 11

    def _on_additional_capex_year_change(index):
        raw = st.session_state[f"spot_additional_capex_year_{index}"]
        value = parse_nok(raw)
        st.session_state.spot_additional_capex_by_year[index] = value
        st.session_state[f"spot_additional_capex_year_{index}"] = format_nok(value)

    _additional_capex_year_cols = st.columns(11)
    for _yi in range(11):
        with _additional_capex_year_cols[_yi]:
            _additional_capex_display_key = f"spot_additional_capex_year_{_yi}"
            if _additional_capex_display_key not in st.session_state:
                st.session_state[_additional_capex_display_key] = format_nok(st.session_state.spot_additional_capex_by_year[_yi])
            st.text_input(
                f"Year {_yi + 2}", key=_additional_capex_display_key,
                on_change=_on_additional_capex_year_change, args=(_yi,), disabled=locked
            )
    spot_additional_capex_by_year = st.session_state.spot_additional_capex_by_year

    st.markdown("**Revenue indexation by segment — Year 2–12** (each year's own %, compounding)")
    st.caption(
        "Each cell is that specific year's escalator versus the year "
        "before — not a single flat rate — so Year 5's nominal rate "
        "depends on every year's own % from Year 2 through Year 5, "
        "compounding. Defaults to 2% everywhere; change any individual "
        "year freely."
    )

    def _init_segment_escalator_years(state_key):
        if state_key not in st.session_state:
            st.session_state[state_key] = [2.0] * 11

    _init_segment_escalator_years("spot_smolt_escalator_by_year")
    _init_segment_escalator_years("spot_harvest_escalator_by_year")
    _init_segment_escalator_years("spot_treatment_escalator_by_year")

    st.markdown("Transport indexation (%/yr)")
    _smolt_esc_cols = st.columns(11)
    for _yi in range(11):
        with _smolt_esc_cols[_yi]:
            st.session_state.spot_smolt_escalator_by_year[_yi] = st.number_input(
                f"Year {_yi + 2}", min_value=-100.0,
                value=st.session_state.spot_smolt_escalator_by_year[_yi], step=0.5,
                key=f"spot_smolt_esc_year_{_yi}", label_visibility="collapsed", disabled=locked
            )
    spot_smolt_escalator_by_year = st.session_state.spot_smolt_escalator_by_year

    st.markdown("Other indexation (%/yr)")
    _harvest_esc_cols = st.columns(11)
    for _yi in range(11):
        with _harvest_esc_cols[_yi]:
            st.session_state.spot_harvest_escalator_by_year[_yi] = st.number_input(
                f"Year {_yi + 2}", min_value=-100.0,
                value=st.session_state.spot_harvest_escalator_by_year[_yi], step=0.5,
                key=f"spot_harvest_esc_year_{_yi}", label_visibility="collapsed", disabled=locked
            )
    spot_harvest_escalator_by_year = st.session_state.spot_harvest_escalator_by_year

    st.markdown("Treatment indexation (%/yr)")
    _treatment_esc_cols = st.columns(11)
    for _yi in range(11):
        with _treatment_esc_cols[_yi]:
            st.session_state.spot_treatment_escalator_by_year[_yi] = st.number_input(
                f"Year {_yi + 2}", min_value=-100.0,
                value=st.session_state.spot_treatment_escalator_by_year[_yi], step=0.5,
                key=f"spot_treatment_esc_year_{_yi}", label_visibility="collapsed", disabled=locked
            )
    spot_treatment_escalator_by_year = st.session_state.spot_treatment_escalator_by_year

    # --- baseline reference: the COMBINED TC-rate (vessel + leased
    # equipment, e.g. the FLS equipment lease) — not vessel-only. Tab 3
    # (Combined TC-rate) computes this but runs AFTER this tab in script
    # order, so it's read one pass behind via session state (same
    # convention used elsewhere in this app, e.g. Tab 1's Sources & Uses
    # guideline). Falls back to vessel-only on the very first load, before
    # Tab 3 has run even once. ---
    spot_baseline_tc_daily = st.session_state.get("_combined_tc_daily", vessel_tc_daily)

    # This is the required NET income per working day — i.e. after voyage
    # opex, matching the TC-equivalent annual return. Gross price (what's
    # actually billed) = this net target + voyage opex. Utilization is now
    # a direct input above, not derived from a table that renders later —
    # so this is computed once, straightforwardly, with no provisional/
    # stale-versus-final distinction needed anymore.
    required_net_rate_at_utilization = (
        spot_baseline_tc_daily / (spot_utilization_pct / 100) if spot_utilization_pct > 0 else 0.0
    )
    required_gross_rate_at_utilization = required_net_rate_at_utilization + spot_opex_rate_nok_day

    st.markdown("**Baseline reference** (imported live from the Combined TC-rate, Tab 3 — vessel + leased equipment)")
    bl1, bl2, bl3, bl4 = st.columns(4)
    bl1.metric("Baseline TC-rate, 100% utilization (vessel + equipment)", fmt(spot_baseline_tc_daily) + "/day")
    bl2.metric("Utilization", f"{spot_utilization_pct:.1f}%")
    bl3.metric("Required NET rate on working days", fmt(required_net_rate_at_utilization) + "/day")
    bl4.metric("Required GROSS price (net + voyage opex)", fmt(required_gross_rate_at_utilization) + "/day")

    # --- Variable Voyage opex (Transport/Other/Treatment build-up tools)
    # is computed further down the page, so it's read one-pass-behind
    # here, same convention as spot_baseline_tc_daily above. ---
    _spot_variable_voyage_opex_prev_pass = st.session_state.get("_spot_variable_voyage_opex_total", 0.0)
    _required_annual_net = spot_baseline_tc_daily * operating_days
    _required_annual_gross = _required_annual_net + spot_opex_annual_nok + _spot_variable_voyage_opex_prev_pass
    st.markdown("**Required spot income, annual** (TC + Lease benchmark, converted to what spot has to earn)")
    bla_cols = st.columns([2.2, 0.3, 2.0, 0.3, 2.0, 0.3, 2.2])
    with bla_cols[0]:
        st.metric("TC + Lease benchmark (net)", fmt(_required_annual_net))
    with bla_cols[1]:
        st.markdown("<div style='padding-top:32px; font-size:28px; text-align:center; color:#888;'>+</div>", unsafe_allow_html=True)
    with bla_cols[2]:
        st.metric("Fixed Voyage opex", fmt(spot_opex_annual_nok))
    with bla_cols[3]:
        st.markdown("<div style='padding-top:32px; font-size:28px; text-align:center; color:#888;'>+</div>", unsafe_allow_html=True)
    with bla_cols[4]:
        st.metric("Variable Voyage opex", fmt(_spot_variable_voyage_opex_prev_pass))
    with bla_cols[5]:
        st.markdown("<div style='padding-top:32px; font-size:28px; text-align:center; color:#888;'>=</div>", unsafe_allow_html=True)
    with bla_cols[6]:
        st.metric("Required spot income (gross)", fmt(_required_annual_gross))
    st.caption(
        "This is the number Transport + Other + Treatment revenue "
        "together need to hit — shown here directly so there's no need "
        "to toggle spot mode off and check the Combined TC-rate tab to "
        "find it. Same figure the Service mix table's own 'Target' line "
        "further down checks your current service mix against. Variable "
        "Voyage opex (fuel, phase costs, customer changeover — from the "
        "Transport/Other/Treatment build-up tools further down) is "
        "computed later on the page, so this figure is one script pass "
        "behind after changing build-up tool inputs — it catches up "
        "automatically; switch tabs or click Refresh in the sidebar if "
        "it looks stale."
    )

    st.caption(
        f"= {fmt(vessel_tc_daily)}/day vessel (Tab 1) + "
        f"{fmt(spot_baseline_tc_daily - vessel_tc_daily)}/day leased equipment "
        f"(Tab 2/3, e.g. FLS) — this is a **100%-utilization** rate (spread "
        f"across every operating day, not just working days), matching how "
        f"a TC charter is priced. Required NET rate = baseline ÷ "
        f"utilization — the net income a job type needs to earn, on the "
        f"days it's actually working, to match the same TC-equivalent "
        f"annual revenue at the utilization set above. Required GROSS "
        f"price = that net target + voyage opex "
        f"({fmt(spot_opex_rate_nok_day)}/day) — this is the day-rate "
        f"actually billed. Services flagged 'priced at baseline' below are "
        f"simply assumed to bill at this gross price exactly (no "
        f"price-list build-up on top) — a simplifying assumption for "
        f"segments that aren't the focus of the pricing work. Baseline "
        f"TC-rate above is one script pass behind Tab 3 — switch tabs once "
        f"after editing lease inputs for it to catch up."
    )

    st.divider()
    st.markdown("**Advanced pricing (optional)** — seasonality & customer mix")
    st.caption(
        "Off by default — with both switched off, each service below "
        "just uses its single flat day-rate exactly as before. Start "
        "with **Seasonality** (the Monthly activity calendar — which "
        "one activity the vessel runs each month), then optionally "
        "layer **Customer mix** on top for any service, blending "
        "several customers/contracts into one effective rate. Both "
        "feed straight into the Rate / Annual revenue columns below "
        "and into the P&L — nothing else in the model needs to change."
    )

    _service_slug_lookup = {
        "Treatment of fish": "treatment",
        "Transport": "smolt",
        "Other": "harvest",
    }

    # One-time migration: anything already stored in this session under
    # the old "Smolt transport"/"Harvest transport" names (from before
    # the rename to generic "Transport"/"Other") is renamed in place,
    # so the live calendar, Service mix table, and any already-booked
    # customer plans aren't silently orphaned by the change. Slugs
    # ("smolt"/"harvest") are untouched throughout — only the
    # user-facing name changes, so customer mix data (keyed by slug)
    # needs no migration at all.
    _service_name_migration = {"Smolt transport": "Transport", "Harvest transport": "Other"}
    for _item in st.session_state.get("spot_service_items", []):
        if _item.get("name") in _service_name_migration:
            _item["name"] = _service_name_migration[_item["name"]]
    for _row in st.session_state.get("spot_calendar_weeks", []):
        if _row.get("activity") in _service_name_migration:
            _row["activity"] = _service_name_migration[_row["activity"]]
    for _row in st.session_state.get("spot_calendar_months", []):
        if _row.get("activity") in _service_name_migration:
            _row["activity"] = _service_name_migration[_row["activity"]]

    def _customer_mix_enabled(service_name):
        slug = _service_slug_lookup.get(service_name)
        return bool(slug) and st.session_state.get(f"spot_customer_mix_enabled_{slug}", False)

    def _calendar_enabled():
        return st.session_state.get("spot_calendar_enabled", False)

    def _calendar_annual_days(service_name):
        """This service's annual days as implied by the Weekly activity
        calendar (sum of each week's days where that week's activity
        matches this service) — 0.0 if the calendar is off or no week
        is currently assigned to this service."""
        if not _calendar_enabled():
            return 0.0
        calendar_weeks = st.session_state.get("spot_calendar_weeks", [])
        days_per_week_equiv = operating_days / 52
        return sum(
            days_per_week_equiv * (row["occupancy_pct"] / 100)
            for row in calendar_weeks if row.get("activity") == service_name
        )

    def _week_to_month_map():
        """Inverse of _month_to_weeks_map() — {week_number: month_index}."""
        result = {}
        for month_idx, weeks in _month_to_weeks_map().items():
            for w in weeks:
                result[w] = month_idx
        return result

    def _eligible_weeks_for_service(service_name):
        """Week numbers available for this service's customer mix — only
        weeks the Weekly activity calendar currently assigns to this
        service, plus any Idle weeks (a potential fill-in opportunity)
        — never weeks assigned to one of the other services. Falls back
        to all 52 weeks if the calendar is off (nothing to constrain by
        yet)."""
        if not _calendar_enabled():
            return list(range(1, 53))
        calendar_weeks = st.session_state.get("spot_calendar_weeks", [])
        return sorted(
            w + 1 for w, row in enumerate(calendar_weeks)
            if row.get("activity") in (service_name, "Idle")
        )

    def _committed_and_idle_weeks_for_service(service_name):
        """Splits _eligible_weeks_for_service(...) into two lists:
        (committed_weeks, idle_weeks) — committed = weeks the calendar
        actually assigns to this service (real, scheduled work); idle =
        weeks the calendar marks Idle (optional fill-in / prospect
        opportunity only). If the calendar is off, everything is
        treated as committed (nothing to distinguish by yet)."""
        if not _calendar_enabled():
            return list(range(1, 53)), []
        calendar_weeks = st.session_state.get("spot_calendar_weeks", [])
        committed = [w + 1 for w, row in enumerate(calendar_weeks) if row.get("activity") == service_name]
        idle = [w + 1 for w, row in enumerate(calendar_weeks) if row.get("activity") == "Idle"]
        return sorted(committed), sorted(idle)

    def _customer_mix_blended_rate(service_name, year=1):
        """Weighted-average Year-1 day-rate across this service's
        customer mix — blended by each customer's share of this
        service's COMMITTED weeks only (i.e. weeks the Monthly activity
        calendar actually schedules for this service). Any weeks ticked
        in an Idle month are prospects only and are never counted here,
        so they can't influence pricing or the plan. No per-customer
        escalation is applied here — Year 2-12 growth comes entirely
        from the aggregate segment escalator (the Year 1-12 planning
        table above), applied on top of this blend by
        _service_day_rate_for_year. The `year` parameter is accepted
        for call-site compatibility but has no effect. Falls back to 0
        if the service isn't one of the three build-up-tool-backed
        services, or has no committed weeks assigned to any customer
        yet."""
        slug = _service_slug_lookup.get(service_name)
        if slug is None:
            return 0.0
        customers = st.session_state.get(f"spot_customers_{slug}", [])
        if not customers:
            return 0.0
        committed_weeks_all, _ = _committed_and_idle_weeks_for_service(service_name)
        committed_set = set(committed_weeks_all)
        committed_counts = {
            id(c): len([w for w in c.get("weeks", []) if w in committed_set]) for c in customers
        }
        total_committed_assigned = sum(committed_counts.values())
        if total_committed_assigned == 0:
            return 0.0
        blended = 0.0
        for c in customers:
            n = committed_counts.get(id(c), 0)
            if n == 0:
                continue
            blended += c["rate_nok_day"] * (n / total_committed_assigned)
        return blended

    def _customer_booking_blocks(service_name):
        """List of (customer_name, week_count) tuples — one entry per
        contiguous booking block of COMMITTED weeks (idle/prospect ticks
        are ignored entirely), ordered chronologically by week number.
        'Contiguous' means consecutive OWNED weeks in week order — an
        unassigned committed week in between doesn't break a block
        (nothing forces a changeover if no one else is using that slot),
        but a different customer's booking in between does, even if the
        original customer returns later (e.g. the same real farmer
        booked again later in the year under a different roster row, or
        even under the same row non-contiguously, both correctly need
        their own changeover on return). Returns [] if customer mix
        isn't enabled for this service, or nothing's booked yet."""
        if not _customer_mix_enabled(service_name):
            return []
        slug = _service_slug_lookup.get(service_name)
        customers = st.session_state.get(f"spot_customers_{slug}", [])
        committed_weeks_all, _ = _committed_and_idle_weeks_for_service(service_name)
        committed_set = set(committed_weeks_all)
        week_owner = {}
        for c in customers:
            for w in c.get("weeks", []):
                if w in committed_set:
                    week_owner[w] = c["name"]
        blocks = []
        for w in sorted(week_owner.keys()):
            owner = week_owner[w]
            if blocks and blocks[-1][0] == owner:
                blocks[-1] = (owner, blocks[-1][1] + 1)
            else:
                blocks.append((owner, 1))
        return blocks

    def _resolved_charged_rate(service_name):
        """The actual Year-1 day-rate in effect for a service, resolving
        priced-at-baseline / customer-mix in the same priority order as
        the Rate column on the Service mix table below — used by each
        service's own Net income check further down the page, so those
        stay in sync rather than reading the raw manual rate_nok_day
        field regardless of these overrides."""
        item = next((it for it in st.session_state.spot_service_items if it["name"] == service_name), None)
        if item is None:
            return 0.0
        if item.get("priced_at_baseline", False):
            return required_gross_rate_at_utilization
        if _customer_mix_enabled(service_name):
            return _customer_mix_blended_rate(service_name, 1)
        return item["rate_nok_day"]

    st.markdown("**1. Seasonality — Weekly activity calendar**")
    st.caption(
        "One vessel, one activity at a time — the vessel can't run "
        "Treatment, Transport, and Other simultaneously, so this is a "
        "week-by-week plan: which one activity (or none — idle) runs "
        "each week, and how full that week is (**Occupancy** — usually "
        "100% in an active week, 0% in a fully idle one). Rounds and "
        "days for that week are then derived automatically from the "
        "activity's own days/round (its build-up tool further down the "
        "page). A month can now mix activities across its own weeks — "
        "e.g. two weeks of Treatment and one idle week in the same "
        "month — everything downstream (Days/year, revenue, voyage "
        "cost) is driven off these 52 weeks directly, not the month "
        "they happen to fall in. Off by default — with it off, each "
        "service's Share of working days (%) on the table below still "
        "works exactly as before, spread flat across the year."
    )

    calendar_enabled = stateful_toggle(
        "Enable weekly activity calendar?", value=True,
        key="spot_calendar_enabled", disabled=locked
    )
    _calendar_days_per_service = {}
    if calendar_enabled:
        _cal_month_labels = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
        _activity_options = ["Treatment of fish", "Transport", "Other", "Idle"]
        _week_to_month = _week_to_month_map()

        if "spot_calendar_weeks" not in st.session_state:
            # Migrate any existing month-level plan (from before this
            # became week-native) into 52 weekly rows, so nothing
            # already configured is lost — each week just inherits its
            # owning month's activity/occupancy as a starting point,
            # fully editable per-week from here on. If there's no prior
            # month-level plan either, fall back directly to
            # _default_calendar_weeks (defined earlier on this page,
            # also used there as the first-load implied-utilization
            # fallback).
            _prior_months = st.session_state.get("spot_calendar_months")
            if _prior_months:
                _month_weeks_map_init = _month_to_weeks_map()
                _source_activities = [row.get("activity", "Idle") for row in _prior_months]
                _source_occupancy = [row.get("occupancy_pct", 0.0) for row in _prior_months]
                _weeks_init = [{"activity": "Idle", "occupancy_pct": 0.0} for _ in range(52)]
                for _mi in range(12):
                    for _w in _month_weeks_map_init.get(_mi, []):
                        if _mi < len(_source_activities):
                            _weeks_init[_w - 1] = {
                                "activity": _source_activities[_mi],
                                "occupancy_pct": _source_occupancy[_mi],
                            }
            else:
                _weeks_init = [
                    {"activity": act, "occupancy_pct": occ} for act, occ in _default_calendar_weeks
                ]
            st.session_state.spot_calendar_weeks = _weeks_init

        _service_round_hours = st.session_state.get("_service_round_hours", {})
        _days_per_week_equiv = operating_days / 52

        cal_hdr = st.columns([0.9, 0.9, 1.6, 1.1, 1.1, 1.1])
        cal_hdr[0].markdown("**Week**")
        cal_hdr[1].markdown("**Month**")
        cal_hdr[2].markdown("**Activity**")
        cal_hdr[3].markdown("**Occupancy (%)**")
        cal_hdr[4].markdown("**Rounds this week**")
        cal_hdr[5].markdown("**Days this week**")

        for wi in range(52):
            _week_num = wi + 1
            _row = st.session_state.spot_calendar_weeks[wi]
            c1, c2, c3, c4, c5, c6 = st.columns([0.9, 0.9, 1.6, 1.1, 1.1, 1.1])
            with c1:
                st.markdown(f"<div style='padding-top:8px'>W{_week_num}</div>", unsafe_allow_html=True)
            with c2:
                _month_idx = _week_to_month.get(_week_num)
                _month_label = _cal_month_labels[_month_idx] if _month_idx is not None else "—"
                st.markdown(f"<div style='padding-top:8px'>{_month_label}</div>", unsafe_allow_html=True)
            with c3:
                _row["activity"] = st.selectbox(
                    "Activity", _activity_options,
                    index=_activity_options.index(_row["activity"]) if _row["activity"] in _activity_options else 3,
                    key=f"spot_calendar_activity_{wi}", label_visibility="collapsed", disabled=locked
                )
            with c4:
                _row["occupancy_pct"] = st.number_input(
                    "Occupancy (%)", min_value=0.0, max_value=100.0, value=_row["occupancy_pct"],
                    step=5.0, key=f"spot_calendar_occ_{wi}", label_visibility="collapsed", disabled=locked
                )
            _days_this_week = _days_per_week_equiv * (_row["occupancy_pct"] / 100)
            _hours_per_round_this = _service_round_hours.get(_row["activity"])
            with c5:
                if _row["activity"] != "Idle" and _hours_per_round_this:
                    _rounds_this_week = (_days_this_week * 24) / _hours_per_round_this
                    st.markdown(f"<div style='padding-top:8px'>{_rounds_this_week:.2f}</div>", unsafe_allow_html=True)
                else:
                    st.markdown("<div style='padding-top:8px'>—</div>", unsafe_allow_html=True)
            with c6:
                if _row["activity"] != "Idle":
                    st.markdown(f"<div style='padding-top:8px'>{_days_this_week:.1f}</div>", unsafe_allow_html=True)
                    _calendar_days_per_service[_row["activity"]] = _calendar_days_per_service.get(_row["activity"], 0.0) + _days_this_week
                else:
                    st.markdown("<div style='padding-top:8px'>—</div>", unsafe_allow_html=True)

        st.caption(
            "Rounds this week needs each activity's days/round from its "
            "own build-up tool further down the page — shows '—' until "
            "that's run at least once this session (self-corrects "
            "automatically right after, no action needed)."
        )

        _cal_total_days = sum(_calendar_days_per_service.values())
        _cal_summary_rows = [
            {
                "Service": name,
                "Annual days (from calendar)": fmt(days),
                "Share of total (%)": f"{(days / _cal_total_days * 100):.1f}%" if _cal_total_days else "0.0%",
            }
            for name, days in _calendar_days_per_service.items()
        ]
        _cal_implied_utilization = (_cal_total_days / operating_days * 100) if operating_days else 0.0
        if _cal_summary_rows:
            _cal_summary_rows.append({"Service": "Total", "Annual days (from calendar)": fmt(_cal_total_days), "Share of total (%)": "100.0%"})
            st.dataframe(pd.DataFrame(_cal_summary_rows).set_index("Service"), width="stretch")
        st.metric("Implied utilization", f"{_cal_implied_utilization:.1f}%", help=f"{fmt(_cal_total_days)} days ÷ {fmt(operating_days)} operating days (Tab 1).")
        st.caption(
            "These annual totals now drive Days/year for the matching "
            "services on the table below (overriding Share of working "
            "days (%) for those three) — everything downstream (Rounds/"
            "year, Annual revenue, each build-up tool, Net income "
            "checks) follows from here automatically. Implied utilization "
            "above is this calendar's effective annual utilization — the "
            "'Utilization (%)' figure near the top of this page now just "
            "displays this number rather than being set independently."
        )

        # --- self-healing: the "Utilization (implied...)" display near
        # the top of this page runs BEFORE this calendar (one script pass
        # behind, same convention used throughout this app — see
        # _service_round_hours above). Store the freshly computed figure
        # and, if it changed from what that earlier display just showed,
        # trigger one more pass so it catches up immediately. ---
        _calendar_util_changed = (
            abs(st.session_state.get("_calendar_implied_utilization_pct", -1.0) - _cal_implied_utilization) > 0.05
        )
        st.session_state["_calendar_implied_utilization_pct"] = _cal_implied_utilization
        _cal_util_retry_count = st.session_state.get("_cal_util_retry_count", 0)
        if _calendar_util_changed and _cal_util_retry_count < 4:
            st.session_state["_cal_util_retry_count"] = _cal_util_retry_count + 1
            _request_rerun()
        else:
            st.session_state["_cal_util_retry_count"] = 0

    st.divider()
    st.markdown("**2. Customer mix**")

    _open_grid_services = [
        _svc for _svc, _slg in _service_slug_lookup.items()
        if st.session_state.get(f"spot_customer_mix_show_grid_{_slg}", False)
    ]
    if _open_grid_services:
        st.warning(
            f"⚠️ **Week-by-week grid is open for: {', '.join(_open_grid_services)}.** "
            "Each open grid renders up to 52 x N-customer checkboxes on "
            "every single interaction anywhere in the app, not just this "
            "tab — this is the single biggest known driver of a slow "
            "reload. Close it below (or with the button) unless you're "
            "actively editing week assignments right now."
        )

    def _close_all_grids():
        for _slug in _service_slug_lookup.values():
            st.session_state[f"spot_customer_mix_show_grid_{_slug}"] = False

    def _clear_all_customer_bookings():
        for _slug in _service_slug_lookup.values():
            _key = f"spot_customers_{_slug}"
            for _cust in st.session_state.get(_key, []):
                _cust["weeks"] = []
            # Also reset every possible week checkbox for every customer
            # row currently defined for this service — same reasoning as
            # the per-customer Clear button: the marker-based sync only
            # resyncs a checkbox that was blocked since last render, so
            # without this, boxes that were already active would keep
            # showing as ticked despite the underlying data now being
            # empty.
            for _ci in range(len(st.session_state.get(_key, []))):
                for _w in range(1, 53):
                    st.session_state[f"spot_cust_week_{_slug}_{_ci}_{_w}"] = False

    _cmbtn1, _cmbtn2 = st.columns(2)
    with _cmbtn1:
        if st.button("🧹 Clear all customer bookings (Treatment, Transport, Other)", disabled=locked):
            _clear_all_customer_bookings()
            _request_rerun()
        st.caption(
            "Wipes every customer's ticked weeks across all three services "
            "— customer names, rates, and the weekly calendar itself are "
            "untouched, only the week assignments are cleared."
        )
    with _cmbtn2:
        if st.button("🔒 Close all week-by-week grids", disabled=locked or not _open_grid_services):
            _close_all_grids()
            _request_rerun()
        st.caption(
            "Forces all three grids shut in one click, without needing "
            "to open each service's expander and find its own toggle."
        )

    st.markdown(
        """
        <style>
        /* Ticked weeks in the customer-mix calendars turn green instead
        of the app's default accent color. Checkboxes are only used for
        this one feature, so this is scoped safely. Several selector
        variants are stacked here (tag-agnostic where possible) since
        Streamlit's internal checkbox markup differs slightly by
        version — this maximizes the chance at least one matches. */
        div[data-testid="stCheckbox"] [aria-checked="true"] {
            background-color: #16a34a !important;
            border-color: #16a34a !important;
        }
        div[data-testid="stCheckbox"] [aria-checked="true"] svg {
            fill: #ffffff !important;
        }
        div[data-testid="stCheckbox"] [data-baseweb="checkbox"] [aria-checked="true"],
        div[data-testid="stCheckbox"] label > div:first-child:has([aria-checked="true"]) {
            background-color: #16a34a !important;
            border-color: #16a34a !important;
        }
        div[data-testid="stCheckbox"] input[type="checkbox"]:checked + div,
        div[data-testid="stCheckbox"] input[type="checkbox"]:checked ~ div:first-of-type {
            background-color: #16a34a !important;
            border-color: #16a34a !important;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )
    for _service_name, _slug in _service_slug_lookup.items():
        with st.expander(f"{_service_name} — customer mix", expanded=False):
            _cm_enabled = stateful_toggle(
                f"Enable customer mix for {_service_name}?", value=True,
                key=f"spot_customer_mix_enabled_{_slug}", disabled=locked
            )
            if _cm_enabled:
                _eligible_weeks = _eligible_weeks_for_service(_service_name)
                st.caption(
                    "Several customers/contracts blended into one "
                    "effective Year-1 rate for this service — each "
                    "assigned specific week numbers, at its own day-rate. "
                    "Year 2-12 growth comes entirely from the aggregate "
                    "segment escalator (the Year 1-12 planning table "
                    "above), applied on top of this blend — no separate "
                    "per-customer escalator. Every customer's calendar "
                    "below shows **all 52 weeks** for context: weeks "
                    "already running Transport or Other are greyed out and "
                    "read-only; open weeks (this service's own committed "
                    "months, plus Idle months as fill-in prospects) start "
                    "blank and turn green when ticked. Only **committed** "
                    "weeks (this service's own months) drive the blended "
                    "rate and Share (%) — Idle ticks are prospects only "
                    "and never affect pricing or days/year. Replaces the "
                    "single 'Rate (NOK/day)' field for this service in "
                    "the table below."
                )
                if _calendar_enabled() and not _eligible_weeks:
                    st.warning(
                        "⚠️ No months are currently assigned to this "
                        "service (or Idle) on the Monthly activity "
                        "calendar above — every week below will show as "
                        "blocked."
                    )
                _customers_key = f"spot_customers_{_slug}"
                if _customers_key not in st.session_state:
                    if _service_name == "Treatment of fish":
                        st.session_state[_customers_key] = [
                            {"name": "Customer A", "weeks": [1, 2], "share_pct": 0.0, "rate_nok_day": 720_000.0},
                            {"name": "Customer B", "weeks": [14, 15, 16], "share_pct": 0.0, "rate_nok_day": 720_000.0},
                            {"name": "Customer C", "weeks": [17, 18, 19], "share_pct": 0.0, "rate_nok_day": 720_000.0},
                            {"name": "Customer D", "weeks": [20, 21, 22], "share_pct": 0.0, "rate_nok_day": 720_000.0},
                            {"name": "Customer E", "weeks": [23, 24], "share_pct": 0.0, "rate_nok_day": 720_000.0},
                            {"name": "Customer F", "weeks": [25, 26], "share_pct": 0.0, "rate_nok_day": 720_000.0},
                            {"name": "Customer G", "weeks": [32, 33], "share_pct": 0.0, "rate_nok_day": 720_000.0},
                            {"name": "Customer H", "weeks": [34, 35], "share_pct": 0.0, "rate_nok_day": 720_000.0},
                            {"name": "Customer I", "weeks": [40, 41, 42], "share_pct": 0.0, "rate_nok_day": 720_000.0},
                            {"name": "Customer J", "weeks": [43, 44, 45], "share_pct": 0.0, "rate_nok_day": 720_000.0},
                            {"name": "Customer K", "weeks": [46, 47, 48], "share_pct": 0.0, "rate_nok_day": 720_000.0},
                            {"name": "Customer L", "weeks": [49, 50, 51], "share_pct": 0.0, "rate_nok_day": 720_000.0},
                        ]
                    elif _service_name == "Transport":
                        st.session_state[_customers_key] = [
                            {"name": "Customer A", "weeks": [10, 11, 12, 13], "share_pct": 0.0, "rate_nok_day": 456_000.0},
                            {"name": "Customer B", "weeks": [36, 37, 38, 39], "share_pct": 0.0, "rate_nok_day": 456_000.0},
                        ]
                    elif _service_name == "Other":
                        st.session_state[_customers_key] = [
                            {"name": "Customer A", "weeks": [40, 41, 42, 43], "share_pct": 0.0, "rate_nok_day": 456_000.0},
                            {"name": "Customer B", "weeks": [44, 45, 46, 47], "share_pct": 0.0, "rate_nok_day": 456_000.0},
                        ]
                    else:
                        st.session_state[_customers_key] = [
                            {"name": "Customer A", "weeks": [], "share_pct": 0.0, "rate_nok_day": 480_000.0},
                            {"name": "Customer B", "weeks": [], "share_pct": 0.0, "rate_nok_day": 420_000.0},
                        ]

                def _next_customer_letter(n):
                    """A, B, ..., Z, AA, AB, ... for the nth (0-indexed)
                    customer about to be added — so the default name
                    always continues the existing A, B, C... sequence
                    regardless of how many customers are already defined."""
                    n += 1
                    letters = ""
                    while n > 0:
                        n, rem = divmod(n - 1, 26)
                        letters = chr(65 + rem) + letters
                    return letters

                def _add_customer(slug=_slug, service_name=_service_name):
                    _existing = st.session_state[f"spot_customers_{slug}"]
                    _new_name = f"Customer {_next_customer_letter(len(_existing))}"
                    _default_rate = next(
                        (it["rate_nok_day"] for it in st.session_state.spot_service_items if it["name"] == service_name),
                        0.0,
                    )
                    _existing.append(
                        {"name": _new_name, "weeks": [], "share_pct": 0.0, "rate_nok_day": _default_rate}
                    )

                def _remove_customer(index, slug=_slug):
                    st.session_state[f"spot_customers_{slug}"].pop(index)

                def _clear_customer_weeks(index, slug=_slug):
                    st.session_state[f"spot_customers_{slug}"][index]["weeks"] = []
                    # Directly force every possible week checkbox for
                    # this customer to False — the marker-based sync in
                    # the render loop only resyncs a checkbox that was
                    # blocked since last render, so without this, boxes
                    # that were already active would keep showing as
                    # ticked despite cust["weeks"] now being empty.
                    for _w in range(1, 53):
                        st.session_state[f"spot_cust_week_{slug}_{index}_{_w}"] = False

                def _on_customer_rate_change(slug, index):
                    raw = st.session_state[f"spot_cust_rate_{slug}_{index}"]
                    value = parse_nok(raw)
                    st.session_state[f"spot_customers_{slug}"][index]["rate_nok_day"] = value
                    st.session_state[f"spot_cust_rate_{slug}_{index}"] = format_nok(value)

                # Self-clean: drop any previously-assigned week that's no
                # longer eligible (e.g. the calendar changed since this
                # customer's weeks were picked).
                for cust in st.session_state[_customers_key]:
                    cust["weeks"] = [w for w in cust.get("weeks", []) if w in _eligible_weeks]

                _calendar_weeks_now = st.session_state.get("spot_calendar_weeks", [])
                # The two OTHER services (relative to whichever expander
                # this is) each get their own grey shade — lighter for
                # whichever comes first in the fixed Treatment/Transport/
                # Other ordering, darker for the other — so e.g. in
                # Treatment's panel, Transport-blocked weeks read lighter
                # grey and Other-blocked weeks read darker grey.
                _other_services = [s for s in _service_slug_lookup if s != _service_name]
                _grey_shades = {
                    _other_services[0]: ("#e5e7eb", "#6b7280"),
                    _other_services[1]: ("#9ca3af", "#f3f4f6"),
                }

                cmhdr = st.columns([2.0, 1.5, 0.5, 0.4])
                cmhdr[0].markdown("**Customer**")
                cmhdr[1].markdown("**Rate (NOK/day), Yr 1**")

                _show_grid = stateful_toggle(
                    f"Show week-by-week grid for {_service_name}?", value=False,
                    key=f"spot_customer_mix_show_grid_{_slug}", disabled=locked
                )
                if not _show_grid:
                    st.caption(
                        "Grid hidden by default — with several customers "
                        "x 52 weeks each, rendering all of them at once "
                        "is the single biggest driver of a slow reload "
                        "(unlike a collapsed expander, this toggle "
                        "actually skips building the grid, not just "
                        "hides it). Turn it on when you need to tick/"
                        "untick weeks; committed/idle counts and the "
                        "blended rate below already reflect whatever's "
                        "saved either way."
                    )

                for ci, cust in enumerate(st.session_state[_customers_key]):
                    top1, top2, top3, top4 = st.columns([2.0, 1.5, 0.5, 0.4])
                    with top1:
                        cust["name"] = st.text_input(
                            "Customer", value=cust["name"], key=f"spot_cust_name_{_slug}_{ci}",
                            label_visibility="collapsed", disabled=locked
                        )
                    with top2:
                        _rate_display_key = f"spot_cust_rate_{_slug}_{ci}"
                        if _rate_display_key not in st.session_state:
                            st.session_state[_rate_display_key] = format_nok(cust["rate_nok_day"])
                        st.text_input(
                            "Rate (NOK/day)", key=_rate_display_key, label_visibility="collapsed",
                            on_change=_on_customer_rate_change, args=(_slug, ci), disabled=locked
                        )
                    with top3:
                        st.button("Clear", key=f"spot_cust_clear_{_slug}_{ci}", on_click=_clear_customer_weeks, args=(ci,), disabled=locked)
                    with top4:
                        st.button("✕", key=f"spot_cust_remove_{_slug}_{ci}", on_click=_remove_customer, args=(ci,), disabled=locked)

                    if not _show_grid:
                        # Cheap path: committed/idle counts only, no
                        # 52-week loop, no checkbox widgets created at
                        # all — this is what actually saves render time,
                        # not the expander's collapsed state.
                        _committed_weeks_all, _idle_weeks_all = _committed_and_idle_weeks_for_service(_service_name)
                        _committed_set = set(_committed_weeks_all)
                        _idle_set = set(_idle_weeks_all)
                        _current_weeks = set(cust.get("weeks", []))
                        _n_committed_ticked = len(_current_weeks & _committed_set)
                        _n_idle_ticked = len(_current_weeks & _idle_set)
                        st.caption(
                            f"{cust['name']}: {_n_committed_ticked} of "
                            f"{len(_committed_weeks_all)} committed weeks · "
                            f"{_n_idle_ticked} of {len(_idle_weeks_all)} idle "
                            f"(prospect) weeks ticked — turn on the grid "
                            f"above to edit."
                        )
                        continue

                    # Weeks already claimed by OTHER customers for this
                    # service are read-only here too (no double-booking),
                    # labeled with whichever customer holds them.
                    _taken_by_others = {}
                    for other_ci, other in enumerate(st.session_state[_customers_key]):
                        if other_ci != ci:
                            for w in other.get("weeks", []):
                                _taken_by_others[w] = other["name"]

                    _committed_weeks_all, _idle_weeks_all = _committed_and_idle_weeks_for_service(_service_name)
                    _committed_set = set(_committed_weeks_all)
                    _idle_set = set(_idle_weeks_all)
                    _current_weeks = set(cust["weeks"])
                    _n_committed_ticked = len(_current_weeks & _committed_set)
                    _n_idle_ticked = len(_current_weeks & _idle_set)
                    st.caption(
                        f"{cust['name']}: {_n_committed_ticked} of "
                        f"{len(_committed_weeks_all)} committed weeks · "
                        f"{_n_idle_ticked} of {len(_idle_weeks_all)} idle "
                        f"(prospect) weeks ticked"
                    )

                    _weeks_per_row = 10
                    for _row_start in range(1, 53, _weeks_per_row):
                        _row_weeks = list(range(_row_start, min(_row_start + _weeks_per_row, 53)))
                        _week_cols = st.columns(_weeks_per_row)
                        for _col, w in zip(_week_cols, _row_weeks):
                            with _col:
                                _cb_key = f"spot_cust_week_{_slug}_{ci}_{w}"
                                _active_marker_key = f"{_cb_key}_was_active"
                                activity = (
                                    _calendar_weeks_now[w - 1]["activity"]
                                    if w - 1 < len(_calendar_weeks_now) else "Idle"
                                )
                                if activity in _grey_shades:
                                    st.session_state[_active_marker_key] = False
                                    _bg, _fg = _grey_shades[activity]
                                    _short = activity.split()[0]
                                    st.markdown(
                                        f"<div style='background:{_bg};color:{_fg};text-align:center;"
                                        f"border-radius:4px;padding:6px 2px;font-size:11px;margin-top:2px;'>"
                                        f"W{w}<br>{_short}</div>",
                                        unsafe_allow_html=True,
                                    )
                                elif w in _taken_by_others:
                                    st.session_state[_active_marker_key] = False
                                    st.markdown(
                                        f"<div style='background:#d1d5db;color:#374151;text-align:center;"
                                        f"border-radius:4px;padding:6px 2px;font-size:11px;margin-top:2px;' "
                                        f"title='Taken by {_taken_by_others[w]}'>W{w}<br>Taken</div>",
                                        unsafe_allow_html=True,
                                    )
                                else:
                                    _should_be_checked = w in _current_weeks
                                    # Only force the widget's stored state
                                    # to match cust["weeks"] if it wasn't
                                    # an active, clickable checkbox on the
                                    # PREVIOUS render (i.e. it just became
                                    # eligible again after being blocked,
                                    # or this is genuinely the first time)
                                    # — never when it was already active,
                                    # since that would silently overwrite
                                    # a click the user just made this very
                                    # render (cust["weeks"] only reflects
                                    # last render's state, not this one's
                                    # in-progress click).
                                    if not st.session_state.get(_active_marker_key, False):
                                        st.session_state[_cb_key] = _should_be_checked
                                    st.session_state[_active_marker_key] = True
                                    _checked = st.checkbox(f"W{w}", key=_cb_key, disabled=locked)
                                    if _checked:
                                        _current_weeks.add(w)
                                    else:
                                        _current_weeks.discard(w)

                    cust["weeks"] = sorted(_current_weeks)
                    st.divider()

                st.button("+ Add customer", key=f"spot_cust_add_{_slug}", on_click=_add_customer, disabled=locked)

                # Share (%) — and everything the blended rate weights by
                # — is based on COMMITTED weeks only (weeks the calendar
                # actually schedules for this service). Idle/prospect
                # ticks are tracked separately below for visibility, but
                # never feed into the share, the blend, or the plan.
                _committed_weeks_all, _idle_weeks_all = _committed_and_idle_weeks_for_service(_service_name)
                _committed_set = set(_committed_weeks_all)
                _idle_set = set(_idle_weeks_all)
                _total_committed_assigned = sum(len([w for w in c["weeks"] if w in _committed_set]) for c in st.session_state[_customers_key])
                _total_idle_assigned = sum(len([w for w in c["weeks"] if w in _idle_set]) for c in st.session_state[_customers_key])
                for cust in st.session_state[_customers_key]:
                    _n_committed = len([w for w in cust["weeks"] if w in _committed_set])
                    cust["share_pct"] = (_n_committed / _total_committed_assigned * 100) if _total_committed_assigned else 0.0

                _all_assigned_weeks = {w for c in st.session_state[_customers_key] for w in c["weeks"]}
                _unassigned_committed = [w for w in _committed_weeks_all if w not in _all_assigned_weeks]
                su1, su2 = st.columns(2)
                su1.metric("Committed weeks assigned", f"{_total_committed_assigned} of {len(_committed_weeks_all)}")
                su2.metric("Unassigned committed weeks", f"{len(_unassigned_committed)}")
                su3, su4 = st.columns(2)
                su3.metric("Idle (prospect) weeks assigned", f"{_total_idle_assigned} of {len(_idle_weeks_all)}")
                su4.metric("Idle weeks still open", f"{len(_idle_weeks_all) - _total_idle_assigned}")
                st.caption(
                    "Only committed weeks drive the blended rate and "
                    "Share (%) below — idle/prospect weeks are shown for "
                    "planning only and never affect pricing or days/year."
                )

                _days_per_week_equiv = operating_days / 52
                _service_round_hours_now = st.session_state.get("_service_round_hours", {})
                _hours_per_round_this_service = _service_round_hours_now.get(_service_name)
                _days_per_round_this_service = (
                    _hours_per_round_this_service / 24 if _hours_per_round_this_service else None
                )

                def _customer_committed_days(c):
                    return sum(
                        _days_per_week_equiv * (_calendar_weeks_now[w - 1]["occupancy_pct"] / 100)
                        for w in c["weeks"]
                        if w in _committed_set and w - 1 < len(_calendar_weeks_now)
                    )

                _cm_summary_rows = []
                for c in st.session_state[_customers_key]:
                    _committed_days_this = _customer_committed_days(c)
                    _amount_paid_this = c["rate_nok_day"] * _committed_days_this
                    if _days_per_round_this_service:
                        _rounds_this = _committed_days_this / _days_per_round_this_service
                        _price_per_round_this = c["rate_nok_day"] * _days_per_round_this_service
                    else:
                        _rounds_this = None
                        _price_per_round_this = None
                    _cm_summary_rows.append({
                        "Customer": c["name"],
                        "Committed weeks": len([w for w in c["weeks"] if w in _committed_set]),
                        "Idle (prospect) weeks": len([w for w in c["weeks"] if w in _idle_set]),
                        "Share (%)": f"{c['share_pct']:.1f}%",
                        "Rounds": f"{_rounds_this:.2f}" if _rounds_this is not None else "—",
                        "Price/round (NOK)": fmt(_price_per_round_this) if _price_per_round_this is not None else "—",
                        "Amount paid, Year 1 (NOK)": fmt(_amount_paid_this),
                    })
                st.dataframe(pd.DataFrame(_cm_summary_rows).set_index("Customer"), width="stretch")
                st.caption(
                    "Amount paid = each committed week's own occupancy-"
                    "weighted days (operating days ÷ 52 x that week's "
                    "occupancy %) x this customer's rate — the same "
                    "day-count logic the P&L uses, so this ties out to "
                    "what actually lands in Year 1 revenue for this "
                    "customer. Rounds = committed days ÷ this service's "
                    "own days/round (from its build-up tool further down "
                    "the page — shows '—' until that's run at least once "
                    "this session); Price/round = rate x days/round, so "
                    "Rounds x Price/round reconciles exactly to Amount "
                    "paid. Idle/prospect weeks contribute nothing here, "
                    "same as everywhere else."
                )

                _blended_yr1 = _customer_mix_blended_rate(_service_name, 1)
                st.metric("Blended rate, Year 1", fmt(_blended_yr1) + "/day")

    def _add_service_item():
        st.session_state.spot_service_items.append(
            {"name": "New service", "share_pct": 0.0, "rate_nok_day": 0.0, "escalator_pct": 0.0, "priced_at_baseline": False}
        )

    def _remove_service_item(index):
        st.session_state.spot_service_items.pop(index)

    def _on_service_rate_change(index):
        raw = st.session_state[f"service_rate_{index}"]
        value = parse_nok(raw)
        st.session_state.spot_service_items[index]["rate_nok_day"] = value
        st.session_state[f"service_rate_{index}"] = format_nok(value)

    shdr1, shdr2, shdr3, shdr4, shdr5, shdr6, shdr7, shdr8, shdr9, shdr10, shdr11 = st.columns(
        [1.3, 1.1, 0.8, 0.7, 1.0, 1.0, 1.2, 0.9, 0.9, 1.2, 0.4]
    )
    shdr1.markdown("**Service**")
    shdr2.markdown("**Share of working days (%)**")
    shdr3.markdown("**Days/year**")
    shdr4.markdown("**% of year**")
    shdr5.markdown("**Priced at baseline?**")
    shdr6.markdown("**Rate (NOK/day)**")
    shdr7.markdown("**Annual revenue (NOK)**")
    shdr8.markdown("**Rounds/year**")
    shdr9.markdown("**Days/round**")
    shdr10.markdown("**Payment/round (NOK)**")

    _sum_share = 0.0
    _sum_days = 0.0
    _sum_annual_revenue = 0.0

    _service_round_hours = st.session_state.get("_service_round_hours", {})

    for i, item in enumerate(st.session_state.spot_service_items):
        item.setdefault("priced_at_baseline", False)
        item.setdefault("share_pct", 0.0)
        item.setdefault("escalator_pct", 0.0)
        c1, c2, c3, c4, c5, c6, c7, c8, c9, c10, c11 = st.columns(
            [1.3, 1.1, 0.8, 0.7, 1.0, 1.0, 1.2, 0.9, 0.9, 1.2, 0.4]
        )
        with c1:
            item["name"] = st.text_input(
                "Service", value=item["name"], key=f"service_name_{i}", label_visibility="collapsed",
                disabled=locked
            )
        with c2:
            _is_calendar_driven = _calendar_enabled() and item["name"] in _service_slug_lookup
            if _is_calendar_driven:
                _cal_total_days_all = sum(_calendar_days_per_service.values())
                _implied_share_pct = (
                    _calendar_days_per_service.get(item["name"], 0.0) / _cal_total_days_all * 100
                ) if _cal_total_days_all else 0.0
                item["share_pct"] = _implied_share_pct  # kept in sync for export/display; not user-editable here
                st.markdown(f"<div style='padding-top:8px'>{_implied_share_pct:.1f}% (calendar)</div>", unsafe_allow_html=True)
            else:
                item["share_pct"] = st.number_input(
                    "Share of working days (%)", min_value=0.0, max_value=100.0, value=item["share_pct"],
                    step=1.0, key=f"service_share_{i}", label_visibility="collapsed", disabled=locked
                )
        with c3:
            if _is_calendar_driven:
                _days_this = _calendar_annual_days(item["name"])
            else:
                _days_this = _working_days_annual_target * (item["share_pct"] / 100)
            item["days_per_year"] = _days_this  # cached/computed — everything downstream still just reads this
            _days_label = f"{fmt(_days_this)} (calendar)" if _is_calendar_driven else fmt(_days_this)
            st.markdown(f"<div style='padding-top:8px'>{_days_label}</div>", unsafe_allow_html=True)
        with c4:
            _pct_of_year = (_days_this / operating_days * 100) if operating_days else 0.0
            st.markdown(f"<div style='padding-top:8px'>{_pct_of_year:.1f}%</div>", unsafe_allow_html=True)
        with c5:
            item["priced_at_baseline"] = st.checkbox(
                "Priced at baseline?", value=item["priced_at_baseline"],
                key=f"service_baseline_{i}", label_visibility="collapsed", disabled=locked
            )
        with c6:
            if item["priced_at_baseline"]:
                _row_rate = required_gross_rate_at_utilization
                st.markdown(f"<div style='padding-top:8px'>{format_nok(_row_rate)}</div>", unsafe_allow_html=True)
            elif _customer_mix_enabled(item["name"]):
                _row_rate = _customer_mix_blended_rate(item["name"], 1)
                st.markdown(f"<div style='padding-top:8px'>{format_nok(_row_rate)} (mix)</div>", unsafe_allow_html=True)
            else:
                _row_rate = item["rate_nok_day"]
                if f"service_rate_{i}" not in st.session_state:
                    st.session_state[f"service_rate_{i}"] = format_nok(item["rate_nok_day"])
                st.text_input(
                    "Rate (NOK/day)", key=f"service_rate_{i}", label_visibility="collapsed",
                    on_change=_on_service_rate_change, args=(i,), disabled=locked
                )
        with c7:
            # Seasonality now lives entirely in _days_this (c3, via the
            # Monthly activity calendar when enabled) — this stays the
            # plain rate x days calc either way.
            _row_annual_revenue = _row_rate * _days_this
            st.markdown(f"<div style='padding-top:8px'>{fmt(_row_annual_revenue)}</div>", unsafe_allow_html=True)
        with c8:
            _hours_per_round_this = _service_round_hours.get(item["name"])
            if _hours_per_round_this:
                _rounds_this = (_days_this * 24) / _hours_per_round_this
                st.markdown(f"<div style='padding-top:8px'>{_rounds_this:.1f}</div>", unsafe_allow_html=True)
            else:
                st.markdown("<div style='padding-top:8px'>—</div>", unsafe_allow_html=True)
        with c9:
            if _hours_per_round_this:
                _days_per_round_this = _hours_per_round_this / 24
                st.markdown(f"<div style='padding-top:8px'>{_days_per_round_this:.1f}</div>", unsafe_allow_html=True)
            else:
                st.markdown("<div style='padding-top:8px'>—</div>", unsafe_allow_html=True)
        with c10:
            if _hours_per_round_this:
                _payment_per_round_this = _row_rate * _hours_per_round_this / 24
                st.markdown(f"<div style='padding-top:8px'>{fmt(_payment_per_round_this)}</div>", unsafe_allow_html=True)
            else:
                st.markdown("<div style='padding-top:8px'>—</div>", unsafe_allow_html=True)
        with c11:
            st.button("✕", key=f"service_remove_{i}", on_click=_remove_service_item, args=(i,), disabled=locked)

        _sum_share += item["share_pct"]
        _sum_days += _days_this
        _sum_annual_revenue += _row_annual_revenue

    st.button("+ Add service", on_click=_add_service_item, disabled=locked)

    if not _calendar_enabled() and abs(_sum_share - 100.0) > 0.5:
        st.warning(
            f"⚠️ Shares sum to {_sum_share:.1f}%, not 100%. Days/year above "
            f"are still computed as entered, but check the shares reflect "
            f"the intended mix."
        )

    _sum_pct_of_year = (_sum_days / operating_days * 100) if operating_days else 0.0
    tcol1, tcol2, tcol3, tcol4, tcol5 = st.columns([1.3, 1.1, 0.8, 0.7, 7.6])
    tcol1.markdown("**Total**")
    tcol2.markdown(f"**{_sum_share:.1f}%**")
    tcol3.markdown(f"**{fmt(_sum_days)}**")
    tcol4.markdown(f"**{_sum_pct_of_year:.1f}%**")
    tcol5.markdown(f"**{fmt(_sum_annual_revenue)}** (revenue only, excl. opex and price-list build-up)")

    _target_net_annual = spot_baseline_tc_daily * operating_days
    _target_gross_annual = _target_net_annual + spot_opex_annual_nok + _spot_variable_voyage_opex_prev_pass
    _target_delta = _sum_annual_revenue - _target_gross_annual
    st.caption(
        f"Target: {fmt(_target_net_annual)} net (baseline TC-rate x operating days) + "
        f"{fmt(spot_opex_annual_nok)} Fixed Voyage opex + "
        f"{fmt(_spot_variable_voyage_opex_prev_pass)} Variable Voyage opex "
        f"= **{fmt(_target_gross_annual)}** gross required. "
        f"Current total above: {fmt(_sum_annual_revenue)} "
        f"({'+' if _target_delta >= 0 else ''}{fmt(_target_delta)} vs. target). "
        f"Variable Voyage opex is one script pass behind (computed further "
        f"down, in the build-up tools) — it catches up automatically."
    )

    spot_service_items_current = [
        {
            "name": item["name"],
            "days_per_year": item["days_per_year"],
            "share_pct": item.get("share_pct", 0.0),
            "rate_nok_day": item["rate_nok_day"],
            "escalator_pct": item.get("escalator_pct", 0.0),
            "priced_at_baseline": item.get("priced_at_baseline", False),
        }
        for item in st.session_state.spot_service_items
    ]

    # Should equal _working_days_annual_target by construction, barring the
    # shares-not-summing-to-100% edge case flagged above.
    _working_days_annual = sum(item["days_per_year"] for item in spot_service_items_current)

    st.divider()
    st.markdown("**Transport voyage cost build-up** (per round trip)")
    st.caption(
        "Bottom-up, phase-by-phase cost for one smolt round trip — build "
        "this properly first, then copy the same structure for Other. "
        "Steaming phases derive fuel from speed (fuel burn scales roughly "
        "with speed cubed — hull resistance physics, hence 'almost double' "
        "going from 9 to 11 knots); Stationary phases (loading/offloading) "
        "use a directly-entered fuel rate instead, since propulsion "
        "physics don't apply while alongside. An additional flat cost/hour "
        "(crew overtime, wear, consumables, etc.) applies across every "
        "phase regardless of type. **Every number below — speed per "
        "Steaming phase, fuel rates, the exponent, fuel price — is a "
        "starting default, not fixed**: try raising the speed on 'Steam "
        "to client' or 'Steam to pens' to see the cost of running faster "
        "flow straight through, live."
    )

    smolt_gcol1, smolt_gcol2, smolt_gcol3, smolt_gcol4, smolt_gcol5 = st.columns(5)
    with smolt_gcol1:
        spot_smolt_ref_speed_kn = stateful_number_input(
            "Reference speed (knots)", min_value=0.1, value=9.0, step=0.5,
            key="spot_smolt_ref_speed", disabled=locked
        )
    with smolt_gcol2:
        spot_smolt_ref_fuel_lhr = stateful_number_input(
            "Fuel rate @ ref. speed, light (L/hr)", min_value=0.0, value=450.0, step=10.0,
            key="spot_smolt_ref_fuel", disabled=locked
        )
    with smolt_gcol3:
        spot_smolt_ref_fuel_lhr_loaded = stateful_number_input(
            "Fuel rate @ ref. speed, loaded (L/hr)", min_value=0.0, value=650.0, step=10.0,
            key="spot_smolt_ref_fuel_loaded", disabled=locked,
            help="Same reference speed as the light rate, but a fully "
                 "loaded vessel simply burns more fuel at that speed — "
                 "applies to phases marked 'Loaded' in the table below."
        )
    with smolt_gcol4:
        spot_smolt_speed_exponent = stateful_number_input(
            "Speed → fuel exponent", min_value=1.0, max_value=5.0, value=1.8, step=0.1,
            key="spot_smolt_speed_exp", disabled=locked,
            help="Fuel rate at any speed = reference rate x (speed / reference speed) ^ this exponent — "
                 "the same exponent applies whether light or loaded, only the reference rate itself differs. "
                 "3.0 is the standard hull-resistance approximation (fuel roughly triples if speed doubles)."
        )
    with smolt_gcol5:
        spot_smolt_fuel_price = stateful_number_input(
            "Fuel price (NOK/liter)", min_value=0.0, value=12.5, step=0.5,
            key="spot_smolt_fuel_price", disabled=locked
        )
    spot_smolt_additional_opex_hr = nok_input(
        "Additional cost per hour in operation (NOK/hr) — all phases",
        "spot_smolt_additional_opex_hr", 600.0,
        key="spot_smolt_additional_opex_input", disabled=locked
    )

    if "spot_smolt_segments" not in st.session_state:
        st.session_state.spot_smolt_segments = [
            {"name": "Steam to client", "type": "Steaming", "duration_hr": 8.0, "speed_kn": 9.0, "fuel_rate_lhr": 0.0, "load": "Light"},
            {"name": "Load smolt", "type": "Stationary", "duration_hr": 8.0, "speed_kn": 0.0, "fuel_rate_lhr": 300.0, "load": "Light"},
            {"name": "Steam to pens", "type": "Steaming", "duration_hr": 8.0, "speed_kn": 9.0, "fuel_rate_lhr": 0.0, "load": "Loaded"},
            {"name": "Offload smolt", "type": "Stationary", "duration_hr": 4.0, "speed_kn": 0.0, "fuel_rate_lhr": 310.0, "load": "Light"},
            {"name": "Others/waiting time", "type": "Stationary", "duration_hr": 4.5, "speed_kn": 0.0, "fuel_rate_lhr": 150.0, "load": "Light"},
        ]

    def _add_smolt_segment():
        st.session_state.spot_smolt_segments.append(
            {"name": "New phase", "type": "Stationary", "duration_hr": 0.0, "speed_kn": 0.0, "fuel_rate_lhr": 0.0, "load": "Light"}
        )

    def _remove_smolt_segment(index):
        st.session_state.spot_smolt_segments.pop(index)

    smhdr = st.columns([1.4, 0.9, 0.8, 0.8, 0.8, 1.0, 0.8, 1.0, 1.0, 1.1, 0.4])
    smhdr[0].markdown("**Phase**")
    smhdr[1].markdown("**Type**")
    smhdr[2].markdown("**Load**")
    smhdr[3].markdown("**Duration (hr)**")
    smhdr[4].markdown("**Speed (kn)**")
    smhdr[5].markdown("**Fuel rate (L/hr)**")
    smhdr[6].markdown("**Fuel (L)**")
    smhdr[7].markdown("**Fuel cost (NOK)**")
    smhdr[8].markdown("**Add'l opex (NOK)**")
    smhdr[9].markdown("**Total cost (NOK)**")

    _smolt_total_hours = 0.0
    _smolt_total_fuel_l = 0.0
    _smolt_total_fuel_cost = 0.0
    _smolt_total_additional_opex = 0.0
    _smolt_total_cost = 0.0

    for si, seg in enumerate(st.session_state.spot_smolt_segments):
        seg.setdefault("load", "Light")
        cols = st.columns([1.4, 0.9, 0.8, 0.8, 0.8, 1.0, 0.8, 1.0, 1.0, 1.1, 0.4])
        with cols[0]:
            seg["name"] = st.text_input(
                "Phase", value=seg["name"], key=f"smolt_seg_name_{si}", label_visibility="collapsed", disabled=locked
            )
        with cols[1]:
            seg["type"] = st.selectbox(
                "Type", ["Steaming", "Stationary"],
                index=0 if seg["type"] == "Steaming" else 1,
                key=f"smolt_seg_type_{si}", label_visibility="collapsed", disabled=locked
            )
        with cols[2]:
            if seg["type"] == "Steaming":
                seg["load"] = st.selectbox(
                    "Load", ["Light", "Loaded"],
                    index=0 if seg["load"] == "Light" else 1,
                    key=f"smolt_seg_load_{si}", label_visibility="collapsed", disabled=locked
                )
            else:
                st.markdown("<div style='padding-top:8px'>—</div>", unsafe_allow_html=True)
        with cols[3]:
            seg["duration_hr"] = st.number_input(
                "Duration (hr)", min_value=0.0, value=seg["duration_hr"], step=0.5,
                key=f"smolt_seg_duration_{si}", label_visibility="collapsed", disabled=locked
            )
        with cols[4]:
            if seg["type"] == "Steaming":
                seg["speed_kn"] = st.number_input(
                    "Speed (kn)", min_value=0.0, value=seg["speed_kn"], step=0.5,
                    key=f"smolt_seg_speed_{si}", label_visibility="collapsed", disabled=locked
                )
            else:
                st.markdown("<div style='padding-top:8px'>—</div>", unsafe_allow_html=True)
        with cols[5]:
            if seg["type"] == "Steaming":
                _ref_fuel_this = spot_smolt_ref_fuel_lhr_loaded if seg["load"] == "Loaded" else spot_smolt_ref_fuel_lhr
                _fuel_rate_this = _ref_fuel_this * (
                    (seg["speed_kn"] / spot_smolt_ref_speed_kn) ** spot_smolt_speed_exponent
                    if spot_smolt_ref_speed_kn else 0.0
                )
                st.markdown(f"<div style='padding-top:8px'>{fmt(_fuel_rate_this)}</div>", unsafe_allow_html=True)
            else:
                seg["fuel_rate_lhr"] = st.number_input(
                    "Fuel rate (L/hr)", min_value=0.0, value=seg["fuel_rate_lhr"], step=5.0,
                    key=f"smolt_seg_fuelrate_{si}", label_visibility="collapsed", disabled=locked
                )
                _fuel_rate_this = seg["fuel_rate_lhr"]
        with cols[6]:
            _fuel_this = _fuel_rate_this * seg["duration_hr"]
            st.markdown(f"<div style='padding-top:8px'>{fmt(_fuel_this)}</div>", unsafe_allow_html=True)
        with cols[7]:
            _fuel_cost_this = _fuel_this * spot_smolt_fuel_price
            st.markdown(f"<div style='padding-top:8px'>{fmt(_fuel_cost_this)}</div>", unsafe_allow_html=True)
        with cols[8]:
            _additional_opex_this = spot_smolt_additional_opex_hr * seg["duration_hr"]
            st.markdown(f"<div style='padding-top:8px'>{fmt(_additional_opex_this)}</div>", unsafe_allow_html=True)
        with cols[9]:
            _total_cost_this = _fuel_cost_this + _additional_opex_this
            st.markdown(f"<div style='padding-top:8px'>{fmt(_total_cost_this)}</div>", unsafe_allow_html=True)
        with cols[10]:
            st.button("✕", key=f"smolt_seg_remove_{si}", on_click=_remove_smolt_segment, args=(si,), disabled=locked)

        _smolt_total_hours += seg["duration_hr"]
        _smolt_total_fuel_l += _fuel_this
        _smolt_total_fuel_cost += _fuel_cost_this
        _smolt_total_additional_opex += _additional_opex_this
        _smolt_total_cost += _total_cost_this

    st.button("+ Add phase", key="smolt_add_phase", on_click=_add_smolt_segment, disabled=locked)

    tcols = st.columns([1.4, 0.9, 0.8, 0.8, 0.8, 1.0, 0.8, 1.0, 1.0, 1.1, 0.4])
    tcols[0].markdown("**Round trip total**")
    tcols[3].markdown(f"**{fmt(_smolt_total_hours)}**")
    tcols[6].markdown(f"**{fmt(_smolt_total_fuel_l)}**")
    tcols[7].markdown(f"**{fmt(_smolt_total_fuel_cost)}**")
    tcols[8].markdown(f"**{fmt(_smolt_total_additional_opex)}**")
    tcols[9].markdown(f"**{fmt(_smolt_total_cost)}**")

    _smolt_days_available = next(
        (item["days_per_year"] for item in spot_service_items_current if item["name"] == "Transport"),
        0.0
    )
    _smolt_hours_available = _smolt_days_available * 24
    _smolt_trips_exact = (_smolt_hours_available / _smolt_total_hours) if _smolt_total_hours else 0.0
    _smolt_trips_whole = int(_smolt_trips_exact)
    _smolt_annual_voyage_cost = _smolt_total_cost * _smolt_trips_exact
    _smolt_implied_day_rate = (_smolt_annual_voyage_cost / _smolt_days_available) if _smolt_days_available else 0.0

    _smolt_liters_per_year = _smolt_total_fuel_l * _smolt_trips_exact

    sm1, sm2, sm3, sm4, sm5 = st.columns(5)
    sm1.metric("Hours per round trip", fmt(_smolt_total_hours))
    sm2.metric(f"Trips available ({fmt(_smolt_days_available)}-day window)", f"{_smolt_trips_exact:.1f}", help=f"{_smolt_trips_whole} whole trips + a partial trip, over {fmt(_smolt_days_available)} days available (from the Service mix table).")
    sm3.metric("Liters consumed per year", fmt(_smolt_liters_per_year) + " L", help=f"{fmt(_smolt_total_fuel_l)} L per round trip x {_smolt_trips_exact:.2f} trips/year.")
    sm4.metric("Implied annual voyage cost (round trips only)", fmt(_smolt_annual_voyage_cost))
    sm5.metric("Implied day-rate (round trips only)", fmt(_smolt_implied_day_rate) + "/day")
    st.caption(
        f"Days available for Transport: {fmt(_smolt_days_available)} (from the Service mix table above) "
        f"x 24 hr = {fmt(_smolt_hours_available)} hours ÷ {fmt(_smolt_total_hours)} hours/round trip "
        f"= {_smolt_trips_exact:.2f} trips/year."
    )

    st.divider()
    st.markdown("**Customer changeover costs** (per year)")
    _smolt_booking_blocks = _customer_booking_blocks("Transport")
    _smolt_booking_driven = _customer_mix_enabled("Transport") and bool(_smolt_booking_blocks)
    if _smolt_booking_driven:
        st.caption(
            "Driven by the actual bookings in Customer mix above — "
            f"**{len(_smolt_booking_blocks)} booking block(s)** currently "
            f"assigned ({', '.join(f'{name} ({wk}w)' for name, wk in _smolt_booking_blocks)}), "
            f"giving **{len(_smolt_booking_blocks)} changeover(s)** — one "
            "deep disinfection per customer arrival, N bookings needs N "
            "changeovers. Intermediate cleans scale with each booking's "
            "own length instead of a flat count — turn off customer mix "
            "above to go back to a manual Customers/year estimate instead."
        )
        _smolt_n_changeovers = len(_smolt_booking_blocks)
        _smolt_cust_metric_col = st.columns(1)[0]
        _smolt_cust_metric_col.metric("Number of customers (= deep disinfections/year)", f"{_smolt_n_changeovers}")
        cc2 = st.columns(1)[0]
        with cc2:
            spot_smolt_weeks_per_intermediate_clean = stateful_number_input(
                "Weeks per intermediate clean", min_value=0.5, value=4.0, step=0.5,
                key="spot_smolt_weeks_per_intermediate_clean", disabled=locked,
                help="1 intermediate clean for every this-many weeks a booking spans — a booking shorter than this gets none."
            )
    else:
        st.caption(
            "Separate from the round-trip cost above — these happen per "
            "customer relationship, not per round trip. Defaults below assume "
            "your restated total (8 deep cleans, 8 intermediate cleans); "
            "'Intermediate cleans per customer' defaults to 1 so 8 customers "
            "x 1 = 8 — bump it to 2 directly if you actually meant twice per "
            "customer (16/year). Enable customer mix above (with weeks "
            "actually booked) to derive these from real bookings instead."
        )
        cc1, cc2 = st.columns(2)
        with cc1:
            spot_smolt_customers_per_year = stateful_number_input(
                "Customers/year", min_value=0.0, value=8.0, step=1.0,
                key="spot_smolt_customers_per_year", disabled=locked
            )
        with cc2:
            _rounds_per_customer = (_smolt_trips_exact / spot_smolt_customers_per_year) if spot_smolt_customers_per_year else 0.0
            st.metric("Rounds per customer (check)", f"{_rounds_per_customer:.1f}", help="Trips available ÷ customers/year — should land near your expected 4-5 rounds per customer.")
        _smolt_n_changeovers = spot_smolt_customers_per_year

    cc3, cc4 = st.columns(2)
    with cc3:
        st.markdown("**Deep disinfection**")
        spot_smolt_deep_clean_days = stateful_number_input(
            "Days at yard", min_value=0.0, value=1.0, step=0.5,
            key="spot_smolt_deep_clean_days", disabled=locked
        )
        spot_smolt_yard_cost_per_day = nok_input(
            "Disinfection opex (NOK/day)", "spot_smolt_yard_cost_per_day", 25_000.0,
            key="spot_smolt_yard_cost_input", disabled=locked
        )
        spot_smolt_drydock_cost_per_day = nok_input(
            "Dry-docking cost (NOK/day)", "spot_smolt_drydock_cost_per_day", 75_000.0,
            key="spot_smolt_drydock_cost_input", disabled=locked
        )
        st.caption(
            "Cost/event = days at yard x (disinfection opex/day + dry-docking "
            "cost/day). Once per customer change — frequency = " +
            ("changeovers derived from bookings." if _smolt_booking_driven else "customers/year.")
        )
    with cc4:
        st.markdown("**Intermediate clean**")
        spot_smolt_intermediate_clean_hr = stateful_number_input(
            "Duration (hr)", min_value=0.0, value=3.0, step=0.5,
            key="spot_smolt_intermediate_clean_hr", disabled=locked
        )
        spot_smolt_intermediate_cost_per_hr = nok_input(
            "Cost per hour (NOK/hr)", "spot_smolt_intermediate_cost_per_hr", 2_500.0,
            key="spot_smolt_intermediate_cost_input", disabled=locked
        )
        if not _smolt_booking_driven:
            spot_smolt_intermediate_cleans_per_customer = stateful_number_input(
                "Intermediate cleans per customer (0 to skip)", min_value=0.0, value=1.0, step=1.0,
                key="spot_smolt_intermediate_cleans_per_customer", disabled=locked
            )
        st.caption(
            "Cost/event = duration x cost/hour — no dry-dock, no fuel "
            "physics, just a blended operational rate."
        )

    spot_smolt_deep_clean_cost = (
        spot_smolt_deep_clean_days * (spot_smolt_yard_cost_per_day + spot_smolt_drydock_cost_per_day)
    )
    spot_smolt_intermediate_clean_cost = (
        spot_smolt_intermediate_clean_hr * spot_smolt_intermediate_cost_per_hr
    )
    st.caption(
        f"Computed cost per event — Deep disinfection: {fmt(spot_smolt_deep_clean_cost)}, "
        f"Intermediate clean: {fmt(spot_smolt_intermediate_clean_cost)}."
    )

    if _smolt_booking_driven:
        spot_smolt_n_intermediate_cleans = sum(
            int(weeks // spot_smolt_weeks_per_intermediate_clean) for _, weeks in _smolt_booking_blocks
        ) if spot_smolt_weeks_per_intermediate_clean else 0
    else:
        spot_smolt_n_intermediate_cleans = spot_smolt_customers_per_year * spot_smolt_intermediate_cleans_per_customer

    cc5, cc6 = st.columns(2)
    with cc5:
        spot_smolt_transport_base_hr = stateful_number_input(
            "Transport back to base — duration (hr)", min_value=0.0, value=8.0, step=0.5,
            key="spot_smolt_transport_base_hr", disabled=locked
        )
    with cc6:
        spot_smolt_transport_base_speed = stateful_number_input(
            "Transport back to base — speed (kn)", min_value=0.0, value=spot_smolt_ref_speed_kn, step=0.5,
            key="spot_smolt_transport_base_speed", disabled=locked
        )
    st.caption(
        "Once per customer change — frequency = " +
        ("changeovers derived from bookings." if _smolt_booking_driven else "customers/year.") +
        " Fuel cost uses the same speed formula and additional cost/hour "
        "as the round-trip phases above."
    )

    _transport_base_fuel_rate = spot_smolt_ref_fuel_lhr * (
        (spot_smolt_transport_base_speed / spot_smolt_ref_speed_kn) ** spot_smolt_speed_exponent
        if spot_smolt_ref_speed_kn else 0.0
    )
    _transport_base_cost_per_event = (
        (_transport_base_fuel_rate * spot_smolt_transport_base_hr * spot_smolt_fuel_price)
        + (spot_smolt_additional_opex_hr * spot_smolt_transport_base_hr)
    )

    _annual_deep_clean_cost = _smolt_n_changeovers * spot_smolt_deep_clean_cost
    _annual_intermediate_clean_cost = spot_smolt_n_intermediate_cleans * spot_smolt_intermediate_clean_cost
    _annual_transport_base_cost = _smolt_n_changeovers * _transport_base_cost_per_event
    _annual_changeover_cost = _annual_deep_clean_cost + _annual_intermediate_clean_cost + _annual_transport_base_cost

    changeover_df = pd.DataFrame([
        {"Item": "Deep disinfection", "Events/year": _smolt_n_changeovers, "Cost/event (NOK)": spot_smolt_deep_clean_cost, "Annual cost (NOK)": _annual_deep_clean_cost},
        {"Item": "Intermediate clean", "Events/year": spot_smolt_n_intermediate_cleans, "Cost/event (NOK)": spot_smolt_intermediate_clean_cost, "Annual cost (NOK)": _annual_intermediate_clean_cost},
        {"Item": "Transport back to base", "Events/year": _smolt_n_changeovers, "Cost/event (NOK)": _transport_base_cost_per_event, "Annual cost (NOK)": _annual_transport_base_cost},
        {"Item": "Total", "Events/year": None, "Cost/event (NOK)": None, "Annual cost (NOK)": _annual_changeover_cost},
    ])
    show_table(changeover_df, "Item", width="stretch")

    _smolt_total_annual_voyage_cost = _smolt_annual_voyage_cost + _annual_changeover_cost
    _smolt_total_implied_day_rate = (
        _smolt_total_annual_voyage_cost / _smolt_days_available if _smolt_days_available else 0.0
    )

    st.markdown("**Combined total (round trips + customer changeover)**")
    ct1, ct2 = st.columns(2)
    ct1.metric("Total implied annual voyage cost", fmt(_smolt_total_annual_voyage_cost))
    ct2.metric("Total implied day-rate", fmt(_smolt_total_implied_day_rate) + "/day")
    st.caption(
        "This is a build-up tool — the resulting day-rate isn't wired into "
        "the Voyage costs table below yet; once you're happy with this and "
        "the Other build-up further down, say so and I'll connect both "
        "into the actual Voyage costs table."
    )

    st.markdown("**Net income check — Transport**")
    st.caption(
        "Links the customer rate set on the Service mix table above "
        "against this build-up's voyage cost, so you can see the "
        "contribution margin move live as you test different rates in "
        "discussion."
    )
    _smolt_charged_rate = _resolved_charged_rate("Transport")
    _smolt_charged_annual_revenue = _smolt_charged_rate * _smolt_days_available
    _smolt_net_income_day = _smolt_charged_rate - _smolt_total_implied_day_rate
    _smolt_net_income_annual = _smolt_charged_annual_revenue - _smolt_total_annual_voyage_cost

    ni1, ni2, ni3, ni4 = st.columns(4)
    ni1.metric("Charged rate (NOK/day)", fmt(_smolt_charged_rate) + "/day")
    ni2.metric("Charged annual revenue (NOK)", fmt(_smolt_charged_annual_revenue))
    ni3.metric("Net income (NOK/day)", fmt(_smolt_net_income_day) + "/day")
    ni4.metric("Net income (NOK/year)", fmt(_smolt_net_income_annual))

    st.divider()
    st.markdown("**Other voyage cost build-up** (per round trip)")
    st.caption(
        "Bottom-up, phase-by-phase cost for one harvest round trip — same "
        "structure as the Transport build-up above, adapted for harvest-size "
        "(~5kg) fish and reversed logistics (picking fish up from the "
        "pens, delivering to the processing plant, rather than delivering "
        "smolt to the pens). Steaming phases derive fuel from speed (fuel "
        "burn scales roughly with speed cubed — hull resistance physics, "
        "hence 'almost double' going from 9 to 11 knots); Stationary "
        "phases (loading/offloading) use a directly-entered fuel rate "
        "instead, since propulsion physics don't apply while alongside. "
        "An additional flat cost/hour (crew overtime, wear, consumables, "
        "etc.) applies across every phase regardless of type. **Every "
        "number below — speed per Steaming phase, fuel rates, the "
        "exponent, fuel price — is a starting default, not fixed**: try "
        "raising the speed on 'Steam to pens' or 'Steam to processing "
        "plant' to see the cost of running faster flow straight through, "
        "live."
    )

    harvest_gcol1, harvest_gcol2, harvest_gcol3, harvest_gcol4, harvest_gcol5 = st.columns(5)
    with harvest_gcol1:
        spot_harvest_ref_speed_kn = stateful_number_input(
            "Reference speed (knots)", min_value=0.1, value=9.0, step=0.5,
            key="spot_harvest_ref_speed", disabled=locked
        )
    with harvest_gcol2:
        spot_harvest_ref_fuel_lhr = stateful_number_input(
            "Fuel rate @ ref. speed, light (L/hr)", min_value=0.0, value=450.0, step=10.0,
            key="spot_harvest_ref_fuel", disabled=locked
        )
    with harvest_gcol3:
        spot_harvest_ref_fuel_lhr_loaded = stateful_number_input(
            "Fuel rate @ ref. speed, loaded (L/hr)", min_value=0.0, value=650.0, step=10.0,
            key="spot_harvest_ref_fuel_loaded", disabled=locked,
            help="Same reference speed as the light rate, but a fully "
                 "loaded vessel simply burns more fuel at that speed — "
                 "applies to phases marked 'Loaded' in the table below."
        )
    with harvest_gcol4:
        spot_harvest_speed_exponent = stateful_number_input(
            "Speed → fuel exponent", min_value=1.0, max_value=5.0, value=1.8, step=0.1,
            key="spot_harvest_speed_exp", disabled=locked,
            help="Fuel rate at any speed = reference rate x (speed / reference speed) ^ this exponent — "
                 "the same exponent applies whether light or loaded, only the reference rate itself differs. "
                 "3.0 is the standard hull-resistance approximation (fuel roughly triples if speed doubles)."
        )
    with harvest_gcol5:
        spot_harvest_fuel_price = stateful_number_input(
            "Fuel price (NOK/liter)", min_value=0.0, value=12.5, step=0.5,
            key="spot_harvest_fuel_price", disabled=locked
        )
    spot_harvest_additional_opex_hr = nok_input(
        "Additional cost per hour in operation (NOK/hr) — all phases",
        "spot_harvest_additional_opex_hr", 600.0,
        key="spot_harvest_additional_opex_input", disabled=locked
    )

    if "spot_harvest_segments" not in st.session_state:
        st.session_state.spot_harvest_segments = [
            {"name": "Steam to pens", "type": "Steaming", "duration_hr": 8.0, "speed_kn": 9.0, "fuel_rate_lhr": 0.0, "load": "Light"},
            {"name": "Load fish (5kg)", "type": "Stationary", "duration_hr": 8.0, "speed_kn": 0.0, "fuel_rate_lhr": 50.0, "load": "Light"},
            {"name": "Steam to processing plant", "type": "Steaming", "duration_hr": 8.0, "speed_kn": 9.0, "fuel_rate_lhr": 0.0, "load": "Loaded"},
            {"name": "Offload fish (5kg)", "type": "Stationary", "duration_hr": 4.0, "speed_kn": 0.0, "fuel_rate_lhr": 50.0, "load": "Light"},
            {"name": "Others/waiting time", "type": "Stationary", "duration_hr": 4.5, "speed_kn": 0.0, "fuel_rate_lhr": 150.0, "load": "Light"},
        ]

    def _add_harvest_segment():
        st.session_state.spot_harvest_segments.append(
            {"name": "New phase", "type": "Stationary", "duration_hr": 0.0, "speed_kn": 0.0, "fuel_rate_lhr": 0.0, "load": "Light"}
        )

    def _remove_harvest_segment(index):
        st.session_state.spot_harvest_segments.pop(index)

    smhdr = st.columns([1.4, 0.9, 0.8, 0.8, 0.8, 1.0, 0.8, 1.0, 1.0, 1.1, 0.4])
    smhdr[0].markdown("**Phase**")
    smhdr[1].markdown("**Type**")
    smhdr[2].markdown("**Load**")
    smhdr[3].markdown("**Duration (hr)**")
    smhdr[4].markdown("**Speed (kn)**")
    smhdr[5].markdown("**Fuel rate (L/hr)**")
    smhdr[6].markdown("**Fuel (L)**")
    smhdr[7].markdown("**Fuel cost (NOK)**")
    smhdr[8].markdown("**Add'l opex (NOK)**")
    smhdr[9].markdown("**Total cost (NOK)**")

    _harvest_total_hours = 0.0
    _harvest_total_fuel_l = 0.0
    _harvest_total_fuel_cost = 0.0
    _harvest_total_additional_opex = 0.0
    _harvest_total_cost = 0.0

    for si, seg in enumerate(st.session_state.spot_harvest_segments):
        seg.setdefault("load", "Light")
        cols = st.columns([1.4, 0.9, 0.8, 0.8, 0.8, 1.0, 0.8, 1.0, 1.0, 1.1, 0.4])
        with cols[0]:
            seg["name"] = st.text_input(
                "Phase", value=seg["name"], key=f"harvest_seg_name_{si}", label_visibility="collapsed", disabled=locked
            )
        with cols[1]:
            seg["type"] = st.selectbox(
                "Type", ["Steaming", "Stationary"],
                index=0 if seg["type"] == "Steaming" else 1,
                key=f"harvest_seg_type_{si}", label_visibility="collapsed", disabled=locked
            )
        with cols[2]:
            if seg["type"] == "Steaming":
                seg["load"] = st.selectbox(
                    "Load", ["Light", "Loaded"],
                    index=0 if seg["load"] == "Light" else 1,
                    key=f"harvest_seg_load_{si}", label_visibility="collapsed", disabled=locked
                )
            else:
                st.markdown("<div style='padding-top:8px'>—</div>", unsafe_allow_html=True)
        with cols[3]:
            seg["duration_hr"] = st.number_input(
                "Duration (hr)", min_value=0.0, value=seg["duration_hr"], step=0.5,
                key=f"harvest_seg_duration_{si}", label_visibility="collapsed", disabled=locked
            )
        with cols[4]:
            if seg["type"] == "Steaming":
                seg["speed_kn"] = st.number_input(
                    "Speed (kn)", min_value=0.0, value=seg["speed_kn"], step=0.5,
                    key=f"harvest_seg_speed_{si}", label_visibility="collapsed", disabled=locked
                )
            else:
                st.markdown("<div style='padding-top:8px'>—</div>", unsafe_allow_html=True)
        with cols[5]:
            if seg["type"] == "Steaming":
                _ref_fuel_this = spot_harvest_ref_fuel_lhr_loaded if seg["load"] == "Loaded" else spot_harvest_ref_fuel_lhr
                _fuel_rate_this = _ref_fuel_this * (
                    (seg["speed_kn"] / spot_harvest_ref_speed_kn) ** spot_harvest_speed_exponent
                    if spot_harvest_ref_speed_kn else 0.0
                )
                st.markdown(f"<div style='padding-top:8px'>{fmt(_fuel_rate_this)}</div>", unsafe_allow_html=True)
            else:
                seg["fuel_rate_lhr"] = st.number_input(
                    "Fuel rate (L/hr)", min_value=0.0, value=seg["fuel_rate_lhr"], step=5.0,
                    key=f"harvest_seg_fuelrate_{si}", label_visibility="collapsed", disabled=locked
                )
                _fuel_rate_this = seg["fuel_rate_lhr"]
        with cols[6]:
            _fuel_this = _fuel_rate_this * seg["duration_hr"]
            st.markdown(f"<div style='padding-top:8px'>{fmt(_fuel_this)}</div>", unsafe_allow_html=True)
        with cols[7]:
            _fuel_cost_this = _fuel_this * spot_harvest_fuel_price
            st.markdown(f"<div style='padding-top:8px'>{fmt(_fuel_cost_this)}</div>", unsafe_allow_html=True)
        with cols[8]:
            _additional_opex_this = spot_harvest_additional_opex_hr * seg["duration_hr"]
            st.markdown(f"<div style='padding-top:8px'>{fmt(_additional_opex_this)}</div>", unsafe_allow_html=True)
        with cols[9]:
            _total_cost_this = _fuel_cost_this + _additional_opex_this
            st.markdown(f"<div style='padding-top:8px'>{fmt(_total_cost_this)}</div>", unsafe_allow_html=True)
        with cols[10]:
            st.button("✕", key=f"harvest_seg_remove_{si}", on_click=_remove_harvest_segment, args=(si,), disabled=locked)

        _harvest_total_hours += seg["duration_hr"]
        _harvest_total_fuel_l += _fuel_this
        _harvest_total_fuel_cost += _fuel_cost_this
        _harvest_total_additional_opex += _additional_opex_this
        _harvest_total_cost += _total_cost_this

    st.button("+ Add phase", key="harvest_add_phase", on_click=_add_harvest_segment, disabled=locked)

    tcols = st.columns([1.4, 0.9, 0.8, 0.8, 0.8, 1.0, 0.8, 1.0, 1.0, 1.1, 0.4])
    tcols[0].markdown("**Round trip total**")
    tcols[3].markdown(f"**{fmt(_harvest_total_hours)}**")
    tcols[6].markdown(f"**{fmt(_harvest_total_fuel_l)}**")
    tcols[7].markdown(f"**{fmt(_harvest_total_fuel_cost)}**")
    tcols[8].markdown(f"**{fmt(_harvest_total_additional_opex)}**")
    tcols[9].markdown(f"**{fmt(_harvest_total_cost)}**")

    _harvest_days_available = next(
        (item["days_per_year"] for item in spot_service_items_current if item["name"] == "Other"),
        0.0
    )
    _harvest_hours_available = _harvest_days_available * 24
    _harvest_trips_exact = (_harvest_hours_available / _harvest_total_hours) if _harvest_total_hours else 0.0
    _harvest_trips_whole = int(_harvest_trips_exact)
    _harvest_annual_voyage_cost = _harvest_total_cost * _harvest_trips_exact
    _harvest_implied_day_rate = (_harvest_annual_voyage_cost / _harvest_days_available) if _harvest_days_available else 0.0

    _harvest_liters_per_year = _harvest_total_fuel_l * _harvest_trips_exact

    sm1, sm2, sm3, sm4, sm5 = st.columns(5)
    sm1.metric("Hours per round trip", fmt(_harvest_total_hours))
    sm2.metric(f"Trips available ({fmt(_harvest_days_available)}-day window)", f"{_harvest_trips_exact:.1f}", help=f"{_harvest_trips_whole} whole trips + a partial trip, over {fmt(_harvest_days_available)} days available (from the Service mix table).")
    sm3.metric("Liters consumed per year", fmt(_harvest_liters_per_year) + " L", help=f"{fmt(_harvest_total_fuel_l)} L per round trip x {_harvest_trips_exact:.2f} trips/year.")
    sm4.metric("Implied annual voyage cost (round trips only)", fmt(_harvest_annual_voyage_cost))
    sm5.metric("Implied day-rate (round trips only)", fmt(_harvest_implied_day_rate) + "/day")
    st.caption(
        f"Days available for Other: {fmt(_harvest_days_available)} (from the Service mix table above) "
        f"x 24 hr = {fmt(_harvest_hours_available)} hours ÷ {fmt(_harvest_total_hours)} hours/round trip "
        f"= {_harvest_trips_exact:.2f} trips/year."
    )

    st.divider()
    st.markdown("**Customer changeover costs** (per year)")
    _harvest_booking_blocks = _customer_booking_blocks("Other")
    _harvest_booking_driven = _customer_mix_enabled("Other") and bool(_harvest_booking_blocks)
    if _harvest_booking_driven:
        st.caption(
            "Driven by the actual bookings in Customer mix above — "
            f"**{len(_harvest_booking_blocks)} booking block(s)** currently "
            f"assigned ({', '.join(f'{name} ({wk}w)' for name, wk in _harvest_booking_blocks)}), "
            f"giving **{len(_harvest_booking_blocks)} changeover(s)** — one "
            "deep disinfection per customer arrival, N bookings needs N "
            "changeovers. Intermediate cleans scale with each booking's "
            "own length instead of a flat count — turn off customer mix "
            "above to go back to a manual Customers/year estimate instead."
        )
        _harvest_n_changeovers = len(_harvest_booking_blocks)
        _harvest_cust_metric_col = st.columns(1)[0]
        _harvest_cust_metric_col.metric("Number of customers (= deep disinfections/year)", f"{_harvest_n_changeovers}")
        cc2 = st.columns(1)[0]
        with cc2:
            spot_harvest_weeks_per_intermediate_clean = stateful_number_input(
                "Weeks per intermediate clean", min_value=0.5, value=4.0, step=0.5,
                key="spot_harvest_weeks_per_intermediate_clean", disabled=locked,
                help="1 intermediate clean for every this-many weeks a booking spans — a booking shorter than this gets none."
            )
    else:
        st.caption(
            "Separate from the round-trip cost above — these happen per "
            "customer relationship, not per round trip. Customers/year "
            "defaults to 4 here (vs. 8 for Transport), since harvest activity "
            "takes roughly half the time. 'Intermediate cleans per customer' "
            "defaults to 1 so 4 customers x 1 = 4/year — adjust directly if "
            "needed. Enable customer mix above (with weeks actually "
            "booked) to derive these from real bookings instead."
        )
        cc1, cc2 = st.columns(2)
        with cc1:
            spot_harvest_customers_per_year = stateful_number_input(
                "Customers/year", min_value=0.0, value=4.0, step=1.0,
                key="spot_harvest_customers_per_year", disabled=locked
            )
        with cc2:
            _rounds_per_customer = (_harvest_trips_exact / spot_harvest_customers_per_year) if spot_harvest_customers_per_year else 0.0
            st.metric("Rounds per customer (check)", f"{_rounds_per_customer:.1f}", help="Trips available ÷ customers/year — a sanity check on the mix, not a fixed target.")
        _harvest_n_changeovers = spot_harvest_customers_per_year

    cc3, cc4 = st.columns(2)
    with cc3:
        st.markdown("**Deep disinfection**")
        spot_harvest_deep_clean_days = stateful_number_input(
            "Days at yard", min_value=0.0, value=1.0, step=0.5,
            key="spot_harvest_deep_clean_days", disabled=locked
        )
        spot_harvest_yard_cost_per_day = nok_input(
            "Disinfection opex (NOK/day)", "spot_harvest_yard_cost_per_day", 25_000.0,
            key="spot_harvest_yard_cost_input", disabled=locked
        )
        spot_harvest_drydock_cost_per_day = nok_input(
            "Dry-docking cost (NOK/day)", "spot_harvest_drydock_cost_per_day", 75_000.0,
            key="spot_harvest_drydock_cost_input", disabled=locked
        )
        st.caption(
            "Cost/event = days at yard x (disinfection opex/day + dry-docking "
            "cost/day). Once per customer change — frequency = " +
            ("changeovers derived from bookings." if _harvest_booking_driven else "customers/year.")
        )
    with cc4:
        st.markdown("**Intermediate clean**")
        spot_harvest_intermediate_clean_hr = stateful_number_input(
            "Duration (hr)", min_value=0.0, value=3.0, step=0.5,
            key="spot_harvest_intermediate_clean_hr", disabled=locked
        )
        spot_harvest_intermediate_cost_per_hr = nok_input(
            "Cost per hour (NOK/hr)", "spot_harvest_intermediate_cost_per_hr", 2_500.0,
            key="spot_harvest_intermediate_cost_input", disabled=locked
        )
        if not _harvest_booking_driven:
            spot_harvest_intermediate_cleans_per_customer = stateful_number_input(
                "Intermediate cleans per customer (0 to skip)", min_value=0.0, value=1.0, step=1.0,
                key="spot_harvest_intermediate_cleans_per_customer", disabled=locked
            )
        st.caption(
            "Cost/event = duration x cost/hour — no dry-dock, no fuel "
            "physics, just a blended operational rate."
        )

    spot_harvest_deep_clean_cost = (
        spot_harvest_deep_clean_days * (spot_harvest_yard_cost_per_day + spot_harvest_drydock_cost_per_day)
    )
    spot_harvest_intermediate_clean_cost = (
        spot_harvest_intermediate_clean_hr * spot_harvest_intermediate_cost_per_hr
    )
    st.caption(
        f"Computed cost per event — Deep disinfection: {fmt(spot_harvest_deep_clean_cost)}, "
        f"Intermediate clean: {fmt(spot_harvest_intermediate_clean_cost)}."
    )

    if _harvest_booking_driven:
        spot_harvest_n_intermediate_cleans = sum(
            int(weeks // spot_harvest_weeks_per_intermediate_clean) for _, weeks in _harvest_booking_blocks
        ) if spot_harvest_weeks_per_intermediate_clean else 0
    else:
        spot_harvest_n_intermediate_cleans = spot_harvest_customers_per_year * spot_harvest_intermediate_cleans_per_customer

    cc5, cc6 = st.columns(2)
    with cc5:
        spot_harvest_transport_base_hr = stateful_number_input(
            "Transport back to base — duration (hr)", min_value=0.0, value=8.0, step=0.5,
            key="spot_harvest_transport_base_hr", disabled=locked
        )
    with cc6:
        spot_harvest_transport_base_speed = stateful_number_input(
            "Transport back to base — speed (kn)", min_value=0.0, value=spot_harvest_ref_speed_kn, step=0.5,
            key="spot_harvest_transport_base_speed", disabled=locked
        )
    st.caption(
        "Once per customer change — frequency = " +
        ("changeovers derived from bookings." if _harvest_booking_driven else "customers/year.") +
        " Fuel cost uses the same speed formula and additional cost/hour "
        "as the round-trip phases above."
    )

    _transport_base_fuel_rate = spot_harvest_ref_fuel_lhr * (
        (spot_harvest_transport_base_speed / spot_harvest_ref_speed_kn) ** spot_harvest_speed_exponent
        if spot_harvest_ref_speed_kn else 0.0
    )
    _transport_base_cost_per_event = (
        (_transport_base_fuel_rate * spot_harvest_transport_base_hr * spot_harvest_fuel_price)
        + (spot_harvest_additional_opex_hr * spot_harvest_transport_base_hr)
    )

    _annual_deep_clean_cost = _harvest_n_changeovers * spot_harvest_deep_clean_cost
    _annual_intermediate_clean_cost = spot_harvest_n_intermediate_cleans * spot_harvest_intermediate_clean_cost
    _annual_transport_base_cost = _harvest_n_changeovers * _transport_base_cost_per_event
    _annual_changeover_cost = _annual_deep_clean_cost + _annual_intermediate_clean_cost + _annual_transport_base_cost

    changeover_df = pd.DataFrame([
        {"Item": "Deep disinfection", "Events/year": _harvest_n_changeovers, "Cost/event (NOK)": spot_harvest_deep_clean_cost, "Annual cost (NOK)": _annual_deep_clean_cost},
        {"Item": "Intermediate clean", "Events/year": spot_harvest_n_intermediate_cleans, "Cost/event (NOK)": spot_harvest_intermediate_clean_cost, "Annual cost (NOK)": _annual_intermediate_clean_cost},
        {"Item": "Transport back to base", "Events/year": _harvest_n_changeovers, "Cost/event (NOK)": _transport_base_cost_per_event, "Annual cost (NOK)": _annual_transport_base_cost},
        {"Item": "Total", "Events/year": None, "Cost/event (NOK)": None, "Annual cost (NOK)": _annual_changeover_cost},
    ])
    show_table(changeover_df, "Item", width="stretch")

    _harvest_total_annual_voyage_cost = _harvest_annual_voyage_cost + _annual_changeover_cost
    _harvest_total_implied_day_rate = (
        _harvest_total_annual_voyage_cost / _harvest_days_available if _harvest_days_available else 0.0
    )

    st.markdown("**Combined total (round trips + customer changeover)**")
    ct1, ct2 = st.columns(2)
    ct1.metric("Total implied annual voyage cost", fmt(_harvest_total_annual_voyage_cost))
    ct2.metric("Total implied day-rate", fmt(_harvest_total_implied_day_rate) + "/day")
    st.caption(
        "This is a build-up tool — the resulting day-rate isn't wired into "
        "the Voyage costs table below yet; once you're happy with both "
        "this and the Transport build-up above, say so and I'll connect both "
        "into the actual Voyage costs table."
    )

    st.markdown("**Net income check — Other**")
    st.caption(
        "Links the customer rate set on the Service mix table above "
        "against this build-up's voyage cost, so you can see the "
        "contribution margin move live as you test different rates in "
        "discussion."
    )
    _harvest_charged_rate = _resolved_charged_rate("Other")
    _harvest_charged_annual_revenue = _harvest_charged_rate * _harvest_days_available
    _harvest_net_income_day = _harvest_charged_rate - _harvest_total_implied_day_rate
    _harvest_net_income_annual = _harvest_charged_annual_revenue - _harvest_total_annual_voyage_cost

    ni1, ni2, ni3, ni4 = st.columns(4)
    ni1.metric("Charged rate (NOK/day)", fmt(_harvest_charged_rate) + "/day")
    ni2.metric("Charged annual revenue (NOK)", fmt(_harvest_charged_annual_revenue))
    ni3.metric("Net income (NOK/day)", fmt(_harvest_net_income_day) + "/day")
    ni4.metric("Net income (NOK/year)", fmt(_harvest_net_income_annual))

    st.divider()
    st.markdown("**Treatment voyage cost build-up** (per round trip)")
    st.caption(
        "Same fuel-speed physics as Transport/Other above. Structure: steam "
        "out to a site (8hr), treat (8hr x however many treatments happen "
        "at that visit — 'typical 1-5', via the Repeats column), steam "
        "home (8hr, disinfection happens during this leg at no extra time "
        "or cost), steam out to a second site (8hr), treat again, steam "
        "home again. Unlike Transport/Other, there's no separate "
        "customer-changeover section here — disinfection is already "
        "embedded in every return-to-base leg, not a discrete event "
        "between customers."
    )

    treatment_gcol1, treatment_gcol2, treatment_gcol3, treatment_gcol4, treatment_gcol5 = st.columns(5)
    with treatment_gcol1:
        spot_treatment_ref_speed_kn = stateful_number_input(
            "Reference speed (knots)", min_value=0.1, value=9.0, step=0.5,
            key="spot_treatment_ref_speed", disabled=locked
        )
    with treatment_gcol2:
        spot_treatment_ref_fuel_lhr = stateful_number_input(
            "Fuel rate @ ref. speed, light (L/hr)", min_value=0.0, value=450.0, step=10.0,
            key="spot_treatment_ref_fuel", disabled=locked
        )
    with treatment_gcol3:
        spot_treatment_ref_fuel_lhr_loaded = stateful_number_input(
            "Fuel rate @ ref. speed, loaded (L/hr)", min_value=0.0, value=650.0, step=10.0,
            key="spot_treatment_ref_fuel_loaded", disabled=locked,
            help="Same reference speed as the light rate, but a fully "
                 "loaded vessel simply burns more fuel at that speed — "
                 "applies to phases marked 'Loaded' in the table below."
        )
    with treatment_gcol4:
        spot_treatment_speed_exponent = stateful_number_input(
            "Speed → fuel exponent", min_value=1.0, max_value=5.0, value=1.8, step=0.1,
            key="spot_treatment_speed_exp", disabled=locked,
            help="Fuel rate at any speed = reference rate x (speed / reference speed) ^ this exponent — "
                 "the same exponent applies whether light or loaded, only the reference rate itself differs."
        )
    with treatment_gcol5:
        spot_treatment_fuel_price = stateful_number_input(
            "Fuel price (NOK/liter)", min_value=0.0, value=12.5, step=0.5,
            key="spot_treatment_fuel_price", disabled=locked
        )
    spot_treatment_additional_opex_hr = nok_input(
        "Additional cost per hour in operation (NOK/hr) — all phases",
        "spot_treatment_additional_opex_hr", 600.0,
        key="spot_treatment_additional_opex_input", disabled=locked
    )

    if "spot_treatment_segments" not in st.session_state:
        st.session_state.spot_treatment_segments = [
            {"name": "Steam to site 1", "type": "Steaming", "duration_hr": 8.0, "speed_kn": 9.0, "fuel_rate_lhr": 0.0, "repeats": 1.0, "load": "Light"},
            {"name": "Loading at site 1", "type": "Stationary", "duration_hr": 8.0, "speed_kn": 0.0, "fuel_rate_lhr": 300.0, "repeats": 1.0, "load": "Light"},
            {"name": "Treatment at site 1", "type": "Stationary", "duration_hr": 8.0, "speed_kn": 0.0, "fuel_rate_lhr": 500.0, "repeats": 3.0, "load": "Light"},
            {"name": "Offloading at site 1", "type": "Stationary", "duration_hr": 4.0, "speed_kn": 0.0, "fuel_rate_lhr": 310.0, "repeats": 1.0, "load": "Light"},
            {"name": "Steam home (incl. disinfection)", "type": "Steaming", "duration_hr": 8.0, "speed_kn": 9.0, "fuel_rate_lhr": 0.0, "repeats": 1.0, "load": "Light"},
            {"name": "Steam to site 2", "type": "Steaming", "duration_hr": 8.0, "speed_kn": 9.0, "fuel_rate_lhr": 0.0, "repeats": 1.0, "load": "Light"},
            {"name": "Loading at site 2", "type": "Stationary", "duration_hr": 8.0, "speed_kn": 0.0, "fuel_rate_lhr": 300.0, "repeats": 1.0, "load": "Light"},
            {"name": "Treatment at site 2", "type": "Stationary", "duration_hr": 8.0, "speed_kn": 0.0, "fuel_rate_lhr": 500.0, "repeats": 3.0, "load": "Light"},
            {"name": "Offloading at site 2", "type": "Stationary", "duration_hr": 4.0, "speed_kn": 0.0, "fuel_rate_lhr": 310.0, "repeats": 1.0, "load": "Light"},
            {"name": "Steam home (incl. disinfection)", "type": "Steaming", "duration_hr": 8.0, "speed_kn": 9.0, "fuel_rate_lhr": 0.0, "repeats": 1.0, "load": "Light"},
        ]

    def _add_treatment_segment():
        st.session_state.spot_treatment_segments.append(
            {"name": "New phase", "type": "Stationary", "duration_hr": 0.0, "speed_kn": 0.0, "fuel_rate_lhr": 0.0, "repeats": 1.0, "load": "Light"}
        )

    def _remove_treatment_segment(index):
        st.session_state.spot_treatment_segments.pop(index)

    thdr = st.columns([1.2, 0.8, 0.6, 0.7, 0.6, 0.7, 0.9, 0.7, 0.8, 0.9, 0.9, 0.9, 0.4])
    thdr[0].markdown("**Phase**")
    thdr[1].markdown("**Type**")
    thdr[2].markdown("**Load**")
    thdr[3].markdown("**Duration (hr)**")
    thdr[4].markdown("**Repeats**")
    thdr[5].markdown("**Speed (kn)**")
    thdr[6].markdown("**Fuel rate (L/hr)**")
    thdr[7].markdown("**Total hrs**")
    thdr[8].markdown("**Fuel (L)**")
    thdr[9].markdown("**Fuel cost**")
    thdr[10].markdown("**Add'l opex**")
    thdr[11].markdown("**Total cost**")

    _treatment_total_hours = 0.0
    _treatment_total_fuel_l = 0.0
    _treatment_total_fuel_cost = 0.0
    _treatment_total_additional_opex = 0.0
    _treatment_total_cost = 0.0

    for ti, seg in enumerate(st.session_state.spot_treatment_segments):
        seg.setdefault("load", "Light")
        cols = st.columns([1.2, 0.8, 0.6, 0.7, 0.6, 0.7, 0.9, 0.7, 0.8, 0.9, 0.9, 0.9, 0.4])
        with cols[0]:
            seg["name"] = st.text_input(
                "Phase", value=seg["name"], key=f"treatment_seg_name_{ti}", label_visibility="collapsed", disabled=locked
            )
        with cols[1]:
            seg["type"] = st.selectbox(
                "Type", ["Steaming", "Stationary"],
                index=0 if seg["type"] == "Steaming" else 1,
                key=f"treatment_seg_type_{ti}", label_visibility="collapsed", disabled=locked
            )
        with cols[2]:
            if seg["type"] == "Steaming":
                seg["load"] = st.selectbox(
                    "Load", ["Light", "Loaded"],
                    index=0 if seg["load"] == "Light" else 1,
                    key=f"treatment_seg_load_{ti}", label_visibility="collapsed", disabled=locked
                )
            else:
                st.markdown("<div style='padding-top:8px'>—</div>", unsafe_allow_html=True)
        with cols[3]:
            seg["duration_hr"] = st.number_input(
                "Duration (hr)", min_value=0.0, value=seg["duration_hr"], step=0.5,
                key=f"treatment_seg_duration_{ti}", label_visibility="collapsed", disabled=locked
            )
        with cols[4]:
            seg["repeats"] = st.number_input(
                "Repeats", min_value=0.0, value=seg.get("repeats", 1.0), step=1.0,
                key=f"treatment_seg_repeats_{ti}", label_visibility="collapsed", disabled=locked
            )
        with cols[5]:
            if seg["type"] == "Steaming":
                seg["speed_kn"] = st.number_input(
                    "Speed (kn)", min_value=0.0, value=seg["speed_kn"], step=0.5,
                    key=f"treatment_seg_speed_{ti}", label_visibility="collapsed", disabled=locked
                )
            else:
                st.markdown("<div style='padding-top:8px'>—</div>", unsafe_allow_html=True)
        with cols[6]:
            if seg["type"] == "Steaming":
                _ref_fuel_this = spot_treatment_ref_fuel_lhr_loaded if seg["load"] == "Loaded" else spot_treatment_ref_fuel_lhr
                _fuel_rate_this = _ref_fuel_this * (
                    (seg["speed_kn"] / spot_treatment_ref_speed_kn) ** spot_treatment_speed_exponent
                    if spot_treatment_ref_speed_kn else 0.0
                )
                st.markdown(f"<div style='padding-top:8px'>{fmt(_fuel_rate_this)}</div>", unsafe_allow_html=True)
            else:
                seg["fuel_rate_lhr"] = st.number_input(
                    "Fuel rate (L/hr)", min_value=0.0, value=seg["fuel_rate_lhr"], step=5.0,
                    key=f"treatment_seg_fuelrate_{ti}", label_visibility="collapsed", disabled=locked
                )
                _fuel_rate_this = seg["fuel_rate_lhr"]
        with cols[7]:
            _effective_hours_this = seg["duration_hr"] * seg["repeats"]
            st.markdown(f"<div style='padding-top:8px'>{fmt(_effective_hours_this)}</div>", unsafe_allow_html=True)
        with cols[8]:
            _fuel_this = _fuel_rate_this * _effective_hours_this
            st.markdown(f"<div style='padding-top:8px'>{fmt(_fuel_this)}</div>", unsafe_allow_html=True)
        with cols[9]:
            _fuel_cost_this = _fuel_this * spot_treatment_fuel_price
            st.markdown(f"<div style='padding-top:8px'>{fmt(_fuel_cost_this)}</div>", unsafe_allow_html=True)
        with cols[10]:
            _additional_opex_this = spot_treatment_additional_opex_hr * _effective_hours_this
            st.markdown(f"<div style='padding-top:8px'>{fmt(_additional_opex_this)}</div>", unsafe_allow_html=True)
        with cols[11]:
            _total_cost_this = _fuel_cost_this + _additional_opex_this
            st.markdown(f"<div style='padding-top:8px'>{fmt(_total_cost_this)}</div>", unsafe_allow_html=True)
        with cols[12]:
            st.button("✕", key=f"treatment_seg_remove_{ti}", on_click=_remove_treatment_segment, args=(ti,), disabled=locked)

        _treatment_total_hours += _effective_hours_this
        _treatment_total_fuel_l += _fuel_this
        _treatment_total_fuel_cost += _fuel_cost_this
        _treatment_total_additional_opex += _additional_opex_this
        _treatment_total_cost += _total_cost_this

    st.button("+ Add phase", key="treatment_add_phase", on_click=_add_treatment_segment, disabled=locked)

    tcols = st.columns([1.2, 0.8, 0.6, 0.7, 0.6, 0.7, 0.9, 0.7, 0.8, 0.9, 0.9, 0.9, 0.4])
    tcols[0].markdown("**Round trip total**")
    tcols[7].markdown(f"**{fmt(_treatment_total_hours)}**")
    tcols[8].markdown(f"**{fmt(_treatment_total_fuel_l)}**")
    tcols[9].markdown(f"**{fmt(_treatment_total_fuel_cost)}**")
    tcols[10].markdown(f"**{fmt(_treatment_total_additional_opex)}**")
    tcols[11].markdown(f"**{fmt(_treatment_total_cost)}**")

    _treatment_days_available = next(
        (item["days_per_year"] for item in spot_service_items_current if item["name"] == "Treatment of fish"),
        0.0
    )
    _treatment_hours_available = _treatment_days_available * 24
    _treatment_trips_exact = (_treatment_hours_available / _treatment_total_hours) if _treatment_total_hours else 0.0
    _treatment_annual_voyage_cost = _treatment_total_cost * _treatment_trips_exact
    _treatment_implied_day_rate = (_treatment_annual_voyage_cost / _treatment_days_available) if _treatment_days_available else 0.0

    _treatment_liters_per_year = _treatment_total_fuel_l * _treatment_trips_exact

    sm1, sm2, sm3, sm4, sm5 = st.columns(5)
    sm1.metric("Hours per round trip", fmt(_treatment_total_hours))
    sm2.metric(f"Rounds available ({fmt(_treatment_days_available)}-day window)", f"{_treatment_trips_exact:.1f}", help=f"Over {fmt(_treatment_days_available)} days available (from the Service mix table).")
    sm3.metric("Liters consumed per year", fmt(_treatment_liters_per_year) + " L", help=f"{fmt(_treatment_total_fuel_l)} L per round trip x {_treatment_trips_exact:.2f} rounds/year.")
    sm4.metric("Implied annual voyage cost", fmt(_treatment_annual_voyage_cost))
    sm5.metric("Implied day-rate", fmt(_treatment_implied_day_rate) + "/day")
    st.caption(
        f"Days available for Treatment: {fmt(_treatment_days_available)} (from the Service mix table above) "
        f"x 24 hr = {fmt(_treatment_hours_available)} hours ÷ {fmt(_treatment_total_hours)} hours/round trip "
        f"= {_treatment_trips_exact:.2f} rounds/year."
    )

    # --- combined Variable Voyage opex across all three build-up tools —
    # stored here (self-healing, one-pass-behind) so the Baseline
    # reference panel and Service mix Target line — both further UP the
    # page, rendering before this point — can include it in the gross
    # required-revenue benchmark. Fixed Voyage opex (the shared/
    # unallocated crew-analog line) is tracked separately and stays out
    # of this total. ---
    _spot_variable_voyage_opex_total = (
        _smolt_total_annual_voyage_cost + _harvest_total_annual_voyage_cost + _treatment_annual_voyage_cost
    )
    # Hours per round trip for each of the three build-up-tool-backed
    # services — read by the Service mix table (further UP the page,
    # same one-pass-behind convention) to derive Rounds/year and
    # Payment/round per service.
    _spot_round_hours = {
        "Transport": _smolt_total_hours,
        "Other": _harvest_total_hours,
        "Treatment of fish": _treatment_total_hours,
    }

    def _round_hours_differ(prev, current, tol=0.01):
        """Tolerance-based comparison instead of exact dict equality —
        exact != on a dict of floats built from chained arithmetic
        (duration x fuel rate x speed exponent, summed across phases)
        can flag 'changed' on nearly every rerun purely from floating-
        point representation noise, even when nothing meaningfully
        changed. That was almost certainly causing this self-healing
        trigger to fire repeatedly, indefinitely, on every interaction
        — each firing re-runs the entire financial model. Matches the
        tolerance-based pattern every other self-healing check in this
        app already uses."""
        if prev is None or set(prev.keys()) != set(current.keys()):
            return True
        return any(abs(prev[k] - current[k]) > tol for k in current)

    _spot_variable_voyage_opex_changed = (
        abs(st.session_state.get("_spot_variable_voyage_opex_total", 0.0) - _spot_variable_voyage_opex_total) > 1.0
        or _round_hours_differ(st.session_state.get("_service_round_hours"), _spot_round_hours)
    )
    st.session_state["_spot_variable_voyage_opex_total"] = _spot_variable_voyage_opex_total
    st.session_state["_service_round_hours"] = _spot_round_hours
    _spot_vvo_retry_count = st.session_state.get("_spot_vvo_retry_count", 0)
    if _spot_variable_voyage_opex_changed and _spot_vvo_retry_count < 4:
        st.session_state["_spot_vvo_retry_count"] = _spot_vvo_retry_count + 1
        _request_rerun()
    else:
        st.session_state["_spot_vvo_retry_count"] = 0

    st.markdown("**Net income check — Treatment of fish**")
    st.caption(
        "Links the customer rate set on the Service mix table above "
        "against this build-up's voyage cost, so you can see the "
        "contribution margin move live as you test different rates in "
        "discussion."
    )
    _treatment_charged_rate = _resolved_charged_rate("Treatment of fish")
    _treatment_charged_annual_revenue = _treatment_charged_rate * _treatment_days_available
    _treatment_net_income_day = _treatment_charged_rate - _treatment_implied_day_rate
    _treatment_net_income_annual = _treatment_charged_annual_revenue - _treatment_annual_voyage_cost

    ni1, ni2, ni3, ni4 = st.columns(4)
    ni1.metric("Charged rate (NOK/day)", fmt(_treatment_charged_rate) + "/day")
    ni2.metric("Charged annual revenue (NOK)", fmt(_treatment_charged_annual_revenue))
    ni3.metric("Net income (NOK/day)", fmt(_treatment_net_income_day) + "/day")
    ni4.metric("Net income (NOK/year)", fmt(_treatment_net_income_annual))


# ===========================================================================
# TAB 2 — Lease spread (customer lease, with optional bank financing leg)
# ===========================================================================
with tab_lease:
    st.subheader("Leased equipment")
    st.caption(
        "Lease equipment out to the customer alongside the vessel. Optionally "
        "model bank financing for the equipment as well, to see your funding "
        "margin — but the customer lease works independently, whether or not "
        "the equipment is bank-financed."
    )

    toggle_col1, toggle_col2 = st.columns(2)
    with toggle_col1:
        lease_enabled = stateful_toggle(
            "Include customer lease", value=True, key="lease_enabled", disabled=locked
        )
    with toggle_col2:
        bank_financing_enabled = stateful_toggle(
            "Include bank financing", value=True, key="bank_financing_enabled", disabled=locked
        )

    if not lease_enabled:
        st.info(
            "Customer lease is currently **off**. Turn it on to add the "
            "equipment's lease payment to the Combined TC-rate. Inputs below "
            "are still editable so it's ready whenever you switch it on."
        )

    left, right = st.columns([1, 1.4], gap="large")

    with left:
        st.markdown("**Equipment**")
        st.caption(
            "Defaults mirror the vessel's own financing as a starting "
            "point: 65,000,000 capex, 6.7% bank rate, 84-month term/payback "
            "— all independently editable below. In TC mode, the customer "
            "lease payment below applies as configured, underpinned by the "
            "TC-rate. **In spot mode** (see the Spot market tab), there's "
            "no secured lease contract — the customer lease payment is "
            "cancelled entirely (set to zero), while the equipment still "
            "gets bought and bank-financed exactly as configured below; "
            "spot revenue has to cover that bank payment like any other "
            "cost, with no dedicated lease income offsetting it."
        )
        lease_capex_nok = nok_input(
            "Capex (NOK)", "lease_capex_nok", 65_000_000.0, key="lease_capex_input", disabled=locked
        )

        st.markdown("**Customer lease (income)**")
        lease_yield_pct = stateful_number_input(
            "Lease-out rate, annual (%)", min_value=0.0, value=12.0, step=0.1, key="lease_yield", disabled=locked
        )
        customer_term_months = stateful_number_input(
            "Customer lease term (months)", min_value=1, max_value=120, value=60, step=1,
            key="customer_term", disabled=locked
        )
        lease_payback_months = stateful_number_input(
            "Lease payback structure (months)", min_value=1, max_value=180, value=84, step=1,
            key="lease_payback_months", disabled=locked
        )
        st.caption(
            "The monthly rental rate is calculated as if amortizing the full "
            "capex over this payback period — which can be longer than the "
            "customer contract above (e.g. price on a 7-year/84-month "
            "payback even though the contract is only 5 years/60 months). "
            "This produces a lower monthly rate than a 5-year-only "
            "amortization would, but leaves a residual, unrecovered lease "
            "value at the end of the customer term — shown below — which is "
            "the residual-value risk being taken on."
        )
        lease_opex_monthly_nok = nok_input(
            "Additional opex billed to customer (NOK/month)", "lease_opex_monthly_nok",
            50_000.0, key="lease_opex_input", disabled=locked
        )
        st.caption(
            "This opex is billed to the customer on top of the lease payment, "
            "and passes straight through: it adds to revenue (the TC) and to "
            "vessel opex by the same amount, with no net EBITDA impact. The "
            "lease payment itself (above) has no offsetting cost — it flows "
            "100% to EBITDA."
        )

        if bank_financing_enabled:
            st.markdown("**Bank financing (cost)**")
            bank_rate_pct = stateful_number_input(
                "Bank interest rate, annual (%)", min_value=0.0, value=6.7, step=0.1, key="bank_rate", disabled=locked
            )
            bank_term_months = stateful_number_input(
                "Bank loan term (months)", min_value=1, max_value=120, value=84, step=1,
                key="bank_term", disabled=locked
            )
            lease_equity_instalment_nok = nok_input(
                "Equity instalment (NOK)", "lease_equity_instalment_nok", 0.0,
                key="lease_equity_instalment_input", disabled=locked
            )
            st.caption(
                "Portion of the equipment capex funded by equity rather than "
                "the bank. Default 0 = 100% debt-financed. Bank loan principal "
                "= equipment capex − equity instalment."
            )
        else:
            bank_rate_pct = 0.0
            bank_term_months = int(lease_payback_months)  # depreciation basis matches the payback structure, not the shorter customer contract
            lease_equity_instalment_nok = 0.0

    bank_loan_principal = max(0.0, lease_capex_nok - lease_equity_instalment_nok)

    # --- calculations ---
    # Monthly rental rate is priced on the (potentially longer) payback
    # structure, not the customer contract length — see the caption above.
    lease_monthly_payment = annuity_monthly_payment(
        lease_capex_nok, lease_yield_pct, int(lease_payback_months)
    )

    if bank_financing_enabled:
        bank_monthly_payment = annuity_monthly_payment(
            bank_loan_principal, bank_rate_pct, int(bank_term_months)
        )
        monthly_surplus = lease_monthly_payment - bank_monthly_payment
        tail_months = max(0, int(bank_term_months) - int(customer_term_months))
        total_term_months = max(int(customer_term_months), int(bank_term_months))
        total_surplus_active = monthly_surplus * customer_term_months
        total_shortfall_tail = bank_monthly_payment * tail_months
        net_result = total_surplus_active - total_shortfall_tail
    else:
        bank_monthly_payment = 0.0
        monthly_surplus = 0.0
        tail_months = 0
        total_term_months = int(customer_term_months)
        total_surplus_active = 0.0
        total_shortfall_tail = 0.0
        net_result = 0.0

    # --- full monthly schedules, always computed so later tabs can use them.
    # The customer-side schedule runs only for customer_term_months, but the
    # payment itself was priced on the longer lease_payback_months — leaving
    # a genuine residual (unrecovered) balance at the end; see
    # residual_lease_value below. ---
    lease_schedule_full = amortization_schedule_full(
        lease_capex_nok, lease_yield_pct, int(customer_term_months),
        payment_basis_months=int(lease_payback_months)
    )
    residual_lease_value = lease_schedule_full[-1]["Closing balance"] if lease_schedule_full else 0.0
    bank_schedule_full = amortization_schedule_full(
        bank_loan_principal, bank_rate_pct, int(bank_term_months)
    )

    with right:
        if lease_enabled:
            st.subheader("Cash flow — customer lease")
            lease_annual_amt = lease_monthly_payment * 12
            lease_daily_amt = lease_annual_amt / 365
            m1, m2, m3 = st.columns(3)
            m1.metric("Daily", fmt(lease_daily_amt))
            m2.metric("Monthly", fmt(lease_monthly_payment))
            m3.metric("Annual", fmt(lease_annual_amt))

            if int(lease_payback_months) > int(customer_term_months):
                st.warning(
                    f"**Residual value risk: {fmt(residual_lease_value)}.** The "
                    f"rental rate above is priced on a {int(lease_payback_months)}-month "
                    f"payback, but the customer only pays for "
                    f"{int(customer_term_months)} months — leaving this much "
                    f"unrecovered equipment value at the end of the contract, "
                    f"with no further lease income committed against it "
                    f"(unless renewed, re-leased, or sold)."
                )

            st.subheader("Additional opex (pass-through, billed to customer)")
            lease_opex_annual_amt = lease_opex_monthly_nok * 12
            lease_opex_daily_amt = lease_opex_annual_amt / 365
            m1b, m2b, m3b = st.columns(3)
            m1b.metric("Daily", fmt(lease_opex_daily_amt))
            m2b.metric("Monthly", fmt(lease_opex_monthly_nok))
            m3b.metric("Annual", fmt(lease_opex_annual_amt))

            if bank_financing_enabled:
                st.subheader("Cash flow — bank financing")
                bank_annual_amt = bank_monthly_payment * 12
                bank_daily_amt = bank_annual_amt / 365
                m4, m5, m6 = st.columns(3)
                m4.metric("Daily", fmt(bank_daily_amt))
                m5.metric("Monthly", fmt(bank_monthly_payment))
                m6.metric("Annual", fmt(bank_annual_amt))

                st.subheader("Monthly surplus (lease − bank)")
                surplus_annual_amt = monthly_surplus * 12
                surplus_daily_amt = surplus_annual_amt / 365
                m7, m8, m9 = st.columns(3)
                m7.metric("Daily", fmt(surplus_daily_amt))
                m8.metric("Monthly", fmt(monthly_surplus))
                m9.metric("Annual", fmt(surplus_annual_amt))
            else:
                st.caption(
                    "Bank financing is off — the figures above are the lease payment "
                    "received from the customer, with no offsetting financing cost. "
                    "Turn on 'Include bank financing' to see your funding margin."
                )

            # --- schedule chart ---
            lease_balances = amortization_balances(
                lease_capex_nok, lease_yield_pct, int(customer_term_months),
                payment_basis_months=int(lease_payback_months)
            )

            if bank_financing_enabled:
                bank_balances = amortization_balances(
                    bank_loan_principal, bank_rate_pct, int(bank_term_months)
                )
                # Pad with the residual value (not zero) — the unrecovered
                # lease balance doesn't vanish at the end of the customer
                # term, it simply stops being paid down (see the residual
                # value metric above).
                lease_balances_padded = lease_balances + [residual_lease_value] * (total_term_months - len(lease_balances))
                bank_balances_padded = bank_balances + [0.0] * (total_term_months - len(bank_balances))
                schedule_df = pd.DataFrame(
                    {
                        "Month": list(range(1, total_term_months + 1)),
                        "Lease balance (customer)": lease_balances_padded,
                        "Bank balance (loan)": bank_balances_padded,
                    }
                )
                formatted_line_chart(schedule_df, "Month", ["Lease balance (customer)", "Bank balance (loan)"])
            else:
                schedule_df = pd.DataFrame(
                    {
                        "Month": list(range(1, total_term_months + 1)),
                        "Lease balance (customer)": lease_balances,
                    }
                )
                formatted_line_chart(schedule_df, "Month", ["Lease balance (customer)"])

            if bank_financing_enabled and tail_months > 0:
                st.warning(
                    f"**Tail period:** the bank loan continues for **{tail_months} months** "
                    f"after the customer lease ends, with a monthly shortfall of "
                    f"**{fmt(bank_monthly_payment)}** and no offsetting lease income "
                    f"(unless the equipment is re-leased)."
                )

            if bank_financing_enabled:
                st.subheader("Summary over the full term")
                summary_df = pd.DataFrame(
                    [
                        {
                            "Period": f"Months 1–{int(customer_term_months)} (both active)",
                            "Monthly": monthly_surplus,
                            "Total": total_surplus_active,
                        },
                        {
                            "Period": f"Months {int(customer_term_months)+1}–{total_term_months} (tail)"
                                      if tail_months > 0 else "—",
                            "Monthly": -bank_monthly_payment if tail_months > 0 else None,
                            "Total": -total_shortfall_tail if tail_months > 0 else None,
                        },
                        {
                            "Period": f"Net result, full {total_term_months} months",
                            "Monthly": None,
                            "Total": net_result,
                        },
                    ]
                )
                show_table(summary_df, "Period", width="stretch")

            # --- full monthly amortization tables ---
            st.subheader("Amortization schedule — customer lease")
            lease_schedule_df = pd.DataFrame(lease_schedule_full)
            show_table(lease_schedule_df, "Month", width="stretch", height=300)

            if bank_financing_enabled:
                st.subheader("Amortization schedule — bank financing")
                bank_schedule_df = pd.DataFrame(bank_schedule_full)
                show_table(bank_schedule_df, "Month", width="stretch", height=300)

                st.subheader("Spread payment schedule")
                st.caption(
                    "Lease payment received from the customer, less bank payment "
                    "paid, month by month — including any tail period after the "
                    "customer lease ends but the bank loan continues."
                )
                spread_rows = []
                for month in range(1, total_term_months + 1):
                    lease_payment = (
                        lease_schedule_full[month - 1]["Payment"]
                        if month <= int(customer_term_months) else 0.0
                    )
                    bank_payment = (
                        bank_schedule_full[month - 1]["Payment"]
                        if month <= int(bank_term_months) else 0.0
                    )
                    spread_rows.append({
                        "Month": month,
                        "Lease payment (in)": lease_payment,
                        "Bank payment (out)": bank_payment,
                        "Spread": lease_payment - bank_payment,
                    })
                spread_df = pd.DataFrame(spread_rows)
                show_table(spread_df, "Month", width="stretch", height=300)
        else:
            st.markdown("&nbsp;")
            st.markdown(
                "*Results will appear here once leased equipment is switched on.*"
            )

# ===========================================================================
# TAB 3 — Combined TC-rate (vessel + customer lease payment)
# ===========================================================================
with tab_combined:
    st.subheader("Combined TC-rate — vessel + leased equipment")
    st.caption(
        "Every revenue stream, always shown together — TC and lease income "
        "(TC mode), plus Transport/Other/Treatment spot income (spot mode), "
        "each on its own line. Whichever aren't relevant to the current "
        "mode simply show 0, matching exactly what the Financial "
        "Statements P&L does for these same lines. For the detailed "
        "per-service breakdown behind the spot lines, see the Spot market "
        "tab directly."
    )

    if not lease_enabled:
        st.info(
            "Leased equipment is currently **off** (see the Lease spread tab). "
            "The Lease income and Lease opex lines below will show 0."
        )

    active_lease_monthly = lease_monthly_payment if lease_enabled else 0.0
    lease_annual = active_lease_monthly * 12
    lease_daily = lease_annual / operating_days if operating_days else 0

    active_lease_opex_monthly = lease_opex_monthly_nok if lease_enabled else 0.0
    lease_opex_annual = active_lease_opex_monthly * 12
    lease_opex_daily = lease_opex_annual / operating_days if operating_days else 0

    # --- TC-equivalent baseline, kept ALWAYS on TC-mode figures regardless
    # of the spot toggle — this feeds the Spot market tab's own "required
    # rate" benchmark, so it must never reflect spot economics itself
    # (that would make the benchmark circular, measuring spot against
    # spot). This is intentionally separate from the DISPLAY figures below. ---
    total_tc_annual = vessel_tc_annual + lease_annual + lease_opex_annual
    total_tc_daily = vessel_tc_daily + lease_daily + lease_opex_daily
    total_tc_monthly = vessel_tc_monthly + active_lease_monthly + active_lease_opex_monthly

    # --- stored for the Spot market tab's baseline reference, which runs
    # BEFORE this tab in script order (same one-pass-behind convention used
    # elsewhere in the app, e.g. Tab 1's Sources & Uses guideline) ---
    st.session_state["_combined_tc_daily"] = total_tc_daily

    # --- the six lines: TC/Lease/Lease-opex only nonzero in TC mode
    # (matching the P&L's TC-revenue/Lease-revenue/Pass-through costs
    # lines, all of which are cancelled under spot mode); Transport/Other/
    # Treatment only nonzero in spot mode. Each segment's revenue uses the
    # same pure manual/baseline-gross rate as the corrected P&L calc — no
    # price-list addition. ---
    def _segment_annual(service_name):
        item = next((it for it in spot_service_items_current if it["name"] == service_name), None)
        if item is None:
            return 0.0
        rate = required_gross_rate_at_utilization if item["priced_at_baseline"] else item["rate_nok_day"]
        return rate * item["days_per_year"]

    _tc_income_annual = vessel_tc_annual if not spot_market_enabled else 0.0
    _lease_income_annual = lease_annual if (lease_enabled and not spot_market_enabled) else 0.0
    _lease_opex_annual_line = lease_opex_annual if (lease_enabled and not spot_market_enabled) else 0.0
    _smolt_spot_annual = _segment_annual("Transport") if spot_market_enabled else 0.0
    _harvest_spot_annual = _segment_annual("Other") if spot_market_enabled else 0.0
    _treatment_spot_annual = _segment_annual("Treatment of fish") if spot_market_enabled else 0.0

    _line_items = [
        ("TC income", _tc_income_annual),
        ("Lease income", _lease_income_annual),
        ("Lease-opex pass-through", _lease_opex_annual_line),
        ("Transport spot", _smolt_spot_annual),
        ("Other spot", _harvest_spot_annual),
        ("Treatment spot", _treatment_spot_annual),
    ]

    _display_total_annual = sum(v for _, v in _line_items)
    _display_total_daily = _display_total_annual / operating_days if operating_days else 0
    _display_total_monthly = _display_total_annual / 12

    combined_df = pd.DataFrame(
        [
            {
                "Component": name,
                "Daily": annual / operating_days if operating_days else 0,
                "Monthly": annual / 12,
                "Annual": annual,
            }
            for name, annual in _line_items
        ] + [
            {
                "Component": "TOTAL",
                "Daily": _display_total_daily,
                "Monthly": _display_total_monthly,
                "Annual": _display_total_annual,
            },
        ]
    )
    show_table(combined_df, "Component", width="stretch")

    m1, m2, m3 = st.columns(3)
    m1.metric("Total, daily", fmt(_display_total_daily))
    m2.metric("Total, monthly", fmt(_display_total_monthly))
    m3.metric("Total, annual", fmt(_display_total_annual))

    if not spot_market_enabled:
        st.caption(
            "**Note:** if the customer lease term is shorter than the bank loan term "
            "(see the Lease spread tab), this combined figure reflects the period "
            "while the lease is active. During any tail period, the vessel's TC-rate "
            "no longer includes the equipment lease payment, but the bank loan "
            "obligation continues separately."
        )
    else:
        st.caption(
            "**Note:** in spot mode, the customer lease payment is cancelled "
            "entirely (see the Lease spread tab) — the equipment is still "
            "bank-financed as configured, but spot revenue has to cover "
            "that cost with no dedicated lease income offsetting it."
        )

# ===========================================================================
# TAB 4 — Financial Statements (monthly & annual, horizontal layout)
# ===========================================================================
with tab_financials:
    def _escalation_factor(rate_pct, month):
        """Pure function, no dependencies on anything else in this tab —
        defined first, before anything else, since _get_vessel_revenue
        (defined further down) calls this internally, and gets called
        itself from the Contract summary block before the rest of this
        tab's variables exist. A previous version of this bug (the same
        ordering issue, for spot_segment_revenue_monthly_base) was already
        fixed once; this is the same class of issue for this function."""
        year_number = (month - 1) // 12 + 1  # Year 1 = months 1-12
        escalation_periods = year_number - 1  # 0 in Year 1, 1 in Year 2 (month 13+), ...
        return (1 + rate_pct / 100) ** escalation_periods

    def _utilization_ratio_for_month(month):
        """Utilization for this month's year, relative to Year 1's
        baseline (the top-level Utilization input — Year 1 has no
        separate entry here, avoiding two editable copies of the same
        number). Year 2-12 come from the Year 2-12 planning row (no
        smooth escalation, no staging — each year just holds whatever
        value was typed for it). Years beyond 12 hold at Year 12's value."""
        year = (month - 1) // 12 + 1
        if year <= 1:
            return 1.0
        year_idx = min(year, 12) - 2  # year 2 -> index 0
        active_util = spot_utilization_by_year[year_idx]
        return (active_util / spot_utilization_pct) if spot_utilization_pct else 1.0

    def _fixed_opex_nominal_monthly(month):
        """Fixed Voyage opex, nominal, for this month. Year 1 uses the
        top-level Fixed Voyage opex input directly (no escalation). Years
        2-12 use that specific year's own real (today's money) value from
        the Year 1-12 planning table, run through the SAME single Fixed
        Voyage opex escalator using the standard Year-1-based factor (not
        a new clock starting from when that year's value was set) — e.g.
        a real value typed for Year 3 is multiplied by (1+esc)^2, exactly
        as if it had been the real figure since Year 1. Years beyond 12
        hold at Year 12's real value, still escalating each year."""
        year = (month - 1) // 12 + 1
        if year <= 1:
            real_value = spot_opex_annual_nok
        else:
            year_idx = min(year, 12) - 2  # year 2 -> index 0, clamp beyond 12 at year 12's value
            real_value = spot_fixed_opex_real_by_year[year_idx]
        return real_value * _escalation_factor(spot_opex_escalator_pct, month) / 12

    _segment_escalator_by_year_lookup = {
        "Transport": spot_smolt_escalator_by_year,
        "Other": spot_harvest_escalator_by_year,
        "Treatment of fish": spot_treatment_escalator_by_year,
    }

    def _maintenance_capex_depreciation_for_month(month):
        """Straight-line depreciation on every maintenance capex vintage
        incurred so far — one vintage per year (Tab 1's annual maintenance
        capex, escalated), each depreciating at the VESSEL's own
        depreciation rate (Tab 1) starting from its own year — same
        'each addition gets its own clock' principle used for TC
        contract renewals elsewhere in this model. Applies unconditionally
        in both TC and spot mode, since the vessel needs maintenance
        either way. This is what's missing from simply capitalizing
        maintenance capex into vessel NBV without ever depreciating it
        back down."""
        year = (month - 1) // 12 + 1
        total_dep = 0.0
        for vintage_year in range(1, min(year, 12) + 1):
            vintage_start_month = (vintage_year - 1) * 12 + 1
            if month < vintage_start_month:
                continue
            _maint_this_vintage = annual_maintenance_capex_nok * _escalation_factor(maintenance_escalator_pct, vintage_start_month)
            total_dep += _maint_this_vintage * (depreciation_rate_pct / 100) / 12
        return total_dep

    def _additional_spot_capex_nominal_for_vintage(vintage_year):
        """Additional spot capex's NOMINAL value for a given vintage year
        — the real (today's money) figure typed on the Spot market tab,
        run through the maintenance capex escalator (Tab 1) using the
        standard Year-1-based factor, same treatment as Fixed Voyage
        opex's real-to-nominal conversion. Uses the maintenance escalator
        specifically since Additional spot capex sits 'on top of'
        maintenance capex in the asset register — both are capex
        additions to the same vessel, so both index the same way."""
        if vintage_year < 2:
            return 0.0
        _real_value = spot_additional_capex_by_year[vintage_year - 2]
        _vintage_start_month = (vintage_year - 1) * 12 + 1
        return _real_value * _escalation_factor(maintenance_escalator_pct, _vintage_start_month)

    def _additional_spot_capex_depreciation_for_month(month):
        """Straight-line depreciation on every Additional spot capex
        vintage incurred so far (Spot market tab's Year 2-12 planning
        table, spot mode only) — each vintage's NOMINAL value (real value
        escalated from its own start year) depreciates at its OWN rate
        (Additional spot capex depreciation rate, default 5%/yr =
        20-year life), separate from the vessel/maintenance capex rate,
        since spot-specific capex may genuinely have a different useful
        life."""
        if not spot_market_enabled:
            return 0.0
        year = (month - 1) // 12 + 1
        total_dep = 0.0
        for vintage_year in range(2, min(year, 12) + 1):
            vintage_start_month = (vintage_year - 1) * 12 + 1
            if month < vintage_start_month:
                continue
            _add_capex_this_vintage = _additional_spot_capex_nominal_for_vintage(vintage_year)
            total_dep += _add_capex_this_vintage * (spot_additional_capex_depreciation_pct / 100) / 12
        return total_dep

    def _segment_revenue_multiplier(segment_name, month):
        """Cumulative compounding multiplier for a segment's revenue,
        relative to its Year 1 rate — each year has its OWN escalator %
        (from the Year 1-12 planning table), not one flat rate, so this
        multiplies together every year's own factor from Year 2 through
        the target year. Years beyond 12 stop compounding further (hold
        at Year 12's cumulative multiplier)."""
        year = (month - 1) // 12 + 1
        year_capped = min(year, 12)
        if year_capped <= 1:
            return 1.0
        escalator_list = _segment_escalator_by_year_lookup.get(segment_name)
        if escalator_list is None:
            return 1.0
        multiplier = 1.0
        for y in range(2, year_capped + 1):
            multiplier *= (1 + escalator_list[y - 2] / 100)
        return multiplier


    st.subheader("Financial statements — vessel + leased equipment")
    st.caption(
        "P&L, cash flow, and balance sheet, laid out horizontally (line items "
        "down the rows, periods across the columns) — toggle Monthly / Annual "
        "within each statement below. Combines the vessel (Tab 1) with the "
        "leased equipment (Tab 2), when switched on. The lease payment "
        "received from the customer flows 100% to EBITDA (no offsetting "
        "cost); the additional opex billed to the customer is a pass-through "
        "(net-zero EBITDA impact); equipment financing (leasing company) is "
        "treated the same way as the vessel's own bank debt, shown as a "
        "separate line throughout. Depreciation is split into two lines: "
        "the vessel depreciates straight-line at the rate set on Tab 1 (% "
        "of original capex per year); leased equipment depreciates "
        "straight-line over the same number of months as its own financing "
        "term (Tab 2), reaching zero NBV once that term ends. Maintenance "
        "capex is capitalized (investing outflow, added to the vessel's "
        "balance sheet value) rather than expensed. The cash flow statement "
        "is an EBITDA-down bridge: EBITDA, less working capital build (from "
        "DSO/DPO), less finance cost, less tax, less amortization and "
        "maintenance capex, leaves cash flow for the period."
    )

    st.subheader("Working capital & tax assumptions")
    wc_col1, wc_col2, wc_col3 = st.columns(3)
    with wc_col1:
        dso_days = stateful_number_input(
            "Days sales outstanding (DSO)", min_value=0, value=30, step=1, key="dso_days", disabled=locked
        )
    with wc_col2:
        dpo_days = stateful_number_input(
            "Days payable outstanding (DPO)", min_value=0, value=20, step=1, key="dpo_days", disabled=locked
        )
    with wc_col3:
        tax_rate_pct = stateful_number_input(
            "Corporate tax rate (%)", min_value=0.0, value=0.0, step=0.5, key="tax_rate", disabled=locked
        )
        st.caption("Default 0% — Norwegian Tonnage Tax regime.")

    st.subheader("Escalators (annual, first adjustment in month 13)")
    st.caption(
        "Each rate compounds once per year starting in month 13 (i.e. Year 2 "
        "onward) — months 1–12 stay at the base level."
    )
    esc_col1, esc_col2, esc_col3 = st.columns(3)
    with esc_col1:
        tc_escalator_pct = stateful_number_input(
            "TC revenue escalator (%/yr)", min_value=-100.0, value=2.0, step=0.5, key="tc_escalator", disabled=locked
        )
        lease_escalator_pct = stateful_number_input(
            "Lease payment escalator (%/yr)", min_value=-100.0, value=0.0, step=0.5, key="lease_escalator", disabled=locked
        )
    with esc_col2:
        maintenance_escalator_pct = stateful_number_input(
            "Maintenance capex escalator (%/yr)", min_value=-100.0, value=2.0, step=0.5,
            key="maintenance_escalator", disabled=locked
        )

    st.markdown("**Vessel opex escalators**")
    opex_escalator_pcts = []
    esc_cols = st.columns(min(len(st.session_state.opex_items), 4) or 1)
    for i, item in enumerate(st.session_state.opex_items):
        default_esc = 3.0 if item["name"].strip().lower() == "crewing" else 2.0
        col = esc_cols[i % len(esc_cols)]
        with col:
            esc_pct = stateful_number_input(
                f"{item['name']} (%/yr)", min_value=-100.0, value=default_esc, step=0.5,
                key=f"opex_escalator_{i}", disabled=locked
            )
        opex_escalator_pcts.append(esc_pct)

    st.subheader("Debt refinancing (vessel)")
    st.caption(
        "Refinance the vessel debt at two points, releveraging to a multiple "
        "of the coming year's projected EBITDA — same interest rate (swap + "
        "spread) and amortization profile as the initial debt. Any excess "
        "over the outstanding balance at that point is released as cash."
    )
    refinancing_enabled = stateful_toggle(
        "Enable debt refinancing", value=True, key="refinancing_enabled", disabled=locked
    )
    if refinancing_enabled:
        refi_col1, refi_col2, refi_col3 = st.columns(3)
        with refi_col1:
            refi_year1 = stateful_number_input(
                "First refinancing (year)", min_value=1, value=4, step=1, key="refi_year1", disabled=locked
            )
        with refi_col2:
            refi_year2 = stateful_number_input(
                "Second refinancing (year)", min_value=1, value=8, step=1, key="refi_year2", disabled=locked
            )
        with refi_col3:
            releverage_multiple = stateful_number_input(
                "Releverage multiple (x next year's EBITDA)", min_value=0.0,
                value=float(debt_multiple), step=0.5, key="releverage_multiple", disabled=locked
            )
        st.caption(
            f"E.g. with defaults: refinance in year {int(refi_year1)}, releveraging to "
            f"{releverage_multiple:.1f}x Year {int(refi_year1)+1}'s projected vessel EBITDA; "
            f"same again in year {int(refi_year2)}, releveraging to Year {int(refi_year2)+1}'s."
        )
        refi_trigger_months = {int(refi_year1) * 12 + 1, int(refi_year2) * 12 + 1}
    else:
        refi_trigger_months = set()
        releverage_multiple = debt_multiple

    # --- horizon: covers vessel debt, and if the lease is on, its customer
    #     and bank terms too, so nothing gets truncated ---
    horizon_months = amortization_months
    if lease_enabled:
        horizon_months = max(horizon_months, int(customer_term_months))
        if bank_financing_enabled:
            horizon_months = max(horizon_months, int(bank_term_months))

    monthly_revenue_vessel_base = vessel_tc_monthly  # only meaningful in TC mode — see the info box below

    st.subheader("TC contract schedule")
    if spot_market_enabled:
        st.info(
            "Spot market is active — this contract schedule is now fully "
            "inactive. Treatment, Transport, and Other revenue are each "
            "tracked directly (Service mix table + their own build-up "
            "tools where available), escalating at their own flat rate "
            "indefinitely — see the Spot market tab. This section (and "
            "the TC revenue escalator/renewals below) only drives the "
            "'TC-revenue' line, which applies in TC mode."
        )
    st.caption(
        "Define the initial contract length and up to 3 renewals over the "
        "horizon. Each new contract starts its own escalation clock — first "
        "adjustment 12 months after that contract begins — using the TC "
        "revenue escalator above. Leave a renewal's length at 0 to skip it; "
        "the final segment automatically runs to the end of the horizon. "
        "Optionally, a capex upgrade or downgrade can be applied at the same "
        "month as each renewal (e.g. equipment added or removed alongside a "
        "new charter) — positive adds to the vessel's book value and is an "
        "investing cash outflow; negative reduces book value and returns cash."
    )

    c1a, c1b = st.columns(2)
    with c1a:
        contract1_length = stateful_number_input(
            "Contract 1 length (months)", min_value=1, value=60,
            step=1, key="contract1_length", disabled=locked
        )
    with c1b:
        st.markdown(f"Rate: vessel TC-rate from Tab 1 ({fmt(monthly_revenue_vessel_base)}/month)")

    def _revenue_for_contracts(contracts_so_far, month):
        """Same lookup as _get_vessel_revenue below, but against a partial
        contracts list — used to compute the LTM yardstick using only
        contracts already known at that point in the form."""
        contract = contracts_so_far[-1]
        for c in contracts_so_far:
            if c["start"] <= month < c["start"] + c["length"]:
                contract = c
                break
        months_into = month - contract["start"] + 1
        periods = (months_into - 1) // 12
        factor = (1 + tc_escalator_pct / 100) ** periods
        return contract["base_monthly"] * factor

    tc_contracts = [{
        "start": 1, "length": int(contract1_length),
        "base_monthly": monthly_revenue_vessel_base, "capex_delta": 0.0,
    }]
    next_start = int(contract1_length) + 1

    # Default renewal schedule: Contract 2 after 36 months at 140m/yr (+10m
    # capex), Contract 3 after a further 24 months at 149m/yr (no capex
    # change), Contract 4 at 155m/yr (+5m capex) running to the horizon.
    _contract_defaults = {
        2: {"length": 36, "rate": 140_000_000.0, "capex": 10_000_000.0},
        3: {"length": 24, "rate": 149_000_000.0, "capex": 0.0},
        4: {"length": None, "rate": 155_000_000.0, "capex": 5_000_000.0},
    }

    contract_renewals = []
    for i in (2, 3, 4):
        upcoming_start = next_start

        # --- LTM yardstick: last 12 months' vessel TC-revenue, using only
        #     the contracts already defined up to this renewal ---
        if upcoming_start > 12:
            ltm_start_month = upcoming_start - 12
            ltm_end_month = upcoming_start - 1
            ltm_total = sum(
                _revenue_for_contracts(tc_contracts, m)
                for m in range(ltm_start_month, ltm_end_month + 1)
            )
            st.caption(
                f"📊 **LTM TC-revenue before Contract {i}** (months {ltm_start_month}–{ltm_end_month}): "
                f"{fmt(ltm_total)} — use this as a yardstick for the new rate."
            )
        else:
            st.caption(
                f"Not enough history yet for a full LTM figure before Contract {i} "
                f"(it would start month {upcoming_start})."
            )

        rcol1, rcol2, rcol3 = st.columns(3)
        with rcol1:
            if i < 4:
                length = stateful_number_input(
                    f"Contract {i} length (months) — 0 to skip", min_value=0,
                    value=_contract_defaults[i]["length"],
                    step=1, key=f"contract{i}_length", disabled=locked
                )
            else:
                length = None  # contract 4 always runs to the end of the horizon
                st.markdown("Contract 4 runs to the end of the horizon (if reached)")
        with rcol2:
            new_annual_rate = nok_input(
                f"Contract {i} new TC-rate (NOK/year)", f"contract{i}_rate_nok",
                _contract_defaults[i]["rate"], key=f"contract{i}_rate_input", disabled=locked
            )
        with rcol3:
            capex_delta = nok_input(
                f"Contract {i} capex adjustment (NOK)", f"contract{i}_capex_delta_nok",
                _contract_defaults[i]["capex"], key=f"contract{i}_capex_delta_input", disabled=locked
            )
        contract_renewals.append({"length": length, "new_annual_rate": new_annual_rate, "capex_delta": capex_delta})

        # extend the known contracts list so the NEXT renewal's LTM sees this one
        if i < 4:
            length_int = int(length)
            if length_int > 0 and upcoming_start <= horizon_months:
                new_monthly = new_annual_rate / 12
                tc_contracts.append({
                    "start": upcoming_start, "length": length_int,
                    "base_monthly": new_monthly, "capex_delta": capex_delta,
                })
                next_start = upcoming_start + length_int
        else:
            remaining = horizon_months - (upcoming_start - 1)
            if remaining > 0:
                new_monthly = new_annual_rate / 12
                tc_contracts.append({
                    "start": upcoming_start, "length": remaining,
                    "base_monthly": new_monthly, "capex_delta": capex_delta,
                })

    # Capex adjustment applied exactly once, at each renewal's start month
    # (contract 1 never has one — it's the vessel's original capex).
    capex_delta_by_month = {
        c["start"]: c["capex_delta"] for c in tc_contracts[1:] if c["capex_delta"] != 0.0
    }

    # --- per-segment revenue basis (Treatment/Transport/Other), kept
    # explicit per segment so the P&L can show each segment's revenue on
    # its own line. Year-1 day-rate resolves to the customer-mix
    # blended rate (Spot market tab) when that's enabled for a service;
    # otherwise the pure manual (or baseline-gross) rate, unchanged from
    # before. Defined here (early) because _get_vessel_revenue below
    # needs it, and that function gets called from the Contract summary
    # block further down — before the main monthly loop even runs. ---
    spot_segment_revenue_monthly_base = []
    for s_idx, item in enumerate(spot_service_items_current):
        if _customer_mix_enabled(item["name"]):
            _seg_rate = _customer_mix_blended_rate(item["name"], 1)
        else:
            _seg_rate = required_gross_rate_at_utilization if item["priced_at_baseline"] else item["rate_nok_day"]
        spot_segment_revenue_monthly_base.append({
            "name": item["name"],
            "monthly_base": _seg_rate * item["days_per_year"] / 12,
            "day_rate_year1": _seg_rate,
            "days_per_year": item["days_per_year"],
            "escalator_pct": item["escalator_pct"],
        })

    def _service_day_rate_for_year(service_name, year):
        """Year-`year` day-rate for a spot service, before seasonality —
        the Year-1 base rate (either the customer-mix blend, or the
        simple manual/baseline rate) escalated via that service's own
        Year 1-12 indexation table (the aggregate 'Revenue indexation
        by segment' planning table above). Customer mix no longer
        carries its own per-customer escalation — the aggregate
        segment escalator is the single source of Year 2-12 growth for
        both modes, so this one function now covers both consistently."""
        if _customer_mix_enabled(service_name):
            base = _customer_mix_blended_rate(service_name, 1)
        else:
            base = next(
                (x["day_rate_year1"] for x in spot_segment_revenue_monthly_base if x["name"] == service_name), 0.0
            )
        year_capped = min(year, 12)
        escalator_list = _segment_escalator_by_year_lookup.get(service_name)
        multiplier = 1.0
        if escalator_list and year_capped > 1:
            for y in range(2, year_capped + 1):
                multiplier *= (1 + escalator_list[y - 2] / 100)
        return base * multiplier

    def _service_monthly_revenue(service_name, month):
        """Monthly revenue for one spot service — folds in the Year
        1-12 utilization schedule and the Year-y day-rate (manual/
        indexed or customer-mix blended, from _service_day_rate_for_year
        above). If the Weekly activity calendar is enabled (Spot market
        tab), revenue for this service comes from summing whichever of
        that calendar month's constituent weeks are actually assigned to
        this service (a month can now mix activities across its weeks —
        e.g. 2 weeks Treatment + 1 idle week — each service only counts
        its own weeks, so nothing double-counts and nothing is lost).
        Otherwise (calendar off), days spread flat across all 12 months,
        exactly as before."""
        year = (month - 1) // 12 + 1
        util_factor = _utilization_ratio_for_month(month)
        day_rate_this_year = _service_day_rate_for_year(service_name, year)

        if _calendar_enabled():
            month_of_year_idx = (month - 1) % 12
            weeks_in_month = _month_to_weeks_map().get(month_of_year_idx, [])
            calendar_weeks = st.session_state.get("spot_calendar_weeks", [])
            days_per_week_equiv = operating_days / 52
            days_this_month = 0.0
            for w in weeks_in_month:
                if w - 1 < len(calendar_weeks):
                    row = calendar_weeks[w - 1]
                    if row.get("activity") == service_name:
                        days_this_month += days_per_week_equiv * (row["occupancy_pct"] / 100)
            days_this_month *= util_factor
            return days_this_month * day_rate_this_year

        days_per_year_base = next(
            (x["days_per_year"] for x in spot_segment_revenue_monthly_base if x["name"] == service_name), 0.0
        )
        annual_days_this_year = days_per_year_base * util_factor
        monthly_days = annual_days_this_year / 12
        return monthly_days * day_rate_this_year

    def _get_vessel_revenue(month):
        """Base monthly revenue for this month. TC mode: from the active
        contract (escalated from that contract's own start — first
        adjustment 12 months in). Spot mode: sum of Treatment/Transport/
        Other revenue via _service_monthly_revenue above (each
        service's own escalation/customer-mix/seasonality already
        folded in) — used here as a convenient 'total vessel revenue'
        figure for refinancing and terminal-value projections; the main
        monthly loop tracks each segment individually and explicitly
        for the P&L."""
        if spot_market_enabled:
            return sum(
                _service_monthly_revenue(item["name"], month)
                for item in spot_segment_revenue_monthly_base
            )
        return _revenue_for_contracts(tc_contracts, month)

    def _get_vessel_opex(month):
        """Vessel opex for this month: Tab 1's fixed crewing/opex items
        (applies every calendar day, unconditional on utilization — crew
        salaries don't stop when idle, and the vessel needs a crew
        whether trading TC or spot), plus, in spot mode, Fixed Voyage
        opex on top (that year's own real value, escalated the standard
        way) — Fixed Voyage opex is additional overhead specific to
        running the spot-trade business, not a replacement for crewing."""
        _crew_and_other_opex = sum(
            item["monthly"] * _escalation_factor(item["escalator_pct"], month)
            for item in opex_line_items_base
        )
        if spot_market_enabled:
            return _crew_and_other_opex + _fixed_opex_nominal_monthly(month)
        return _crew_and_other_opex

    if len(tc_contracts) > 1:
        st.markdown("**Contract summary** (annualized rate, and uplift vs. the TC-rate just before renewal)")
        for i, c in enumerate(tc_contracts):
            annual_rate = c["base_monthly"] * 12
            if i == 0:
                st.markdown(
                    f"- Contract 1 (months {c['start']}–{c['start']+c['length']-1}): "
                    f"{fmt(c['base_monthly'])}/month · {fmt(annual_rate)}/year"
                )
            else:
                prev_rate_monthly = _get_vessel_revenue(c["start"] - 1)
                uplift_pct = (c["base_monthly"] / prev_rate_monthly - 1) * 100 if prev_rate_monthly else 0.0
                end_label = c["start"] + c["length"] - 1
                st.markdown(
                    f"- Contract {i+1} (months {c['start']}–{end_label}): "
                    f"{fmt(c['base_monthly'])}/month · {fmt(annual_rate)}/year "
                    f"— uplift vs. prior rate: **{uplift_pct:+.1f}%**"
                )

    st.divider()
    st.subheader("Lease contract schedule (leased equipment)")
    st.caption(
        "The equipment (already purchased — see the Lease spread tab) can "
        "be re-leased at a new price and its own indexation once the "
        "current contract ends — up to 3 further terms, same mechanism as "
        "the TC contract schedule above. Contract 1 is the customer lease "
        "already defined on the Lease spread tab (length, rate, and "
        "escalator all sourced from there — see the Escalators section "
        "above for its escalator). Leave any renewal's length at 0 to "
        "skip it — unlike the TC schedule, the lease does **not** "
        "auto-continue at the last rate once every configured contract "
        "has ended; it simply goes off-lease (zero revenue) from that "
        "point, exactly as long as the contracts you've actually set. "
        "This only affects the 'Lease-revenue' line — financing (Tab 2's "
        "bank loan) is unaffected and keeps running on its own separate "
        "schedule."
    )

    lease_contracts = []
    if lease_enabled:
        lease_contracts.append({
            "start": 1, "length": int(customer_term_months),
            "base_monthly": lease_monthly_payment, "escalator_pct": lease_escalator_pct,
        })
    lease_next_start = int(customer_term_months) + 1

    _lease_contract_defaults = {
        2: {"length": 0, "rate": lease_monthly_payment * 12, "escalator": 0.0},
        3: {"length": 0, "rate": lease_monthly_payment * 12, "escalator": 0.0},
        4: {"length": 0, "rate": lease_monthly_payment * 12, "escalator": 0.0},
    }

    if lease_enabled:
        for i in (2, 3, 4):
            lrcol1, lrcol2, lrcol3 = st.columns(3)
            with lrcol1:
                lease_renewal_length = stateful_number_input(
                    f"Lease contract {i} length (months) — 0 to skip", min_value=0,
                    value=_lease_contract_defaults[i]["length"],
                    step=1, key=f"lease_contract{i}_length", disabled=locked
                )
            with lrcol2:
                lease_renewal_rate = nok_input(
                    f"Lease contract {i} new rate (NOK/year)", f"lease_contract{i}_rate_nok",
                    _lease_contract_defaults[i]["rate"], key=f"lease_contract{i}_rate_input", disabled=locked
                )
            with lrcol3:
                lease_renewal_escalator = stateful_number_input(
                    f"Lease contract {i} escalator (%/yr)", min_value=-100.0,
                    value=_lease_contract_defaults[i]["escalator"], step=0.5,
                    key=f"lease_contract{i}_escalator", disabled=locked
                )

            _length_int = int(lease_renewal_length)
            if _length_int > 0 and lease_next_start <= horizon_months:
                lease_contracts.append({
                    "start": lease_next_start, "length": _length_int,
                    "base_monthly": lease_renewal_rate / 12, "escalator_pct": lease_renewal_escalator,
                })
                lease_next_start += _length_int

    def _revenue_for_lease_contracts(month):
        """Same lookup pattern as _revenue_for_contracts (TC), but against
        the lease_contracts schedule — each lease contract escalates from
        its own start month, at its own escalator."""
        for c in lease_contracts:
            if c["start"] <= month < c["start"] + c["length"]:
                months_into = month - c["start"] + 1
                periods = (months_into - 1) // 12
                factor = (1 + c["escalator_pct"] / 100) ** periods
                return c["base_monthly"] * factor
        return 0.0

    def _lease_contract_active(month):
        return any(c["start"] <= month < c["start"] + c["length"] for c in lease_contracts)

    if lease_enabled and len(lease_contracts) > 1:
        st.markdown("**Lease contract summary** (annualized rate, and uplift vs. the rate just before renewal)")
        for i, c in enumerate(lease_contracts):
            annual_rate = c["base_monthly"] * 12
            if i == 0:
                st.markdown(
                    f"- Contract 1 (months {c['start']}–{c['start']+c['length']-1}): "
                    f"{fmt(c['base_monthly'])}/month · {fmt(annual_rate)}/year"
                )
            else:
                prev_rate_monthly = _revenue_for_lease_contracts(c["start"] - 1)
                uplift_pct = (c["base_monthly"] / prev_rate_monthly - 1) * 100 if prev_rate_monthly else 0.0
                end_label = c["start"] + c["length"] - 1
                st.markdown(
                    f"- Contract {i+1} (months {c['start']}–{end_label}): "
                    f"{fmt(c['base_monthly'])}/month · {fmt(annual_rate)}/year "
                    f"— uplift vs. prior rate: **{uplift_pct:+.1f}%**"
                )

    monthly_opex_vessel_base = opex_total / 12
    monthly_vessel_depreciation = (capex_nok * (depreciation_rate_pct / 100)) / 12
    monthly_maintenance_base = annual_maintenance_capex_nok / 12

    opex_line_items_base = [
        {"name": item["name"], "monthly": item["value_nok"] / 12, "escalator_pct": esc}
        for item, esc in zip(st.session_state.opex_items, opex_escalator_pcts)
    ]

    # --- spot market basis: voyage costs/recovery scale with working days
    # (idle days don't incur fuel/port fees), unchanged regardless of the
    # revenue-side simplification above. Vessel opex (crewing — Tab 1)
    # stays fixed on the full operating_days basis in TC mode; in spot
    # mode it's a fixed ANNUAL budget (spot_opex_annual_nok) spread evenly
    # across months — NOT tied to working days here, since the annual
    # figure is the true fixed input; only its day-rate equivalent (shown
    # on the Spot tab) derives from working days. ---
    spot_working_days_annual = sum(item["days_per_year"] for item in spot_service_items_current)
    spot_opex_monthly_base = spot_opex_annual_nok / 12

    # --- per-segment voyage cost basis. Treatment, Transport, and Other all
    # now source their direct voyage cost from their own phase-by-phase
    # build-up tools above (fuel physics, cleaning, customer changeover) —
    # all three now fully self-contained, no generic fallback table needed
    # anymore. ---
    _treatment_idx = next(
        (i for i, it in enumerate(spot_service_items_current) if it["name"] == "Treatment of fish"), None
    )
    _smolt_idx_lookup = next(
        (i for i, it in enumerate(spot_service_items_current) if it["name"] == "Transport"), None
    )
    _harvest_idx_lookup = next(
        (i for i, it in enumerate(spot_service_items_current) if it["name"] == "Other"), None
    )

    spot_treatment_cost_monthly_base = _treatment_annual_voyage_cost / 12
    spot_treatment_cost_escalator_pct = spot_variable_opex_escalator_pct
    spot_smolt_cost_monthly_base = _smolt_total_annual_voyage_cost / 12
    spot_smolt_cost_escalator_pct = spot_variable_opex_escalator_pct
    spot_harvest_cost_monthly_base = _harvest_total_annual_voyage_cost / 12
    spot_harvest_cost_escalator_pct = spot_variable_opex_escalator_pct

    def _service_monthly_voyage_cost(service_name, month, annual_voyage_cost_year1, cost_escalator_pct):
        """Escalates a service's Year-1 annual voyage cost to this
        month's year, then distributes it across the year the same way
        revenue is distributed: via the Weekly activity calendar when
        enabled (cost only lands in the specific weeks this service
        actually runs, weighted by each week's day-count, aggregated up
        to whichever month those weeks fall in — so a bigger week
        carries a proportionally bigger share of the annual cost), or
        spread flat across all 12 months otherwise (unchanged default
        behaviour)."""
        util_factor = _utilization_ratio_for_month(month)
        annual_cost_this_year = annual_voyage_cost_year1 * _escalation_factor(cost_escalator_pct, month) * util_factor

        if _calendar_enabled():
            month_of_year_idx = (month - 1) % 12
            weeks_in_month = _month_to_weeks_map().get(month_of_year_idx, [])
            calendar_weeks = st.session_state.get("spot_calendar_weeks", [])
            days_per_week_equiv = operating_days / 52
            days_this_month = sum(
                days_per_week_equiv * (calendar_weeks[w - 1]["occupancy_pct"] / 100)
                for w in weeks_in_month
                if w - 1 < len(calendar_weeks) and calendar_weeks[w - 1].get("activity") == service_name
            )
            total_calendar_days_this_service = sum(
                days_per_week_equiv * (row["occupancy_pct"] / 100)
                for row in calendar_weeks if row.get("activity") == service_name
            ) or 1.0
            return annual_cost_this_year * (days_this_month / total_calendar_days_this_service)

        return annual_cost_this_year / 12

    equipment_capex = lease_capex_nok if lease_enabled else 0.0
    equipment_debt_initial = bank_loan_principal if (lease_enabled and bank_financing_enabled) else 0.0
    equipment_equity_initial = equipment_capex - equipment_debt_initial

    # Equipment is depreciated straight-line over the same number of months
    # as its financing term (bank_term_months) — whether or not bank
    # financing is actually switched on, since bank_term_months already
    # falls back to the customer lease term in that case (see Tab 2).
    equipment_depreciation_months = int(bank_term_months) if (lease_enabled and equipment_capex > 0) else 0
    monthly_equipment_depreciation = (
        equipment_capex / equipment_depreciation_months if equipment_depreciation_months else 0.0
    )

    if not spot_market_enabled and monthly_revenue_vessel_base < (monthly_opex_vessel_base + debt_schedule[0]["Monthly finance cost"]):
        st.warning(
            "**Note:** the vessel TC-rate (Tab 1) is built from required EBITDA + "
            "vessel opex only — it does not include debt finance cost or tax. "
            "Depending on your inputs, monthly revenue may not fully cover opex + "
            "finance cost + tax; check the P&L for negative net income if that "
            "matters for your analysis."
        )
    elif spot_market_enabled:
        st.caption(
            "ℹ️ Spot market is active — vessel revenue and voyage costs below "
            "come from the Spot market tab, not the TC-rate on Tab 1 (Tab 1's "
            "capex/debt sizing still applies)."
        )

    def _row_or_zero(schedule, month, term_months):
        if month <= term_months and month <= len(schedule):
            return schedule[month - 1]
        return {"Finance cost": 0.0, "Amortization": 0.0, "Closing balance": 0.0}

    # --- operational funding, decided on Tab 1's Sources & Uses (runs earlier
    #     in the script, so this is already set for the current pass) ---
    _op_funding = st.session_state.get("_operational_funding", {"equity": 0.0, "debt": 0.0})
    operational_equity_nok = _op_funding["equity"]
    operational_debt_nok = _op_funding["debt"]

    vessel_equity_initial = capex_nok - debt_nok

    def _run_monthly_model(op_equity_nok, op_debt_nok):
        """Runs the full monthly P&L / cash flow / balance sheet model for a
        given operational-funding split. Pulled into a function so it can be
        run twice: once with whatever funding the user actually chose (for
        the real statements and IRR), and once forced to zero funding, purely
        to read off a stable, funding-independent guideline — see the call
        site below for why that removes the circularity entirely rather than
        iterating toward it."""
        pnl_rows_ = []
        cf_rows_ = []
        bs_rows_ = []

        cumulative_cash = op_equity_nok + op_debt_nok  # injected upfront, month 0
        cumulative_vessel_depreciation = 0.0
        cumulative_equipment_depreciation = 0.0
        cumulative_maintenance_capex = 0.0
        cumulative_additional_spot_capex = 0.0
        cumulative_maintenance_capex_depreciation = 0.0
        cumulative_additional_capex_depreciation = 0.0
        cumulative_capex_adjustment = 0.0
        equity = vessel_equity_initial + equipment_equity_initial + op_equity_nok

        bs_rows_.append({
            "Month": 0,
            "Vessel (NBV)": capex_nok,
            "Leased equipment (NBV)": equipment_capex,
            "Accounts receivable": 0.0,
            "Cash": cumulative_cash,
            "Total assets": capex_nok + equipment_capex + cumulative_cash,
            "Debt — vessel (bank)": debt_nok,
            "Debt — equipment (leasing company)": equipment_debt_initial,
            "Debt — operational funding": op_debt_nok,
            "Accounts payable": 0.0,
            "Equity": equity,
            "Total liabilities + equity": debt_nok + equipment_debt_initial + op_debt_nok + equity,
        })

        # --- month 0 on the cash flow statement: the operational funding
        # injected upfront (equity + debt), before month 1's operations begin.
        # Shown here so it's possible to see, side by side, the opening cash
        # available going into month 1 — and to compare "Cash flow for the
        # period" month-by-month with and without covering the liquidity gap,
        # since none of the other flow lines depend on this injection (see
        # "Cash flow for the period" below, month 1 onward, which is
        # independent of operational funding whenever it's debt-free).
        cf_rows_.append({
            "Month": 0,
            "EBITDA": 0.0,
            "Working capital change": 0.0,
            "Finance cost — vessel (bank)": 0.0,
            "Finance cost — equipment (leasing company)": 0.0,
            "Finance cost — operational funding": 0.0,
            "Tax": 0.0,
            "Amortization — vessel (bank)": 0.0,
            "Amortization — equipment (leasing company)": 0.0,
            "Amortization — operational funding": 0.0,
            "Maintenance capex": 0.0,
            "Additional spot capex": 0.0,
            "Capex adjustment (vessel upgrade/downgrade)": 0.0,
            "Refinancing proceeds (vessel)": 0.0,
            "Operational funding injected (equity + debt)": op_equity_nok + op_debt_nok,
            "Cash flow for the period": op_equity_nok + op_debt_nok,
            "Cash balance": cumulative_cash,
        })

        prev_nwc = 0.0
        vessel_debt_balance = debt_nok
        vessel_quarterly_amort = quarterly_amortization_nok
        vessel_cycle_month = 0
        vessel_monthly_rate = (finance_cost_rate_pct / 100) / 12

        operational_debt_balance = op_debt_nok
        operational_quarterly_amort = op_debt_nok / (amortization_years * 4) if amortization_years else 0.0
        operational_cycle_month = 0

        for month in range(1, horizon_months + 1):
            vessel_cycle_month += 1
            operational_cycle_month += 1
            refinancing_proceeds_this_month = 0.0

            if month in refi_trigger_months:
                # trigger month is the first month of the "coming year" itself
                # (e.g. refi_year=4 -> trigger month 49 -> that's the first month
                # of Year 5), so no extra +1 is needed here.
                target_month_for_projection = month
                projected_monthly_revenue = _get_vessel_revenue(target_month_for_projection)
                projected_monthly_opex = _get_vessel_opex(target_month_for_projection)
                if spot_market_enabled:
                    # _get_vessel_revenue already sums Treatment/Transport/Other
                    # revenue (incl. the per-year utilization factor); add
                    # all three segments' own build-up-tool voyage costs
                    # here, with the same factor applied (crew/vessel opex
                    # above is the shared line only, unaffected).
                    _util_factor_proj = _utilization_ratio_for_month(target_month_for_projection)
                    projected_monthly_opex += (
                        spot_treatment_cost_monthly_base * _escalation_factor(spot_treatment_cost_escalator_pct, target_month_for_projection) * _util_factor_proj
                        + spot_smolt_cost_monthly_base * _escalation_factor(spot_smolt_cost_escalator_pct, target_month_for_projection) * _util_factor_proj
                        + spot_harvest_cost_monthly_base * _escalation_factor(spot_harvest_cost_escalator_pct, target_month_for_projection) * _util_factor_proj
                    )
                projected_annual_ebitda_vessel = (projected_monthly_revenue - projected_monthly_opex) * 12
                new_principal = releverage_multiple * projected_annual_ebitda_vessel
                refinancing_proceeds_this_month = new_principal - vessel_debt_balance
                vessel_debt_balance = new_principal
                vessel_quarterly_amort = new_principal / (amortization_years * 4)
                vessel_cycle_month = 1  # this month is month 1 of the new amortization cycle

            vessel_opening_balance = vessel_debt_balance
            vessel_finance_cost = vessel_opening_balance * vessel_monthly_rate
            vessel_is_quarter_end = (vessel_cycle_month % 3 == 0)
            vessel_amortization = vessel_quarterly_amort if vessel_is_quarter_end else 0.0
            vessel_debt_balance = vessel_opening_balance - vessel_amortization
            vessel_debt_closing = vessel_debt_balance

            # --- operational funding debt tranche (same rate/terms as vessel debt) ---
            operational_opening_balance = operational_debt_balance
            operational_finance_cost = operational_opening_balance * vessel_monthly_rate
            operational_is_quarter_end = (operational_cycle_month % 3 == 0)
            operational_amortization = operational_quarterly_amort if operational_is_quarter_end else 0.0
            operational_debt_balance = operational_opening_balance - operational_amortization
            operational_debt_closing = operational_debt_balance

            # --- escalated revenue & opex for this month ---
            # TC-mode revenue uses the existing contract mechanism; in spot
            # mode, this line goes to zero and Treatment/Transport/Other
            # revenue are tracked explicitly below instead, via
            # _service_monthly_revenue (folds in each service's own
            # escalation, customer mix, and seasonality — Spot market tab).
            if spot_market_enabled:
                _util_factor_this_month = _utilization_ratio_for_month(month)
                monthly_revenue_vessel = 0.0
                monthly_treatment_revenue = (
                    _service_monthly_revenue("Treatment of fish", month) if _treatment_idx is not None else 0.0
                )
                monthly_smolt_revenue = (
                    _service_monthly_revenue("Transport", month) if _smolt_idx_lookup is not None else 0.0
                )
                monthly_harvest_revenue = (
                    _service_monthly_revenue("Other", month) if _harvest_idx_lookup is not None else 0.0
                )
                monthly_treatment_voyage_cost = _service_monthly_voyage_cost(
                    "Treatment of fish", month, _treatment_annual_voyage_cost, spot_treatment_cost_escalator_pct
                )
                monthly_smolt_voyage_cost = _service_monthly_voyage_cost(
                    "Transport", month, _smolt_total_annual_voyage_cost, spot_smolt_cost_escalator_pct
                )
                monthly_harvest_voyage_cost = _service_monthly_voyage_cost(
                    "Other", month, _harvest_total_annual_voyage_cost, spot_harvest_cost_escalator_pct
                )
            else:
                monthly_revenue_vessel = _get_vessel_revenue(month)
                monthly_treatment_revenue = 0.0
                monthly_smolt_revenue = 0.0
                monthly_harvest_revenue = 0.0
                monthly_treatment_voyage_cost = 0.0
                monthly_smolt_voyage_cost = 0.0
                monthly_harvest_voyage_cost = 0.0

            # In spot mode, the equipment's customer lease payment is
            # cancelled — there's no secured lease contract underpinning
            # it under spot trading. The equipment still gets bought and
            # bank-financed exactly as configured (finance cost/
            # amortization below are computed separately, from
            # bank_schedule_full, and are unaffected by this); only the
            # lease REVENUE side goes to zero, so spot revenue has to
            # cover that cost like everything else. In TC mode, revenue
            # follows the Lease contract schedule above (Contract 1 = the
            # customer lease from Tab 2; renewals 2-4 as configured).
            if lease_enabled and not spot_market_enabled and _lease_contract_active(month):
                lease_revenue_this_month = _revenue_for_lease_contracts(month)
                lease_opex_this_month = lease_opex_monthly_nok  # pass-through, not escalated
            else:
                lease_revenue_this_month = 0.0
                lease_opex_this_month = 0.0

            # --- equipment depreciation: straight-line, stops once the
            # equipment financing term (equipment_depreciation_months) has
            # elapsed, leaving NBV at exactly 0 rather than going negative ---
            if equipment_depreciation_months and month <= equipment_depreciation_months:
                equipment_depreciation_this_month = monthly_equipment_depreciation
            else:
                equipment_depreciation_this_month = 0.0

            escalated_opex_items = []
            for item in opex_line_items_base:
                factor = _escalation_factor(item["escalator_pct"], month)
                escalated_value = item["monthly"] * factor
                escalated_opex_items.append({"name": item["name"], "value": escalated_value})

            if spot_market_enabled:
                _fixed_voyage_opex_this_month = _fixed_opex_nominal_monthly(month)
                spot_vessel_opex_this_month = _fixed_voyage_opex_this_month  # Fixed Voyage opex specifically, for its own P&L line
                monthly_opex_vessel = _fixed_voyage_opex_this_month + sum(x["value"] for x in escalated_opex_items)
            else:
                monthly_opex_vessel = sum(x["value"] for x in escalated_opex_items)
                spot_vessel_opex_this_month = 0.0

            # --- spot market: Treatment/Transport/Other voyage costs, each
            # from their own build-up tool — all three now fully
            # self-contained, no generic recovery mechanic needed anymore. ---
            if spot_market_enabled:
                monthly_opex_vessel += monthly_treatment_voyage_cost + monthly_smolt_voyage_cost + monthly_harvest_voyage_cost

            maintenance_factor = _escalation_factor(maintenance_escalator_pct, month)
            monthly_maintenance = monthly_maintenance_base * maintenance_factor

            # --- additional spot capex: spread evenly across the year it's
            # incurred, same convention as maintenance capex above. Uses
            # the NOMINAL value (real value escalated from its own start
            # year via the maintenance escalator) — the real figure typed
            # on the Spot market tab isn't what actually hits cash/the
            # balance sheet, its escalated nominal equivalent is. Spot
            # mode only — this represents the extra wear spot trading
            # puts on the vessel beyond what a steady TC charter would. ---
            if spot_market_enabled:
                _year_now = (month - 1) // 12 + 1
                if _year_now >= 2:
                    monthly_additional_spot_capex = _additional_spot_capex_nominal_for_vintage(min(_year_now, 12)) / 12
                else:
                    monthly_additional_spot_capex = 0.0
            else:
                monthly_additional_spot_capex = 0.0

            # --- asset register depreciation: maintenance capex and
            # Additional spot capex each depreciate on their OWN schedule
            # (different rates/useful lives), starting from their own
            # vintage year — see each function's docstring. This is IN
            # ADDITION to the vessel's own depreciation on original capex
            # (monthly_vessel_depreciation, computed once above). ---
            monthly_maintenance_capex_depreciation = _maintenance_capex_depreciation_for_month(month)
            monthly_additional_capex_depreciation = _additional_spot_capex_depreciation_for_month(month)

            if lease_enabled and bank_financing_enabled:
                eq_debt_row = _row_or_zero(bank_schedule_full, month, int(bank_term_months))
                equipment_finance_cost = eq_debt_row["Finance cost"]
                equipment_amortization = eq_debt_row["Amortization"]
                equipment_debt_closing = eq_debt_row["Closing balance"]
            else:
                equipment_finance_cost = 0.0
                equipment_amortization = 0.0
                equipment_debt_closing = 0.0

            # --- combined P&L ---
            # monthly_revenue_vessel is 0 in spot mode (Treatment/Transport/
            # Other carry the actual revenue instead); EBITDA stays a
            # single combined figure (no per-segment EBITDA split — direct
            # voyage costs are netted per segment on the Spot market tab's
            # own 'Net income check' sections instead, since crew opex
            # stays shared/unallocated here).
            revenue = (
                monthly_revenue_vessel + monthly_treatment_revenue + monthly_smolt_revenue + monthly_harvest_revenue
                + lease_revenue_this_month + lease_opex_this_month
            )
            ebitda_vessel = (
                monthly_revenue_vessel + monthly_treatment_revenue + monthly_smolt_revenue + monthly_harvest_revenue
                - monthly_opex_vessel
            )
            ebitda_equipment = lease_revenue_this_month  # pass-through opex nets to zero
            ebitda = ebitda_vessel + ebitda_equipment
            ebit = (
                ebitda - monthly_vessel_depreciation - monthly_maintenance_capex_depreciation
                - monthly_additional_capex_depreciation - equipment_depreciation_this_month
            )
            finance_cost_total = vessel_finance_cost + equipment_finance_cost + operational_finance_cost
            ebt = ebit - finance_cost_total
            tax = ebt * (tax_rate_pct / 100)
            net_income = ebt - tax

            pnl_row = {"Month": month}
            pnl_row["TC-revenue"] = monthly_revenue_vessel
            pnl_row["Lease-revenue"] = lease_revenue_this_month
            if spot_market_enabled:
                pnl_row["Transport revenue"] = monthly_smolt_revenue
                pnl_row["Other revenue"] = monthly_harvest_revenue
                pnl_row["Treatment revenue (spot-income)"] = monthly_treatment_revenue
            pnl_row["Pass-through costs"] = lease_opex_this_month
            pnl_row["Total revenue"] = revenue
            for item in escalated_opex_items:
                pnl_row[item["name"]] = -item["value"]
            if spot_market_enabled:
                pnl_row["Fixed voyage opex (spot — shared, unallocated)"] = -spot_vessel_opex_this_month
                pnl_row["Transport voyage costs"] = -monthly_smolt_voyage_cost
                pnl_row["Other voyage costs"] = -monthly_harvest_voyage_cost
                pnl_row["Treatment voyage costs"] = -monthly_treatment_voyage_cost
                pnl_row["Treatment voyage costs"] = -monthly_treatment_voyage_cost
            pnl_row["Equipment opex (pass-through)"] = -lease_opex_this_month
            pnl_row["EBITDA — vessel"] = ebitda_vessel
            pnl_row["EBITDA — equipment"] = ebitda_equipment
            pnl_row["EBITDA"] = ebitda
            pnl_row["Depreciation — vessel"] = -monthly_vessel_depreciation
            pnl_row["Depreciation — maintenance capex (asset register)"] = -monthly_maintenance_capex_depreciation
            pnl_row["Depreciation — additional spot capex (asset register)"] = -monthly_additional_capex_depreciation
            pnl_row["Depreciation — equipment"] = -equipment_depreciation_this_month
            pnl_row["EBIT"] = ebit
            pnl_row["Finance cost — vessel (bank)"] = -vessel_finance_cost
            pnl_row["Finance cost — equipment (leasing company)"] = -equipment_finance_cost
            pnl_row["Finance cost — operational funding"] = -operational_finance_cost
            pnl_row["EBT"] = ebt
            pnl_row["Tax"] = -tax
            pnl_row["Net income"] = net_income
            pnl_rows_.append(pnl_row)

            # --- working capital: tracks the current (escalated) TOTAL
            # billed revenue run-rate — vessel/spot revenue, voyage cost
            # recovery, and lease revenue (incl. its pass-through opex) all
            # get the same DSO treatment, since they're all genuinely
            # invoiced to the customer the same way. ---
            daily_revenue_now = revenue * 12 / 365
            daily_opex_now = monthly_opex_vessel * 12 / 365
            ar_balance = daily_revenue_now * dso_days
            ap_balance = daily_opex_now * dpo_days
            this_nwc = ar_balance - ap_balance
            wc_change = this_nwc - prev_nwc
            prev_nwc = this_nwc

            cf_after_wc = ebitda - wc_change
            cf_after_finance = cf_after_wc - finance_cost_total
            cf_after_tax = cf_after_finance - tax
            capex_adjustment_this_month = capex_delta_by_month.get(month, 0.0)
            cash_flow_for_period = (
                cf_after_tax - vessel_amortization - equipment_amortization - operational_amortization
                - monthly_maintenance - monthly_additional_spot_capex
                + refinancing_proceeds_this_month - capex_adjustment_this_month
            )
            cumulative_cash += cash_flow_for_period

            cf_rows_.append({
                "Month": month,
                "EBITDA": ebitda,
                "Working capital change": -wc_change,
                "Finance cost — vessel (bank)": -vessel_finance_cost,
                "Finance cost — equipment (leasing company)": -equipment_finance_cost,
                "Finance cost — operational funding": -operational_finance_cost,
                "Tax": -tax,
                "Amortization — vessel (bank)": -vessel_amortization,
                "Amortization — equipment (leasing company)": -equipment_amortization,
                "Amortization — operational funding": -operational_amortization,
                "Maintenance capex": -monthly_maintenance,
                "Additional spot capex": -monthly_additional_spot_capex,
                "Capex adjustment (vessel upgrade/downgrade)": -capex_adjustment_this_month,
                "Refinancing proceeds (vessel)": refinancing_proceeds_this_month,
                "Operational funding injected (equity + debt)": 0.0,
                "Cash flow for the period": cash_flow_for_period,
                "Cash balance": cumulative_cash,
            })

            cumulative_vessel_depreciation += monthly_vessel_depreciation
            cumulative_equipment_depreciation += equipment_depreciation_this_month
            cumulative_maintenance_capex += monthly_maintenance
            cumulative_additional_spot_capex += monthly_additional_spot_capex
            cumulative_maintenance_capex_depreciation += monthly_maintenance_capex_depreciation
            cumulative_additional_capex_depreciation += monthly_additional_capex_depreciation
            cumulative_capex_adjustment += capex_adjustment_this_month
            vessel_nbv = (
                capex_nok - cumulative_vessel_depreciation
                + cumulative_maintenance_capex + cumulative_additional_spot_capex
                - cumulative_maintenance_capex_depreciation - cumulative_additional_capex_depreciation
                + cumulative_capex_adjustment
            )
            equipment_nbv = max(0.0, equipment_capex - cumulative_equipment_depreciation)
            equity += net_income
            total_assets = vessel_nbv + equipment_nbv + ar_balance + cumulative_cash
            total_liab_equity = vessel_debt_closing + equipment_debt_closing + operational_debt_closing + ap_balance + equity

            bs_rows_.append({
                "Month": month,
                "Vessel (NBV)": vessel_nbv,
                "Leased equipment (NBV)": equipment_nbv,
                "Accounts receivable": ar_balance,
                "Cash": cumulative_cash,
                "Total assets": total_assets,
                "Debt — vessel (bank)": vessel_debt_closing,
                "Debt — equipment (leasing company)": equipment_debt_closing,
                "Debt — operational funding": operational_debt_closing,
                "Accounts payable": ap_balance,
                "Equity": equity,
                "Total liabilities + equity": total_liab_equity,
            })

        return pnl_rows_, cf_rows_, bs_rows_

    # --- run twice: the REAL statements use whatever funding the user chose
    # on Tab 1 (possibly a mix of equity and debt); a separate, throwaway
    # BASELINE run forced to zero funding gives a stable guideline that
    # never moves depending on how much was funded — see the caption on
    # Tab 1's Sources & Uses for why this replaces the old iterative
    # approach entirely. ---
    pnl_rows, cf_rows, bs_rows = _run_monthly_model(operational_equity_nok, operational_debt_nok)
    _baseline_pnl_rows, _baseline_cf_rows, _baseline_bs_rows = _run_monthly_model(0.0, 0.0)
    _baseline_cash_series = [row["Cash balance"] for row in _baseline_cf_rows]
    _baseline_month_series = [row["Month"] for row in _baseline_cf_rows]
    _baseline_min_idx = _baseline_cash_series.index(min(_baseline_cash_series))
    baseline_min_cash_balance = _baseline_cash_series[_baseline_min_idx]
    baseline_min_cash_month = _baseline_month_series[_baseline_min_idx]

    pnl_df = pd.DataFrame(pnl_rows)
    cf_df = pd.DataFrame(cf_rows)
    bs_df = pd.DataFrame(bs_rows)

    # --- annual roll-ups ---
    pnl_df_yr = pnl_df.copy()
    pnl_df_yr["Year"] = ((pnl_df_yr["Month"] - 1) // 12 + 1)
    pnl_annual = pnl_df_yr.drop(columns=["Month"]).groupby("Year").sum()

    cf_df_yr = cf_df.copy()
    cf_df_yr["Year"] = ((cf_df_yr["Month"] - 1) // 12 + 1)
    flow_cols = [c for c in cf_df_yr.columns if c not in ("Month", "Year", "Cash balance")]
    cf_annual = cf_df_yr.groupby("Year")[flow_cols].sum()
    cf_annual["Cash balance"] = cf_df_yr.groupby("Year")["Cash balance"].last()  # year-end stock, not summed

    bs_year_end = bs_df[bs_df["Month"] % 12 == 0].copy()
    bs_year_end["Year"] = bs_year_end["Month"] // 12
    bs_opening = bs_df[bs_df["Month"] == 0].copy()
    bs_opening["Year"] = 0
    bs_annual = pd.concat([bs_opening, bs_year_end]).drop(columns=["Month"]).set_index("Year")

    def to_horizontal(df: pd.DataFrame, index_col: str = "Month", prefix: str = "Month") -> pd.DataFrame:
        t = df.set_index(index_col).T if index_col in df.columns else df.T
        t.columns = [f"{prefix} {c}" for c in t.columns]
        t.index.name = "Line item"
        return t

    def to_horizontal_indexed(df: pd.DataFrame, prefix: str) -> pd.DataFrame:
        t = df.T
        t.columns = [f"{prefix} {c}" for c in t.columns]
        t.index.name = "Line item"
        return t

    st.caption(f"All figures in {currency}.")

    if not lease_enabled:
        st.info(
            "Leased equipment is currently **off** (see the Lease spread tab). "
            "The statements below reflect the vessel only."
        )

    pnl_tab, cf_tab, bs_tab = st.tabs(["P&L", "Cash flow", "Balance sheet"])

    with pnl_tab:
        view = st.radio("View", ["Monthly", "Annual"], horizontal=True, key="pnl_view")
        if view == "Monthly":
            st.markdown(f"**Monthly P&L** — months 1 to {horizon_months}")
            show_table(to_horizontal(pnl_df, "Month", "Month"), width="stretch", height=460)
        else:
            st.markdown(f"**Annual P&L** — Year 1 to Year {pnl_annual.index.max()}")
            show_table(to_horizontal_indexed(pnl_annual, "Year"), width="stretch", height=460)

        chart_df = pnl_df[["Month", "EBITDA", "EBIT", "Net income"]]
        formatted_line_chart(chart_df, "Month", ["EBITDA", "EBIT", "Net income"])

    with cf_tab:
        view = st.radio("View", ["Monthly", "Annual"], horizontal=True, key="cf_view")
        if view == "Monthly":
            st.markdown(f"**Cash flow (EBITDA bridge)** — month 0 (opening) to month {horizon_months}")
            show_table(to_horizontal(cf_df, "Month", "Month"), width="stretch", height=380)
        else:
            st.markdown(f"**Annual cash flow (EBITDA bridge)** — Year 0 (opening) to Year {cf_annual.index.max()}")
            show_table(to_horizontal_indexed(cf_annual, "Year"), width="stretch", height=380)

        chart_df = pd.DataFrame({
            "Month": cf_df["Month"],
            "Cash balance": cf_df["Cash flow for the period"].cumsum(),
        })
        formatted_line_chart(chart_df, "Month", ["Cash balance"])

    with bs_tab:
        view = st.radio("View", ["Monthly", "Annual"], horizontal=True, key="bs_view")
        if view == "Monthly":
            st.markdown(f"**Balance sheet** — month 0 (opening) to month {horizon_months}")
            show_table(to_horizontal(bs_df, "Month", "Month"), width="stretch", height=420)
        else:
            st.markdown(f"**Balance sheet, year-end** — Year 0 (opening) to Year {bs_annual.index.max()}")
            show_table(to_horizontal_indexed(bs_annual, "Year"), width="stretch", height=420)

        chart_df = bs_df[["Month", "Vessel (NBV)", "Cash", "Equity"]]
        formatted_line_chart(chart_df, "Month", ["Vessel (NBV)", "Cash", "Equity"])

        max_imbalance = (bs_df["Total assets"] - bs_df["Total liabilities + equity"]).abs().max()
        if max_imbalance < 1.0:
            st.success("Balance sheet ties out: Total assets = Total liabilities + equity, every month.")
        else:
            st.error(
                f"Balance sheet does not tie out — largest imbalance is "
                f"{format_nok(max_imbalance)} NOK. This shouldn't happen; flag it if you see it."
            )

    # =======================================================================
    # Asset register — full depreciation build-up across all four streams:
    # vessel (original capex), equipment (lease), maintenance capex
    # vintages, and additional spot capex vintages — each on its own
    # schedule, shown explicitly so every P&L depreciation line can be
    # traced back to its source rather than trusted blind.
    # =======================================================================
    st.divider()
    st.subheader("Asset register — depreciation build-up")
    st.caption(
        "Every depreciation line in the P&L above, traced back to its "
        "own schedule — four separate streams, each with its own rate "
        "and useful life."
    )

    st.markdown("**1) Vessel** (original capex — Tab 1)")
    _vessel_useful_life = (100 / depreciation_rate_pct) if depreciation_rate_pct else 0.0
    vc1, vc2, vc3, vc4 = st.columns(4)
    vc1.metric("Capex", fmt(capex_nok))
    vc2.metric("Depreciation rate", f"{depreciation_rate_pct:.1f}%/yr")
    vc3.metric("Implied useful life", f"{_vessel_useful_life:.0f} years")
    vc4.metric("Annual depreciation", fmt(monthly_vessel_depreciation * 12))

    st.markdown("**2) Leased equipment** (Tab 2 — depreciated over the bank financing term)")
    if lease_enabled and equipment_capex > 0:
        ec1, ec2, ec3, ec4 = st.columns(4)
        ec1.metric("Capex", fmt(equipment_capex))
        ec2.metric("Financing term", f"{equipment_depreciation_months} months ({equipment_depreciation_months/12:.0f} years)")
        ec3.metric("Monthly depreciation", fmt(monthly_equipment_depreciation))
        ec4.metric("Annual depreciation", fmt(monthly_equipment_depreciation * 12))
        st.caption(
            "Straight-line over the bank financing term (currently "
            f"{int(bank_term_months)} months) — reaches zero NBV exactly "
            "when that term ends, unlike the vessel/maintenance/"
            "additional-capex streams, which keep depreciating at a flat "
            "rate indefinitely rather than fully writing off within the "
            "model horizon."
        )
    else:
        st.caption("Leased equipment is currently off (see the Lease spread tab) — no equipment depreciation.")

    st.markdown(f"**3) Maintenance capex** (Tab 1, {fmt(annual_maintenance_capex_nok)}/yr baseline — same rate/life as the vessel)")
    st.caption(
        f"Each year's maintenance capex becomes its own vintage, "
        f"depreciating at the vessel's own rate ({depreciation_rate_pct:.1f}%/yr, "
        f"{_vessel_useful_life:.0f}-year implied life) from its own start year — "
        f"applies unconditionally in both TC and spot mode."
    )
    _maint_vintage_rows = []
    for _vintage_year in range(1, 13):
        _vintage_start_month = (_vintage_year - 1) * 12 + 1
        _maint_this_vintage = annual_maintenance_capex_nok * _escalation_factor(maintenance_escalator_pct, _vintage_start_month)
        _annual_dep_this_vintage = _maint_this_vintage * (depreciation_rate_pct / 100)
        _maint_vintage_rows.append({
            "Vintage year": f"Year {_vintage_year}",
            "Maintenance capex added (nominal)": _maint_this_vintage,
            "Annual depreciation (this vintage, ongoing every year from here)": _annual_dep_this_vintage,
        })
    maint_vintage_df = pd.DataFrame(_maint_vintage_rows)
    show_table(maint_vintage_df, "Vintage year", width="stretch")

    st.markdown("Total maintenance capex depreciation, by reporting year")
    _maint_total_by_year_rows = []
    for _reporting_year in range(1, 13):
        _total_this_year = sum(
            _maint_vintage_rows[_vy - 1]["Annual depreciation (this vintage, ongoing every year from here)"]
            for _vy in range(1, _reporting_year + 1)
        )
        _maint_total_by_year_rows.append({"Year": f"Year {_reporting_year}", "Total maintenance capex depreciation": _total_this_year})
    maint_total_by_year_df = pd.DataFrame(_maint_total_by_year_rows)
    show_table(maint_total_by_year_df, "Year", width="stretch")

    if spot_market_enabled:
        _additional_useful_life = (100 / spot_additional_capex_depreciation_pct) if spot_additional_capex_depreciation_pct else 0.0
        st.markdown(
            f"**4) Additional spot capex** (Spot market tab, Year 2-12 — "
            f"own rate: {spot_additional_capex_depreciation_pct:.1f}%/yr, "
            f"{_additional_useful_life:.0f}-year implied life)"
        )
        st.caption(
            "Each year's Additional spot capex is entered in today's "
            "money and escalated to nominal terms via the maintenance "
            "capex escalator (same 'real value → nominal' treatment as "
            "Fixed Voyage opex) before becoming its own vintage, "
            "depreciating at its own separately-defined rate (set on the "
            "Spot market tab) — genuinely decoupled from the vessel/"
            "maintenance capex rate, since spot-specific capex may have "
            "a different useful life. Spot mode only."
        )
        _additional_vintage_rows = []
        for _vintage_year in range(1, 13):
            _add_capex_this_vintage = _additional_spot_capex_nominal_for_vintage(_vintage_year)
            _annual_dep_this_vintage = _add_capex_this_vintage * (spot_additional_capex_depreciation_pct / 100)
            _additional_vintage_rows.append({
                "Vintage year": f"Year {_vintage_year}",
                "Additional spot capex added (nominal)": _add_capex_this_vintage,
                "Annual depreciation (this vintage, ongoing every year from here)": _annual_dep_this_vintage,
            })
        additional_vintage_df = pd.DataFrame(_additional_vintage_rows)
        show_table(additional_vintage_df, "Vintage year", width="stretch")

        st.markdown("Total additional spot capex depreciation, by reporting year")
        _additional_total_by_year_rows = []
        for _reporting_year in range(1, 13):
            _total_this_year = sum(
                _additional_vintage_rows[_vy - 1]["Annual depreciation (this vintage, ongoing every year from here)"]
                for _vy in range(1, _reporting_year + 1)
            )
            _additional_total_by_year_rows.append({"Year": f"Year {_reporting_year}", "Total additional spot capex depreciation": _total_this_year})
        additional_total_by_year_df = pd.DataFrame(_additional_total_by_year_rows)
        show_table(additional_total_by_year_df, "Year", width="stretch")
    else:
        st.markdown("**4) Additional spot capex**")
        st.caption("Spot market is off — no additional spot capex depreciation applies.")

    # =======================================================================
    # Net revenue by segment — deliberately separate from the P&L's EBITDA
    # waterfall above (crew/vessel opex stays shared and unallocated there,
    # by design), showing each segment's own direct contribution instead.
    # =======================================================================
    if spot_market_enabled:
        st.divider()
        st.subheader("Net revenue by segment")
        st.caption(
            "Revenue less each segment's own direct voyage cost — kept "
            "separate from the EBITDA waterfall above, since crew/vessel "
            "opex there stays one shared, unallocated line rather than "
            "being split across segments (allocating a shared fixed cost "
            "is inherently arbitrary; this net revenue view is a cleaner "
            "'contribution' read instead). TC and Lease revenue have no "
            "directly-tracked offsetting cost in this model, so they're "
            "shown as-is."
        )
        _net_rev_rows = []
        for _yr in pnl_annual.index:
            _row = {"Year": f"Year {int(_yr)}"}
            _row["TC-revenue"] = pnl_annual.loc[_yr, "TC-revenue"]
            _row["Lease-revenue"] = pnl_annual.loc[_yr, "Lease-revenue"]
            if "Treatment revenue (spot-income)" in pnl_annual.columns:
                _row["Treatment net revenue"] = (
                    pnl_annual.loc[_yr, "Treatment revenue (spot-income)"]
                    + pnl_annual.loc[_yr, "Treatment voyage costs"]
                )
            if "Transport revenue" in pnl_annual.columns:
                _row["Transport net revenue"] = (
                    pnl_annual.loc[_yr, "Transport revenue"] + pnl_annual.loc[_yr, "Transport voyage costs"]
                )
            if "Other revenue" in pnl_annual.columns:
                _row["Other net revenue"] = (
                    pnl_annual.loc[_yr, "Other revenue"] + pnl_annual.loc[_yr, "Other voyage costs"]
                )
            _row["Total net revenue"] = sum(v for k, v in _row.items() if k != "Year")
            _net_rev_rows.append(_row)
        net_rev_by_segment_df = pd.DataFrame(_net_rev_rows).set_index("Year").T
        net_rev_by_segment_df.index.name = "Segment"
        show_table(net_rev_by_segment_df, width="stretch")

    # =======================================================================
    # Operational funding summary
    # =======================================================================
    st.divider()
    st.subheader("Operational funding")
    st.caption(
        "The vessel's implied equity (Tab 1) covers the purchase price only — "
        "it assumes the cash balance starts at zero and simply accumulates "
        "from there. If the monthly cash flow ever dips negative before it "
        "recovers, that's a real shortfall that needs covering so the vessel "
        "never actually runs out of cash. **Decide how to cover it on Tab 1's "
        "Sources & Uses section** — this is just a summary of what's "
        "currently applied there. The guideline below is always the deficit "
        "that would exist with **zero** operational funding — a fixed target "
        "that doesn't move depending on how much you choose to fund, so "
        "there's nothing to iterate towards."
    )

    # Actual, as-funded minimum (reflects whatever funding is currently applied)
    min_cash_balance = cf_df["Cash balance"].min()
    min_cash_month = int(cf_df.loc[cf_df["Cash balance"].idxmin(), "Month"])

    # The guideline: computed above from a SEPARATE run forced to zero
    # funding, so it's a fixed number that never depends on the funding
    # decision itself — no circularity, no staleness, no iteration needed.
    deficit_guideline_fixed = abs(baseline_min_cash_balance) if baseline_min_cash_balance < 0 else 0.0

    st.metric(
        "Maximum cash deficit (guideline, zero funding)",
        fmt(deficit_guideline_fixed),
        help=(
            f"With no operational funding at all, cumulative cash balance "
            f"bottoms out at {fmt(baseline_min_cash_balance)} in month "
            f"{baseline_min_cash_month}. This is the fixed target used to "
            f"size the equity/debt split on Tab 1 — it does not change "
            f"based on how much of it you actually fund."
        ) if deficit_guideline_fixed > 0 else "Cash flow never dips negative even with zero funding."
    )

    if min_cash_balance < 0:
        st.caption(
            f"⚠️ **With the funding currently applied**, cash balance still goes "
            f"as low as {fmt(min_cash_balance)} in month {min_cash_month}."
        )
    else:
        st.caption(
            f"✅ **With the funding currently applied**, cash balance never goes "
            f"negative — lowest point is {fmt(min_cash_balance)} in month {min_cash_month}."
        )

    op_col1, op_col2 = st.columns(2)
    op_col1.metric("Operational funding — equity portion", fmt(operational_equity_nok))
    op_col2.metric("Operational funding — debt portion", fmt(operational_debt_nok))

    total_equity_required = vessel_equity_initial + equipment_equity_initial + operational_equity_nok
    st.markdown(
        f"**Total equity required:** {fmt(vessel_equity_initial)} (vessel) + "
        f"{fmt(equipment_equity_initial)} (equipment) + {fmt(operational_equity_nok)} (operational, equity portion) "
        f"= **{fmt(total_equity_required)}**"
    )

    # --- store for the Sources & Uses summary on Tab 1 (which runs earlier in
    #     the script and can't compute this itself — see the note there).
    #     Crucially, this is the FIXED, zero-funding guideline, not the
    #     as-funded minimum — so it's stable pass to pass regardless of what
    #     funding decision was made, eliminating the old circularity. ---
    _previous_su = st.session_state.get("_sources_uses")
    st.session_state["_sources_uses"] = {
        "vessel_equity": vessel_equity_initial,
        "equipment_equity": equipment_equity_initial,
        "total_equity": total_equity_required,
        "min_cash_balance": baseline_min_cash_balance,
        "min_cash_month": baseline_min_cash_month,
        "vessel_debt": debt_nok,
        "equipment_debt": equipment_debt_initial,
        "lease_enabled": lease_enabled,
    }

    # --- self-healing refresh: Tab 1 is one script pass behind by design
    # (it runs before this tab, so it always shows what was stored on the
    # PREVIOUS pass). If what was just freshly computed here differs
    # meaningfully from what Tab 1 is currently displaying, that number is
    # stale — trigger a deferred rerun so it catches up automatically,
    # without needing a manual Refresh click or a tab switch. Capped at a
    # few retries per interaction as a safety net against any edge case
    # that doesn't settle. ---
    _guideline_changed = (
        _previous_su is None
        or abs(_previous_su.get("min_cash_balance", 0.0) - baseline_min_cash_balance) > 1.0
    )
    _stale_refresh_count = st.session_state.get("_stale_guideline_refresh_count", 0)
    if _guideline_changed and _stale_refresh_count < 4:
        st.session_state["_stale_guideline_refresh_count"] = _stale_refresh_count + 1
        _request_rerun()
    else:
        st.session_state["_stale_guideline_refresh_count"] = 0

# ===========================================================================
# TAB 5 — Investment Analysis (equity IRR, terminal value on forward EBITDA)
# ===========================================================================
with tab_investment:
    st.subheader("Investment analysis — equity IRR")
    st.caption(
        "Equity cash flows from the initial investment (month 0) through "
        "monthly operations, plus a terminal value at exit. Each month's "
        "'Cash flow for the period' (after all debt service, tax, and capex — "
        "see Financial Statements) is treated as available to equity as it's "
        "generated. At exit, enterprise value is set at a multiple of the "
        "**coming year's** projected EBITDA, and outstanding debt (vessel + "
        "equipment) is deducted to arrive at terminal equity value."
    )

    terminal_multiple = stateful_number_input(
        "Terminal EBITDA multiple (x forward EBITDA)", min_value=0.0, value=10.0,
        step=0.5, key="terminal_multiple", disabled=locked
    )

    def _project_ebitda_month(month):
        """EBITDA (vessel + equipment) for any month, including beyond the
        modeled horizon — used to project the exit year's EBITDA."""
        revenue = _get_vessel_revenue(month)
        opex = _get_vessel_opex(month)
        if spot_market_enabled:
            # _get_vessel_revenue already sums Treatment/Transport/Other
            # revenue (incl. calendar/utilization); add all three
            # segments' own build-up-tool voyage costs here too, via the
            # same calendar-aware distributor used in the main monthly
            # loop (crew/vessel opex above is the shared line only,
            # unaffected).
            opex += (
                _service_monthly_voyage_cost("Treatment of fish", month, _treatment_annual_voyage_cost, spot_treatment_cost_escalator_pct)
                + _service_monthly_voyage_cost("Transport", month, _smolt_total_annual_voyage_cost, spot_smolt_cost_escalator_pct)
                + _service_monthly_voyage_cost("Other", month, _harvest_total_annual_voyage_cost, spot_harvest_cost_escalator_pct)
            )
        ebitda_v = revenue - opex
        if lease_enabled and not spot_market_enabled and _lease_contract_active(month):
            ebitda_e = _revenue_for_lease_contracts(month)
        else:
            ebitda_e = 0.0
        return ebitda_v + ebitda_e

    forward_year_start = horizon_months + 1
    forward_annual_ebitda = sum(
        _project_ebitda_month(m) for m in range(forward_year_start, forward_year_start + 12)
    )

    outstanding_debt_at_exit = (
        bs_df.iloc[-1]["Debt — vessel (bank)"] + bs_df.iloc[-1]["Debt — equipment (leasing company)"]
    )
    enterprise_value_at_exit = terminal_multiple * forward_annual_ebitda
    terminal_equity_value = enterprise_value_at_exit - outstanding_debt_at_exit

    tv_col1, tv_col2, tv_col3 = st.columns(3)
    tv_col1.metric("Forward EBITDA (next 12 months)", fmt(forward_annual_ebitda))
    tv_col2.metric("Enterprise value at exit", fmt(enterprise_value_at_exit))
    tv_col3.metric("Outstanding debt at exit", fmt(outstanding_debt_at_exit))
    st.metric("Terminal equity value", fmt(terminal_equity_value))

    # --- build the monthly equity cash flow stream ---
    st.subheader("Equity invested")
    st.caption(
        "How the vessel, equipment, and any operational funding are financed "
        "is decided on Tab 1's Sources & Uses section — the equity portion "
        "chosen there is included in the month 0 investment below, matching "
        "the cash flow model exactly (any debt portion instead shows up as "
        "ongoing interest and amortization, already reflected in the monthly "
        "cash flow). To raise this IRR, cover less of the operational funding "
        "with equity on Tab 1 and let debt (or group liquidity) absorb more "
        "of it."
    )
    if operational_debt_nok > 0:
        st.caption(
            f"Currently: {fmt(operational_equity_nok)} equity + {fmt(operational_debt_nok)} "
            f"debt covering the operational funding requirement."
        )

    initial_equity_investment = vessel_equity_initial + equipment_equity_initial + operational_equity_nok
    equity_cf_rows = [{"Month": 0, "Equity cash flow": -initial_equity_investment}]
    for _, row in cf_df.iterrows():
        m = int(row["Month"])
        cf = row["Cash flow for the period"]
        if m == horizon_months:
            cf += terminal_equity_value
        equity_cf_rows.append({"Month": m, "Equity cash flow": cf})

    equity_cf_df = pd.DataFrame(equity_cf_rows)

    def _irr_bisection(cashflows, low=-0.5, high=1.0, tol=1e-9, max_iter=200):
        def npv(r):
            return sum(cf / (1 + r) ** t for t, cf in enumerate(cashflows))
        f_lo, f_hi = npv(low), npv(high)
        if f_lo * f_hi > 0:
            return None
        for _ in range(max_iter):
            mid = (low + high) / 2
            f_mid = npv(mid)
            if abs(f_mid) < tol:
                return mid
            if f_lo * f_mid < 0:
                high, f_hi = mid, f_mid
            else:
                low, f_lo = mid, f_mid
        return (low + high) / 2

    monthly_irr = _irr_bisection(equity_cf_df["Equity cash flow"].tolist())

    st.subheader("Equity IRR")
    if monthly_irr is None:
        st.error(
            "IRR could not be computed — the cash flow pattern doesn't cross "
            "zero NPV within a plausible rate range (e.g. total cash returned "
            "may be less than invested, or the pattern is unusual)."
        )
    else:
        annual_irr = (1 + monthly_irr) ** 12 - 1
        total_distributed = equity_cf_df["Equity cash flow"].iloc[1:].sum()
        moic = total_distributed / initial_equity_investment if initial_equity_investment else 0.0

        irr_col1, irr_col2, irr_col3 = st.columns(3)
        irr_col1.metric("Equity IRR, annual", f"{annual_irr:+.1%}")
        irr_col2.metric("Equity IRR, monthly", f"{monthly_irr:+.2%}")
        irr_col3.metric("MOIC (multiple on invested capital)", f"{moic:.2f}x")

        st.caption(
            f"Initial equity invested: {fmt(initial_equity_investment)} at month 0. "
            f"Total cash returned to equity over the life of the investment "
            f"(including the terminal value): {fmt(total_distributed)}."
        )

    st.divider()
    st.subheader("Instant equity value creation (Day 1 mark-to-market)")
    st.caption(
        "A quick check owners look at: if the vessel could be sold "
        "immediately at a market EBITDA multiple, using Year 1 EBITDA, "
        "what's the resulting equity value versus what's actually been "
        "put in? EV = Year 1 EBITDA x multiple. Instant equity value = EV "
        "less starting debt (vessel + leased equipment only — excludes "
        "any operational funding debt). Shown against equity exposed both "
        "with and without covering the operational funding gap via "
        "equity, to see the immediate value-creation MOIC either way — "
        "distinct from the full-horizon Equity IRR/MOIC above, which "
        "accounts for cash flows over the whole investment life plus the "
        "terminal value at actual exit."
    )

    _year1_ebitda_ia = pnl_annual.loc[1, "EBITDA"] if 1 in pnl_annual.index else 0.0
    _debt_vessel_and_lease_at_close_ia = debt_nok + equipment_debt_initial
    _equity_excl_opfunding_ia = vessel_equity_initial + equipment_equity_initial
    _equity_incl_opfunding_ia = vessel_equity_initial + equipment_equity_initial + operational_equity_nok

    instant_value_rows_ia = []
    for _multiple in (10.0, 11.0, 12.0):
        _ev = _year1_ebitda_ia * _multiple
        _instant_equity_value = _ev - _debt_vessel_and_lease_at_close_ia
        _moic_excl = (_instant_equity_value / _equity_excl_opfunding_ia) if _equity_excl_opfunding_ia else None
        _moic_incl = (_instant_equity_value / _equity_incl_opfunding_ia) if _equity_incl_opfunding_ia else None
        instant_value_rows_ia.append({
            "EBITDA multiple": f"{_multiple:.0f}x",
            "Enterprise value (EV)": fmt(_ev),
            "Less: debt (vessel + lease)": fmt(_debt_vessel_and_lease_at_close_ia),
            "Instant equity value": fmt(_instant_equity_value),
            "Equity exposed, excl. op. funding gap": fmt(_equity_excl_opfunding_ia),
            "Instant MOIC, excl. gap (x)": f"{_moic_excl:.2f}x" if _moic_excl is not None else "—",
            "Equity exposed, incl. op. funding gap": fmt(_equity_incl_opfunding_ia),
            "Instant MOIC, incl. gap (x)": f"{_moic_incl:.2f}x" if _moic_incl is not None else "—",
        })
    instant_value_df_ia = pd.DataFrame(instant_value_rows_ia).set_index("EBITDA multiple")
    st.dataframe(instant_value_df_ia, width="stretch")
    st.caption(
        f"Operational funding gap currently modeled: {fmt(operational_equity_nok)} "
        f"of equity allocated to it (see Tab 1's Sources & Uses to adjust)."
    )

    st.subheader("Equity cash flow schedule")
    chart_df = equity_cf_df.copy()
    chart_df["Cumulative equity cash flow"] = chart_df["Equity cash flow"].cumsum()
    formatted_line_chart(chart_df, "Month", ["Cumulative equity cash flow"])

    show_table(equity_cf_df, "Month", width="stretch", height=320)

# ===========================================================================
# TAB 6 — Summary (key financial ratios & KPIs)
# ===========================================================================
with tab_summary:
    st.subheader("Summary — key financial ratios & KPIs")
    st.caption(
        "First working version — definitions below are a starting point; "
        "flag anything you'd like calculated differently (e.g. NIBD scope, "
        "ICR basis, LTV basis) and we'll refine it."
    )

    total_capex_at_close = capex_nok + equipment_capex
    total_debt_at_close = debt_nok + equipment_debt_initial + operational_debt_nok
    total_equity_at_close = vessel_equity_initial + equipment_equity_initial + operational_equity_nok
    ltv_at_close = (total_debt_at_close / total_capex_at_close) if total_capex_at_close else 0.0
    gearing_at_close = (total_debt_at_close / total_equity_at_close) if total_equity_at_close else 0.0

    st.markdown("**Financial close snapshot (Month 0)**")
    c1, c2, c3, c4, c5 = st.columns(5)
    c1.metric("Total capex", fmt(total_capex_at_close))
    c2.metric("Total debt at close", fmt(total_debt_at_close))
    c3.metric("Total equity at close", fmt(total_equity_at_close))
    c4.metric("LTV at close", f"{ltv_at_close:.1%}")
    c5.metric("Gearing (debt / equity)", f"{gearing_at_close:.2f}x")

    st.divider()
    st.markdown("**Investment return KPIs**")
    i1, i2, i3, i4 = st.columns(4)
    if monthly_irr is None:
        i1.metric("Equity IRR, annual", "n/a")
        i2.metric("MOIC", "n/a")
    else:
        i1.metric("Equity IRR, annual", f"{annual_irr:+.1%}")
        i2.metric("MOIC", f"{moic:.2f}x")

    _year1_ebitda = pnl_annual.loc[1, "EBITDA"] if 1 in pnl_annual.index else 0.0
    _yield_on_capex_yr1 = (_year1_ebitda / total_capex_at_close) if total_capex_at_close else 0.0
    _year1_nbv = (
        (bs_annual.loc[1, "Vessel (NBV)"] + bs_annual.loc[1, "Leased equipment (NBV)"])
        if 1 in bs_annual.index else 0.0
    )
    _yield_on_nbv_yr1 = (_year1_ebitda / _year1_nbv) if _year1_nbv else 0.0
    i3.metric("EBITDA yield on capex (Yr 1)", f"{_yield_on_capex_yr1:.1%}")
    i4.metric("EBITDA yield on NBV (Yr 1)", f"{_yield_on_nbv_yr1:.1%}")

    st.divider()
    st.markdown("**Instant equity value creation** (Day 1 mark-to-market)")
    st.caption(
        "A quick check owners look at: if the vessel could be sold "
        "immediately at a market EBITDA multiple, using Year 1 EBITDA, "
        "what's the resulting equity value versus what's actually been "
        "put in? EV = Year 1 EBITDA x multiple. Instant equity value = EV "
        "less starting debt (vessel + leased equipment only — excludes "
        "any operational funding debt). Shown against equity exposed both "
        "with and without covering the operational funding gap via "
        "equity, to see the immediate value-creation MOIC either way."
    )

    _debt_vessel_and_lease_at_close = debt_nok + equipment_debt_initial
    _equity_excl_opfunding = vessel_equity_initial + equipment_equity_initial
    _equity_incl_opfunding = total_equity_at_close  # already includes operational_equity_nok

    instant_value_rows = []
    for _multiple in (10.0, 11.0, 12.0):
        _ev = _year1_ebitda * _multiple
        _instant_equity_value = _ev - _debt_vessel_and_lease_at_close
        _moic_excl = (_instant_equity_value / _equity_excl_opfunding) if _equity_excl_opfunding else None
        _moic_incl = (_instant_equity_value / _equity_incl_opfunding) if _equity_incl_opfunding else None
        instant_value_rows.append({
            "EBITDA multiple": f"{_multiple:.0f}x",
            "Enterprise value (EV)": fmt(_ev),
            "Less: debt (vessel + lease)": fmt(_debt_vessel_and_lease_at_close),
            "Instant equity value": fmt(_instant_equity_value),
            "Equity exposed, excl. op. funding gap": fmt(_equity_excl_opfunding),
            "Instant MOIC, excl. gap (x)": f"{_moic_excl:.2f}x" if _moic_excl is not None else "—",
            "Equity exposed, incl. op. funding gap": fmt(_equity_incl_opfunding),
            "Instant MOIC, incl. gap (x)": f"{_moic_incl:.2f}x" if _moic_incl is not None else "—",
        })
    instant_value_df = pd.DataFrame(instant_value_rows).set_index("EBITDA multiple")
    st.dataframe(instant_value_df, width="stretch")
    st.caption(
        f"Operational funding gap currently modeled: {fmt(operational_equity_nok)} "
        f"of equity allocated to it (see Tab 1's Sources & Uses to adjust)."
    )

    st.divider()
    st.markdown("**Annual ratios**")
    st.caption(
        "NIBD = total debt (vessel + equipment + operational funding) less "
        "cash, at year-end. ICR = EBITDA / total finance cost. LTV is shown "
        "on both a declining (current NBV) and a fixed (original capex) "
        "basis. ROE = net income (annual) / book equity net of cash — as "
        "if all cash on the balance sheet had been distributed out as a "
        "dividend, so the equity base only reflects capital still tied up "
        "in the vessel, equipment, and working capital."
    )

    ratio_rows = []
    for y in pnl_annual.index:
        if y not in bs_annual.index:
            continue
        ebitda_y = pnl_annual.loc[y, "EBITDA"]
        finance_cost_y = -(
            pnl_annual.loc[y, "Finance cost — vessel (bank)"]
            + pnl_annual.loc[y, "Finance cost — equipment (leasing company)"]
            + pnl_annual.loc[y, "Finance cost — operational funding"]
        )
        debt_total_y = (
            bs_annual.loc[y, "Debt — vessel (bank)"]
            + bs_annual.loc[y, "Debt — equipment (leasing company)"]
            + bs_annual.loc[y, "Debt — operational funding"]
        )
        cash_y = bs_annual.loc[y, "Cash"]
        nibd_y = debt_total_y - cash_y
        vessel_nbv_y = bs_annual.loc[y, "Vessel (NBV)"]
        equipment_nbv_y = bs_annual.loc[y, "Leased equipment (NBV)"]
        nbv_total_y = vessel_nbv_y + equipment_nbv_y
        book_equity_y = bs_annual.loc[y, "Equity"]
        net_income_y = pnl_annual.loc[y, "Net income"]

        nibd_ebitda = (nibd_y / ebitda_y) if ebitda_y else None
        icr = (ebitda_y / finance_cost_y) if finance_cost_y else None
        ltv_nbv_basis = (debt_total_y / nbv_total_y) if nbv_total_y else None
        ltv_capex_basis = (debt_total_y / total_capex_at_close) if total_capex_at_close else None
        yield_nbv = (ebitda_y / nbv_total_y) if nbv_total_y else None
        yield_capex = (ebitda_y / total_capex_at_close) if total_capex_at_close else None
        # As if all cash had been distributed out as a dividend: book equity
        # net of cash, used as the denominator for ROE below.
        book_equity_ex_cash_y = book_equity_y - cash_y
        roe = (net_income_y / book_equity_ex_cash_y) if book_equity_ex_cash_y else None

        ratio_rows.append({
            "Year": f"Year {int(y)}",
            "EBITDA": fmt(ebitda_y),
            "NIBD": fmt(nibd_y),
            "NIBD / EBITDA (x)": f"{nibd_ebitda:.2f}x" if nibd_ebitda is not None else "—",
            "Finance cost, total": fmt(finance_cost_y),
            "ICR — EBITDA / finance cost (x)": f"{icr:.2f}x" if icr is not None else "—",
            "Total debt, period-end": fmt(debt_total_y),
            "Fixed assets NBV, period-end": fmt(nbv_total_y),
            "LTV — debt / NBV (%)": f"{ltv_nbv_basis:.1%}" if ltv_nbv_basis is not None else "—",
            "LTV — debt / original capex (%)": f"{ltv_capex_basis:.1%}" if ltv_capex_basis is not None else "—",
            "EBITDA yield on NBV (%)": f"{yield_nbv:.1%}" if yield_nbv is not None else "—",
            "EBITDA yield on original capex (%)": f"{yield_capex:.1%}" if yield_capex is not None else "—",
            "Vessel book value (NBV), period-end": fmt(vessel_nbv_y),
            "Equipment book value (NBV), period-end": fmt(equipment_nbv_y),
            "Book equity, period-end": fmt(book_equity_y),
            "Book equity excl. cash (as if dividended out)": fmt(book_equity_ex_cash_y),
            "ROE — net income / book equity excl. cash (%)": f"{roe:.1%}" if roe is not None else "—",
        })

    ratio_df = pd.DataFrame(ratio_rows).set_index("Year").T
    ratio_df.index.name = "Metric"
    st.dataframe(ratio_df, width="stretch", height=460)

# ===========================================================================
# TAB 7 — Inputs (read-only summary of every input currently set, grouped
#          by section — a quick audit view: check everything is as it
#          should be before sharing the model, without hunting across
#          nine tabs. Purely a display of values already computed/stored
#          above; never sets or edits anything itself.)
# ===========================================================================
with tab_inputs:
    st.subheader("Input summary")
    st.caption(
        "Every input currently set in the model, grouped by section — "
        "read-only, for a quick check before sharing or saving this as "
        "the default. To change anything, go to that input's own tab; "
        "nothing here is editable. Figures shown reflect the current "
        "unlocked/locked state and currently active mode (TC vs. spot "
        "market)."
    )

    def _bool_label(value):
        return "On" if value else "Off"

    def _show_input_table(df_or_rows, label_col, width="stretch", height=None):
        """Like show_table(), but without its NumberColumn formatting —
        every cell in this tab is already a formatted string (e.g.
        '819 000 NOK'), not a raw number, so show_table's numeric column
        config isn't the right fit here."""
        _df = df_or_rows if isinstance(df_or_rows, pd.DataFrame) else pd.DataFrame(df_or_rows)
        display_df = _df.set_index(label_col) if label_col else _df
        st.dataframe(display_df, width=width, height=height)

    st.markdown("### Vessel — capital & return (Tab 1)")
    vessel_capital_rows = [
        {"Input": "Capex", "Value": f"{fmt(capex_nok)} {currency}"},
        {"Input": "EBITDA-yield", "Value": f"{ebitda_yield_pct:.2f}%"},
        {"Input": "Operating days / year", "Value": f"{fmt(operating_days)}"},
        {"Input": "Vessel depreciation rate", "Value": f"{depreciation_rate_pct:.2f}%/yr"},
        {"Input": "Annual maintenance capex", "Value": f"{fmt(annual_maintenance_capex_nok)} {currency}"},
    ]
    _show_input_table(pd.DataFrame(vessel_capital_rows), "Input", width="stretch", height=210)

    st.markdown("### Vessel opex (Tab 1 / Vessel opex)")
    opex_rows = [
        {"Line item": item["name"], "Annual value": f"{fmt(item['value_nok'])} {currency}"}
        for item in st.session_state.opex_items
    ] + [
        {"Line item": "Total vessel opex", "Annual value": f"{fmt(opex_total)} {currency}"},
        {"Line item": "Linked to Vessel opex tab build-up?", "Annual value": _bool_label(opex_linked_to_vessel_opex_tab)},
    ]
    _show_input_table(pd.DataFrame(opex_rows), "Line item", width="stretch", height=min(300, 45 + 38 * len(opex_rows)))

    st.markdown("### Debt financing — vessel (Tab 1)")
    debt_rows = [
        {"Input": "Debt multiple", "Value": f"{debt_multiple:.1f}x Year 1 EBITDA"},
        {"Input": "Amortization profile", "Value": f"{fmt(amortization_years)} years"},
        {"Input": "Swap rate", "Value": f"{swap_rate_pct:.2f}%"},
        {"Input": "Credit spread", "Value": f"{credit_spread_pct:.2f}%"},
        {"Input": "Implied debt (Year 1 EBITDA x multiple)", "Value": f"{fmt(debt_nok)} {currency}"},
        {"Input": "Implied LTV", "Value": f"{implied_ltv_pct:.1f}%"},
    ]
    _show_input_table(pd.DataFrame(debt_rows), "Input", width="stretch", height=250)

    st.markdown("### Debt refinancing — vessel (Financial Statements)")
    if refinancing_enabled:
        refi_rows = [
            {"Input": "Refinancing enabled?", "Value": "On"},
            {"Input": "First refinancing", "Value": f"Year {int(refi_year1)}"},
            {"Input": "Second refinancing", "Value": f"Year {int(refi_year2)}"},
            {"Input": "Releverage multiple", "Value": f"{releverage_multiple:.1f}x next year's EBITDA"},
        ]
    else:
        refi_rows = [{"Input": "Refinancing enabled?", "Value": "Off"}]
    _show_input_table(pd.DataFrame(refi_rows), "Input", width="stretch", height=min(210, 45 + 38 * len(refi_rows)))

    st.markdown("### Construction capex (Construction capex tab)")
    construction_fx_rows = [
        {"Currency": "EUR/NOK", "Rate": f"{construction_fx_eur:.4f}"},
        {"Currency": "USD/NOK", "Rate": f"{construction_fx_usd:.4f}"},
        {"Currency": "GBP/NOK", "Rate": f"{construction_fx_gbp:.4f}"},
        {"Currency": "CAD/NOK", "Rate": f"{construction_fx_cad:.4f}"},
        {"Currency": "CLP/NOK", "Rate": f"{construction_fx_clp:.4f}"},
    ]
    _show_input_table(pd.DataFrame(construction_fx_rows), "Currency", width="stretch", height=210)

    construction_capex_item_rows = [
        {"Item": it["name"], "Value": f"{fmt(it['nok_value'])} {currency}"}
        for it in _construction_capex_items_nok
    ] + [
        {"Item": "Total installment capex", "Value": f"{fmt(_construction_installment_capex_nok)} {currency}"},
        {"Item": "Total non-installment capex (PM, legal, etc.)", "Value": f"{fmt(_construction_other_capex_nok)} {currency}"},
        {"Item": "Total capitalized cost (excl. finance cost & guarantee premium)", "Value": f"{fmt(_construction_total_capitalized_nok)} {currency}"},
        {"Item": "Total debt", "Value": f"{fmt(construction_total_debt_nok)} {currency}"},
        {"Item": "Guarantee premium rate", "Value": f"{construction_guarantee_premium_pct:.2f}%/yr"},
        {"Item": "Construction swap rate", "Value": f"{construction_swap_rate_pct:.2f}%"},
        {"Item": "Construction credit spread", "Value": f"{construction_credit_spread_pct:.2f}%"},
    ]
    _show_input_table(
        pd.DataFrame(construction_capex_item_rows), "Item", width="stretch",
        height=min(360, 45 + 38 * len(construction_capex_item_rows))
    )

    installment_rows = [
        {
            "Installment": ("Take-out-financing" if idx == _construction_n_installments - 1 else f"{idx + 1}{'st' if idx == 0 else 'nd' if idx == 1 else 'rd' if idx == 2 else 'th'} yard-installment"),
            "Share (%)": f"{inst['share_pct']:.1f}%",
            "Month": f"{inst['month']:.1f}",
        }
        for idx, inst in enumerate(st.session_state.construction_installments)
    ]
    _show_input_table(pd.DataFrame(installment_rows), "Installment", width="stretch", height=min(250, 45 + 38 * len(installment_rows)))

    st.markdown("### Spot market (Spot market tab)")
    spot_mode_rows = [
        {"Input": "Spot market active?", "Value": _bool_label(spot_market_enabled)},
        {"Input": "Utilization (baseline)", "Value": f"{spot_utilization_pct:.1f}%"},
        {"Input": "Fixed Voyage opex (Year 1)", "Value": f"{fmt(spot_opex_annual_nok)} {currency}/yr"},
        {"Input": "Fixed Voyage opex escalator", "Value": f"{spot_opex_escalator_pct:.2f}%/yr"},
        {"Input": "Variable Voyage opex escalator", "Value": f"{spot_variable_opex_escalator_pct:.2f}%/yr"},
        {"Input": "Monthly activity calendar active?", "Value": _bool_label(_calendar_enabled())},
    ]
    _show_input_table(pd.DataFrame(spot_mode_rows), "Input", width="stretch", height=250)

    spot_service_rows = [
        {
            "Service": item["name"],
            "Share of working days (%)": f"{item.get('share_pct', 0.0):.1f}%" if not _calendar_enabled() else "— (calendar)",
            "Priced at baseline?": _bool_label(item.get("priced_at_baseline", False)),
            "Customer mix on?": _bool_label(_customer_mix_enabled(item["name"])),
            "Rate (NOK/day), Yr 1": fmt(_resolved_charged_rate(item["name"])),
        }
        for item in spot_service_items_current
    ]
    _show_input_table(
        pd.DataFrame(spot_service_rows), "Service", width="stretch",
        height=min(250, 45 + 38 * len(spot_service_rows))
    )

    if _calendar_enabled():
        st.markdown("**Weekly activity calendar**")
        _week_to_month_inputs = _week_to_month_map()
        _month_labels_inputs = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
        _calendar_rows = [
            {
                "Week": f"W{wi + 1}",
                "Month": _month_labels_inputs[_week_to_month_inputs.get(wi + 1, 0)],
                "Activity": row.get("activity", "Idle"),
                "Occupancy (%)": f"{row.get('occupancy_pct', 0.0):.1f}%",
            }
            for wi, row in enumerate(st.session_state.get("spot_calendar_weeks", []))
        ]
        if _calendar_rows:
            _show_input_table(pd.DataFrame(_calendar_rows), "Week", width="stretch", height=min(420, 45 + 38 * 12))

    for _svc_name, _svc_slug in _service_slug_lookup.items():
        if _customer_mix_enabled(_svc_name):
            st.markdown(f"**{_svc_name} — customer mix**")
            _cust_rows = [
                {
                    "Customer": c["name"],
                    "Weeks": len(c.get("weeks", [])),
                    "Share (%)": f"{c['share_pct']:.1f}%",
                    "Rate (NOK/day), Yr 1": fmt(c["rate_nok_day"]),
                }
                for c in st.session_state.get(f"spot_customers_{_svc_slug}", [])
            ]
            if _cust_rows:
                _show_input_table(
                    pd.DataFrame(_cust_rows), "Customer", width="stretch",
                    height=min(220, 45 + 38 * len(_cust_rows))
                )

    st.markdown("### Leased equipment (Lease spread tab)")
    lease_rows = [
        {"Input": "Customer lease enabled?", "Value": _bool_label(lease_enabled)},
        {"Input": "Bank financing enabled?", "Value": _bool_label(bank_financing_enabled)},
        {"Input": "Equipment capex", "Value": f"{fmt(lease_capex_nok)} {currency}"},
        {"Input": "Lease-out rate", "Value": f"{lease_yield_pct:.2f}%"},
        {"Input": "Customer lease term", "Value": f"{fmt(customer_term_months)} months"},
        {"Input": "Lease payback structure", "Value": f"{fmt(lease_payback_months)} months"},
        {"Input": "Additional opex billed to customer", "Value": f"{fmt(lease_opex_monthly_nok)} {currency}/month"},
        {"Input": "Bank interest rate", "Value": f"{bank_rate_pct:.2f}%" if bank_financing_enabled else "n/a (bank financing off)"},
        {"Input": "Bank loan term", "Value": f"{fmt(bank_term_months)} months" if bank_financing_enabled else "n/a (bank financing off)"},
        {"Input": "Equity instalment", "Value": f"{fmt(lease_equity_instalment_nok)} {currency}" if bank_financing_enabled else "n/a (bank financing off)"},
    ]
    _show_input_table(pd.DataFrame(lease_rows), "Input", width="stretch", height=380)

    st.markdown("### Lease contract schedule (Financial Statements)")
    if lease_enabled and lease_contracts:
        lease_contract_rows = [
            {
                "Contract": f"Contract {idx + 1}",
                "Months": f"{c['start']}–{c['start'] + c['length'] - 1}",
                "Rate (NOK/year)": fmt(c["base_monthly"] * 12),
                "Escalator (%/yr)": f"{c['escalator_pct']:.2f}%",
            }
            for idx, c in enumerate(lease_contracts)
        ]
        _show_input_table(
            pd.DataFrame(lease_contract_rows), "Contract", width="stretch",
            height=min(220, 45 + 38 * len(lease_contract_rows))
        )
    else:
        st.caption("Leased equipment is currently off — no lease contract schedule applies.")

    st.markdown("### TC contract schedule (Financial Statements)")
    tc_contract_rows = [
        {
            "Contract": f"Contract {idx + 1}",
            "Months": f"{c['start']}–{c['start'] + c['length'] - 1}",
            "Rate (NOK/year)": fmt(c["base_monthly"] * 12),
            "Capex adjustment": fmt(c["capex_delta"]) if c["capex_delta"] else "—",
        }
        for idx, c in enumerate(tc_contracts)
    ]
    _show_input_table(
        pd.DataFrame(tc_contract_rows), "Contract", width="stretch",
        height=min(220, 45 + 38 * len(tc_contract_rows))
    )

    st.markdown("### Escalators (Financial Statements)")
    escalator_rows = [
        {"Escalator": "TC revenue", "Rate": f"{tc_escalator_pct:.2f}%/yr"},
        {"Escalator": "Lease payment (Contract 1)", "Rate": f"{lease_escalator_pct:.2f}%/yr"},
        {"Escalator": "Maintenance capex", "Rate": f"{maintenance_escalator_pct:.2f}%/yr"},
    ] + [
        {"Escalator": f"Vessel opex — {item['name']}", "Rate": f"{item['escalator_pct']:.2f}%/yr"}
        for item in opex_line_items_base
    ]
    _show_input_table(
        pd.DataFrame(escalator_rows), "Escalator", width="stretch",
        height=min(320, 45 + 38 * len(escalator_rows))
    )

    st.markdown("### Working capital, tax & investment assumptions")
    other_assumption_rows = [
        {"Input": "Days sales outstanding (DSO)", "Value": f"{fmt(dso_days)} days"},
        {"Input": "Days payable outstanding (DPO)", "Value": f"{fmt(dpo_days)} days"},
        {"Input": "Corporate tax rate", "Value": f"{tax_rate_pct:.2f}%"},
        {"Input": "Terminal EBITDA multiple (exit)", "Value": f"{terminal_multiple:.1f}x forward EBITDA"},
    ]
    _show_input_table(pd.DataFrame(other_assumption_rows), "Input", width="stretch", height=210)

    st.divider()
    st.caption(
        "This tab reflects whatever is currently loaded in the app "
        "(either the saved default_config.json, or anything you've "
        "changed this session while unlocked) — it does not itself "
        "save or change anything. To lock in what's shown here as the "
        "new default for every visitor, use 'Save current inputs as "
        "default' in the sidebar."
    )

# ===========================================================================
# Excel export — every table in the app, on one workbook, one click
# ===========================================================================
st.divider()
st.subheader("Export to Excel")
st.caption(
    "Downloads every table in the app as a single .xlsx workbook — one "
    "sheet per table, raw numbers (no text formatting), ready to use "
    "directly in Excel."
)


def _style_workbook(workbook):
    """Bold header row, sensible column widths, frozen header/label column,
    and space-separated number formatting — Excel supports a literal space
    as a thousands separator natively, unlike Streamlit's charts/tables."""
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    label_font = Font(bold=True)

    for ws in workbook.worksheets:
        if ws.max_row < 1 or ws.max_column < 1:
            continue

        is_percent_sheet = ws.title == "Finance cost summary"

        # header row styling
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # column widths + number formatting + bold label column
        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            max_len = 8
            for row_idx in range(1, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                value = cell.value
                if value is not None:
                    max_len = max(max_len, len(str(value)))
                if row_idx > 1 and isinstance(value, (int, float)):
                    cell.number_format = "0.00" if is_percent_sheet else "#,##0"
                if row_idx > 1 and col_idx == 1 and not isinstance(value, (int, float)):
                    cell.font = label_font
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 45)

        ws.freeze_panes = "B2"
        ws.sheet_view.showGridLines = False


def _build_workbook() -> bytes:
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        results_df.to_excel(writer, sheet_name="Vessel TC-rate")
        finance_cost_summary_df.to_excel(writer, sheet_name="Finance cost summary", index=False)
        debt_schedule_df.to_excel(writer, sheet_name="Vessel debt schedule", index=False)
        if spot_market_enabled:
            pd.DataFrame(spot_service_items_current).to_excel(writer, sheet_name="Spot service mix", index=False)
            pd.DataFrame(st.session_state.spot_smolt_segments).to_excel(writer, sheet_name="Transport voyage build-up", index=False)
            pd.DataFrame(st.session_state.spot_harvest_segments).to_excel(writer, sheet_name="Other voyage build-up", index=False)
            pd.DataFrame(st.session_state.spot_treatment_segments).to_excel(writer, sheet_name="Treatment voyage build-up", index=False)
        if lease_enabled:
            lease_schedule_df.to_excel(writer, sheet_name="Lease schedule", index=False)
            if bank_financing_enabled:
                bank_schedule_df.to_excel(writer, sheet_name="Bank financing schedule", index=False)
                spread_df.to_excel(writer, sheet_name="Spread payment schedule", index=False)
        combined_df.to_excel(writer, sheet_name="Combined TC-rate", index=False)
        pnl_df.to_excel(writer, sheet_name="P&L (monthly)", index=False)
        pnl_annual.to_excel(writer, sheet_name="P&L (annual)")
        cf_df.to_excel(writer, sheet_name="Cash flow (monthly)", index=False)
        cf_annual.to_excel(writer, sheet_name="Cash flow (annual)")
        bs_df.to_excel(writer, sheet_name="Balance sheet (monthly)", index=False)
        bs_annual.to_excel(writer, sheet_name="Balance sheet (annual)")
        equity_cf_df.to_excel(writer, sheet_name="Equity cash flow (IRR)", index=False)
        instant_value_df_ia.to_excel(writer, sheet_name="Instant equity value (IA)")
        ratio_df.to_excel(writer, sheet_name="Summary ratios")
        instant_value_df.to_excel(writer, sheet_name="Instant equity value")
        _style_workbook(writer.book)
    buffer.seek(0)
    return buffer.getvalue()


excel_bytes = _build_workbook()
st.download_button(
    label="Download Excel workbook (.xlsx)",
    data=excel_bytes,
    file_name="tc_rate_model.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
)

# ---------------------------------------------------------------------------
# All reruns are triggered here, at the very end of the script, after every
# tab's widgets have fully rendered in this pass — never mid-script. Two
# triggers share this one spot:
#   1. Forced exactly once after the very first execution — Tab 1's
#      "Sources & Uses" reads a value computed on Tab 4 (which runs later
#      in this script), so on a true first load Tab 1 renders before that
#      value exists; this makes sure even a passive viewer sees the
#      correct number immediately, without needing to touch anything.
#   2. Any button elsewhere that called _request_rerun() instead of
#      st.rerun() directly, specifically to avoid cutting a pass short
#      before later tabs' widgets have run (see _request_rerun()'s
#      docstring near the top of the script).
# ---------------------------------------------------------------------------

if not st.session_state.get("_initial_rerun_done"):
    st.session_state["_initial_rerun_done"] = True
    st.rerun()
elif st.session_state.get("_pending_rerun"):
    st.session_state["_pending_rerun"] = False
    st.rerun()
